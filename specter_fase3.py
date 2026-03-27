from __future__ import annotations

import os
import re
import json
import html
import hashlib
import csv
import io
import logging
import threading
import sqlite3
import time
import openpyxl
from urllib.parse import urlparse, parse_qs, unquote_plus, quote
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, Optional, List, Tuple

import requests
from flask import Flask, request, jsonify, abort, render_template_string, redirect, Response

# ----------------------------
# Load .env (dev)
# ----------------------------
try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv()
except Exception:
    pass

# ----------------------------
# Logging with secret masking
# ----------------------------
SENSITIVE_ENV_KEYS = {"ESCAVADOR_TOKEN", "WEBHOOK_AUTH_TOKEN"}


def _mask(s: str) -> str:
    if not s:
        return s
    if len(s) <= 12:
        return "*" * len(s)
    return s[:6] + "…" + s[-6:]


def redact_secrets(text: str) -> str:
    if not text:
        return text
    out = text
    for k in SENSITIVE_ENV_KEYS:
        v = os.getenv(k, "")
        if v:
            out = out.replace(v, _mask(v))
    out = re.sub(r"Bearer\s+[A-Za-z0-9\-\._]+", "Bearer ***", out)
    return out


class RedactingFormatter(logging.Formatter):
    def format(self, record: logging.LogRecord) -> str:
        msg = super().format(record)
        return redact_secrets(msg)


LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
handler = logging.StreamHandler()
handler.setFormatter(RedactingFormatter("%(asctime)s [%(levelname)s] %(message)s"))
logger = logging.getLogger("escavador_monitor")
logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
logger.handlers.clear()
logger.addHandler(handler)
logger.propagate = False

# ----------------------------
# Config
# ----------------------------
ESCAVADOR_BASE = os.getenv("ESCAVADOR_BASE", "https://api.escavador.com/api/v2").rstrip("/")
ESCAVADOR_TOKEN = os.getenv("ESCAVADOR_TOKEN", "").strip()
WEBHOOK_AUTH_TOKEN = os.getenv("WEBHOOK_AUTH_TOKEN", "").strip()

DB_PATH = os.getenv("DB_PATH", "./specter.db")
DB_PATH_ABS = os.path.abspath(DB_PATH)
POLL_INTERVAL_SECONDS = int(os.getenv("POLL_INTERVAL_SECONDS", "300"))
AUTO_DISCOVER_ENABLED = os.getenv("AUTO_DISCOVER_ENABLED", "0").strip().lower() in ("1", "true", "yes", "on")
DISCOVER_INTERVAL_SECONDS = int(os.getenv("DISCOVER_INTERVAL_SECONDS", "3600"))
DISCOVER_LIMIT_PER_DOC = int(os.getenv("DISCOVER_LIMIT_PER_DOC", "50"))
DISCOVER_ONLY_IF_NO_LINKS = os.getenv("DISCOVER_ONLY_IF_NO_LINKS", "1").strip().lower() in ("1", "true", "yes", "on")
DISCOVER_MAX_DOCS_PER_CYCLE = int(os.getenv("DISCOVER_MAX_DOCS_PER_CYCLE", "50"))
CAPA_CACHE_TTL_SECONDS = int(os.getenv("CAPA_CACHE_TTL_SECONDS", "86400"))  # 1 dia
HOST = os.getenv("HOST", "0.0.0.0")
PORT = int(os.getenv("PORT", "5000"))

logger.info("DB configured: %s | abs=%s | cwd=%s", DB_PATH, DB_PATH_ABS, os.getcwd())

# ----------------------------
# Auto-discover runtime status (last cycle, manual trigger, etc.)
# ----------------------------
DISCOVER_STATE_LOCK = threading.Lock()
DISCOVER_STATE: Dict[str, Any] = {
    "running": False,
    "last_trigger": None,          # "auto" | "manual"
    "last_started": None,          # ISO UTC
    "last_finished": None,         # ISO UTC
    "last_totals": None,           # dict totals
    "last_error": None,            # str
}

def _set_discover_state(**kwargs: Any) -> None:
    with DISCOVER_STATE_LOCK:
        for k, v in kwargs.items():
            DISCOVER_STATE[k] = v

def _get_discover_state() -> Dict[str, Any]:
    with DISCOVER_STATE_LOCK:
        return dict(DISCOVER_STATE)

# ----------------------------
# Poll runtime status (last cycle, last totals, last error)
# ----------------------------
POLL_STATE_LOCK = threading.Lock()
POLL_STATE: Dict[str, Any] = {
    "running": False,
    "last_started": None,    # ISO UTC
    "last_finished": None,   # ISO UTC
    "last_totals": None,     # dict totals
    "last_error": None,      # str
}

def _set_poll_state(**kwargs: Any) -> None:
    with POLL_STATE_LOCK:
        for k, v in kwargs.items():
            POLL_STATE[k] = v

def _get_poll_state() -> Dict[str, Any]:
    with POLL_STATE_LOCK:
        return dict(POLL_STATE)

# ----------------------------
# Last Escavador API error (observability)
# ----------------------------
API_ERROR_LOCK = threading.Lock()
LAST_API_ERROR: Dict[str, Any] = {
    "at": None,        # ISO UTC
    "method": None,
    "path": None,
    "status": None,
    "message": None,
}

def _set_last_api_error(method: str, path: str, status: Optional[int], message: str) -> None:
    with API_ERROR_LOCK:
        LAST_API_ERROR["at"] = utcnow_iso()
        LAST_API_ERROR["method"] = method
        LAST_API_ERROR["path"] = path
        LAST_API_ERROR["status"] = status
        LAST_API_ERROR["message"] = (message or "")[:800]

def _get_last_api_error() -> Dict[str, Any]:
    with API_ERROR_LOCK:
        return dict(LAST_API_ERROR)

def run_discover_cycle(client: 'EscavadorClient', trigger: str = "manual") -> Dict[str, Any]:
    """Executa 1 ciclo de auto-discover e atualiza DISCOVER_STATE."""
    if client is None:
        raise RuntimeError("Client não inicializado.")
    with DISCOVER_STATE_LOCK:
        if DISCOVER_STATE.get("running"):
            return {"ok": False, "error": "ALREADY_RUNNING", "state": dict(DISCOVER_STATE)}
        DISCOVER_STATE["running"] = True
        DISCOVER_STATE["last_trigger"] = trigger
        DISCOVER_STATE["last_started"] = utcnow_iso()
        DISCOVER_STATE["last_error"] = None
        DISCOVER_STATE["last_totals"] = None
        DISCOVER_STATE["last_finished"] = None

    totals = {"docs": 0, "ran_docs": 0, "discovered": 0, "inserted_processos": 0, "linked": 0, "skipped": 0, "errors": 0}
    try:
        conn = db_connect()
        try:
            docs = _get_watchlist_docs(conn)[: max(DISCOVER_MAX_DOCS_PER_CYCLE, 1)]
        finally:
            conn.close()

        for doc in docs:
            if stop_flag.is_set():
                break
            totals["docs"] += 1
            try:
                conn = db_connect()
                try:
                    if DISCOVER_ONLY_IF_NO_LINKS and _doc_has_links(conn, doc):
                        totals["skipped"] += 1
                        continue
                finally:
                    conn.close()

                st = _discover_link_for_doc(client, doc, limit=max(DISCOVER_LIMIT_PER_DOC, 1))
                totals["ran_docs"] += 1
                totals["discovered"] += st["discovered"]
                totals["inserted_processos"] += st["inserted_processos"]
                totals["linked"] += st["linked"]
            except Exception:
                totals["errors"] += 1
                logger.exception("Discover cycle failed for doc=%s", doc)

        logger.info(
            "Discover cycle (%s): ran_docs=%s total_docs=%s skipped=%s discovered=%s inserted_processos=%s linked=%s errors=%s",
            trigger,
            totals["ran_docs"],
            totals["docs"],
            totals["skipped"],
            totals["discovered"],
            totals["inserted_processos"],
            totals["linked"],
            totals["errors"],
        )
        _set_discover_state(last_totals=totals, last_finished=utcnow_iso())
        return {"ok": True, "totals": totals, "state": _get_discover_state()}
    except Exception as e:
        logger.exception("Discover cycle failed")
        _set_discover_state(last_error=str(e), last_finished=utcnow_iso())
        return {"ok": False, "error": "CYCLE_FAILED", "message": str(e), "state": _get_discover_state()}
    finally:
        _set_discover_state(running=False)

logger.info(
    "Discover configured: enabled=%s interval=%ss limit_per_doc=%s only_if_no_links=%s max_docs_per_cycle=%s",
    AUTO_DISCOVER_ENABLED,
    DISCOVER_INTERVAL_SECONDS,
    DISCOVER_LIMIT_PER_DOC,
    DISCOVER_ONLY_IF_NO_LINKS,
    DISCOVER_MAX_DOCS_PER_CYCLE,
)

# ----------------------------
# Helpers
# ----------------------------
def utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def stable_hash(obj: Any) -> str:
    s = json.dumps(obj, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(s.encode("utf-8")).hexdigest()


def require_token_configured():
    if not ESCAVADOR_TOKEN:
        abort(500, description="ESCAVADOR_TOKEN não configurado no ambiente/.env.")


def normalize_doc(doc: str) -> str:
    digits = re.sub(r"\D", "", doc or "")
    if len(digits) in (11, 14):
        return doc.strip()
    raise ValueError("Documento inválido. Informe CPF (11 dígitos) ou CNPJ (14 dígitos).")


def doc_type(doc: str) -> str:
    digits = re.sub(r"\D", "", doc)
    return "CNPJ" if len(digits) == 14 else "CPF"


def normalize_cnj(value: str) -> str:
    cnj = (value or "").strip()
    digits = re.sub(r"\D", "", cnj)
    if len(digits) != 20:
        raise ValueError("CNJ inválido. Informe no padrão 0000000-00.0000.0.00.0000.")
    if not re.match(r"^\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}$", cnj):
        cnj = f"{digits[:7]}-{digits[7:9]}.{digits[9:13]}.{digits[13]}.{digits[14:16]}.{digits[16:20]}"
    return cnj


CNJ_REGEX = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}")
DOC_REGEX = re.compile(r"\b(\d{3}\.?\d{3}\.?\d{3}-?\d{2}|\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2})\b")

# ----------------------------
# Escavador Client
# ----------------------------
class EscavadorAlreadyMonitored(Exception):
    """Raised when Escavador reports the process is already being monitored."""


class EscavadorUnauthorized(Exception):
    """Raised when the token is not authorized for the requested endpoint."""


class EscavadorUpdateAlreadyRunning(Exception):
    """Raised when Escavador reports an update is already running for the process."""

    def __init__(self, payload: Optional[Dict[str, Any]] = None, message: str = "Atualização já em andamento."):
        super().__init__(message)
        self.payload = payload or {}


class EscavadorClient:
    def __init__(self, base: str, token: str):
        self.base = base.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Authorization": f"Bearer {token}",
                "X-Requested-With": "XMLHttpRequest",
                "Content-Type": "application/json",
                "Accept": "application/json",
                "User-Agent": "Specter/4.3",
            }
        )

    def _url(self, path: str) -> str:
        return f"{self.base}{path}"

    def _request_json(
        self,
        method: str,
        path: str,
        *,
        params: Optional[Dict[str, Any]] = None,
        payload: Optional[Dict[str, Any]] = None,
        timeout: int = 45,
        retries: int = 3,
        backoff_seconds: float = 1.0,
    ) -> Dict[str, Any]:
        """Requests JSON with small retry/backoff for transient network/API issues.

        We retry on:
          - network errors (RemoteDisconnected, timeouts, connection resets)
          - HTTP 429 and 5xx (Escavador or edge hiccups)
        """
        url = self._url(path)
        last_exc: Optional[Exception] = None
        for attempt in range(1, retries + 1):
            try:
                if method.upper() == "GET":
                    r = self.session.get(url, params=params, timeout=timeout)
                else:
                    r = self.session.request(method.upper(), url, params=params, json=payload, timeout=timeout)

                if r.status_code == 429 or 500 <= r.status_code <= 599:
                    # Transient server-side or rate-limit.
                    logger.warning("Escavador %s %s attempt=%s status=%s: %s", method, path, attempt, r.status_code, r.text[:400])
                    if attempt < retries:
                        time.sleep(backoff_seconds * attempt)
                        continue

                if r.status_code == 401:
                    msg = (r.text or "").strip()[:600]
                    logger.warning("Escavador %s %s unauthorized: %s", method, path, msg)
                    _set_last_api_error(method, path, r.status_code, r.text)
                    raise EscavadorUnauthorized(msg or f"401 Unauthorized em {path}")

                if r.status_code == 422 and method.upper() == "POST" and path == "/monitoramentos/processos":
                    body_txt = (r.text or "").lower()
                    if "já monitora este processo" in body_txt or "ja monitora este processo" in body_txt:
                        logger.info("Escavador %s %s: processo já monitorado, sem retry.", method, path)
                        _set_last_api_error(method, path, r.status_code, r.text)
                        raise EscavadorAlreadyMonitored((r.text or "").strip())

                if r.status_code == 422 and method.upper() == "POST" and path.endswith("/solicitar-atualizacao"):
                    body_json = {}
                    try:
                        body_json = r.json() if (r.text or "").strip() else {}
                    except Exception:
                        body_json = {}
                    body_txt = ((body_json.get("message") or r.text or "")).lower()
                    if "já está sendo atualizado" in body_txt or "ja está sendo atualizado" in body_txt or "ja esta sendo atualizado" in body_txt:
                        logger.info("Escavador %s %s: atualização já em andamento, sem retry.", method, path)
                        _set_last_api_error(method, path, r.status_code, r.text)
                        raise EscavadorUpdateAlreadyRunning(body_json, (body_json.get("message") or "Atualização já em andamento."))

                if r.status_code >= 400:
                    logger.error("Escavador %s %s failed %s: %s", method, path, r.status_code, r.text[:1200])
                    _set_last_api_error(method, path, r.status_code, r.text)
                    r.raise_for_status()

                # Some endpoints may reply with empty body (rare). Guard it.
                if not (r.text or "").strip():
                    return {}

                return r.json()

            except requests.exceptions.RequestException as e:
                last_exc = e
                logger.warning("Escavador %s %s attempt=%s network_error=%s", method, path, attempt, repr(e)[:400])
                _set_last_api_error(method, path, None, repr(e))
                if attempt < retries:
                    time.sleep(backoff_seconds * attempt)
                    continue
                raise

        # Should never reach here.
        if last_exc:
            raise last_exc
        return {}

    def post(self, path: str, payload: Dict[str, Any], params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        return self._request_json("POST", path, params=params, payload=payload)

    def get(self, path: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        return self._request_json("GET", path, params=params)

    def criar_monitor_novos_processos(self, termo: str) -> Dict[str, Any]:
        return self.post("/monitoramentos/novos-processos", {"termo": termo})

    def criar_monitor_processo(self, numero_cnj: str) -> Dict[str, Any]:
        return self.post("/monitoramentos/processos", {"numero": numero_cnj})

    def listar_movimentacoes(self, numero_cnj: str, limit: int = 100) -> Dict[str, Any]:
        service_key = "v2_movimentacoes_processo"
        endpoint = f"/processos/numero_cnj/{numero_cnj}/movimentacoes"
        status = None
        data: Dict[str, Any] = {}
        try:
            data = self.get(endpoint, {"limit": limit})
            status = 200
            return data
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            raise
        finally:
            items = 0
            try:
                items = len(extract_list(data))
            except Exception:
                items = 0
            cost = _estimate_cost_brl(service_key, items_upto=items)
            record_api_usage(doc=None, cnj=numero_cnj, service_key=service_key, endpoint=endpoint, http_status=status, items_count=items, cost_brl=cost)

    def obter_capa_processo(self, numero_cnj: str) -> Dict[str, Any]:
        service_key = "v2_capa_processo"
        endpoint = f"/processos/numero_cnj/{numero_cnj}"
        status = None
        data: Dict[str, Any] = {}
        try:
            data = self.get(endpoint, params=None)
            status = 200
            return data
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            raise
        finally:
            # capa não é lista; itens_count=0
            cost = _estimate_cost_brl(service_key, items_upto=0)
            record_api_usage(doc=None, cnj=numero_cnj, service_key=service_key, endpoint=endpoint, http_status=status, items_count=0, cost_brl=cost)


    def solicitar_atualizacao_processo(
        self,
        numero_cnj: str,
        *,
        documentos_publicos: bool = False,
        autos: bool = False,
        enviar_callback: bool = False,
        utilizar_certificado: Optional[bool] = None,
        certificado_id: Optional[int] = None,
        usuario: Optional[str] = None,
        senha: Optional[str] = None,
        documentos_especificos: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Solicita atualização do processo no tribunal (assíncrono).

        Docs Escavador API v2:
          POST /processos/numero_cnj/{numero}/solicitar-atualizacao
          - documentos_publicos=1 OU autos=1 (não simultâneos)
        """
        service_key = "v2_atualizacao_processo"
        endpoint = f"/processos/numero_cnj/{numero_cnj}/solicitar-atualizacao"
        payload: Dict[str, Any] = {}
        if enviar_callback:
            payload["enviar_callback"] = 1
        if documentos_publicos:
            payload["documentos_publicos"] = 1
        if autos:
            payload["autos"] = 1
        if utilizar_certificado is not None:
            payload["utilizar_certificado"] = 1 if utilizar_certificado else 0
        if certificado_id is not None:
            payload["certificado_id"] = int(certificado_id)
        if usuario:
            payload["usuario"] = str(usuario)
        if senha:
            payload["senha"] = str(senha)
        if documentos_especificos:
            payload["documentos_especificos"] = str(documentos_especificos)

        status = None
        data: Dict[str, Any] = {}
        try:
            data = self.post(endpoint, payload if payload else {})
            status = 200
            return data
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            raise
        finally:
            cost = _estimate_cost_brl(service_key, items_upto=0)
            record_api_usage(doc=None, cnj=numero_cnj, service_key=service_key, endpoint=endpoint, http_status=status, items_count=0, cost_brl=cost)

    def status_atualizacao_processo(self, numero_cnj: str) -> Dict[str, Any]:
        """Consulta o status da última solicitação de atualização (se existir)."""
        service_key = "v2_status_atualizacao_processo"
        endpoint = f"/processos/numero_cnj/{numero_cnj}/status-atualizacao"
        status = None
        data: Dict[str, Any] = {}
        try:
            data = self.get(endpoint, params=None)
            status = 200
            return data
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            raise
        finally:
            cost = _estimate_cost_brl(service_key, items_upto=0)
            record_api_usage(doc=None, cnj=numero_cnj, service_key=service_key, http_status=status, items_count=0, cost_brl=cost)

    def listar_documentos_publicos(self, numero_cnj: str, limit: int = 100) -> Dict[str, Any]:
        """Lista documentos públicos do processo (requer atualização com documentos_publicos=1 para ficar 'cheio')."""
        service_key = "v2_documentos_publicos"
        endpoint = f"/processos/numero_cnj/{numero_cnj}/documentos-publicos"
        status = None
        data: Dict[str, Any] = {}
        try:
            data = self.get(endpoint, {"limit": limit})
            status = 200
            return data
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            raise
        finally:
            cost = _estimate_cost_brl(service_key, items_upto=limit)
            record_api_usage(doc=None, cnj=numero_cnj, service_key=service_key, http_status=status, items_count=limit, cost_brl=cost)

    def listar_autos(self, numero_cnj: str, limit: int = 50) -> Dict[str, Any]:
        """Lista autos (públicos + restritos). Requer atualização prévia com autos=1 e status SUCESSO."""
        service_key = "v2_autos"
        endpoint = f"/processos/numero_cnj/{numero_cnj}/autos"
        status = None
        data: Dict[str, Any] = {}
        try:
            data = self.get(endpoint, {"limit": limit})
            status = 200
            return data
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            raise
        finally:
            cost = _estimate_cost_brl(service_key, items_upto=limit)
            record_api_usage(doc=None, cnj=numero_cnj, service_key=service_key, http_status=status, items_count=limit, cost_brl=cost)

    def obter_documento_por_key(self, numero_cnj: str, key: str) -> Dict[str, Any]:
        """Obtém metadados/links de um documento (key) para download."""
        service_key = "v2_documento_key"
        endpoint = f"/processos/numero_cnj/{numero_cnj}/documentos/{key}"
        status = None
        data: Dict[str, Any] = {}
        try:
            data = self.get(endpoint, params=None)
            status = 200
            return data
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            raise
        finally:
            cost = _estimate_cost_brl(service_key, items_upto=0)
            record_api_usage(doc=None, cnj=numero_cnj, service_key=service_key, http_status=status, items_count=0, cost_brl=cost)


    def listar_processos_envolvido(self, cpf_cnpj: str, limit: int = 50, page: Optional[int] = None) -> Dict[str, Any]:
        service_key = "v2_processos_envolvido"
        params: Dict[str, Any] = {"cpf_cnpj": cpf_cnpj, "limit": limit}
        if page is not None:
            params["page"] = page
        endpoint = "/envolvido/processos"
        status = None
        data: Dict[str, Any] = {}
        try:
            data = self.get(endpoint, params=params)
            status = 200
            return data
        except requests.exceptions.HTTPError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            raise
        finally:
            items = 0
            try:
                items = len(extract_list(data))
            except Exception:
                items = 0
            # estimativa por lotes de 200 itens cumulativos
            p = int(page or 1)
            items_upto = (max(p-1, 0) * int(limit)) + items
            cost = _estimate_cost_brl(service_key, items_upto=items_upto)
            record_api_usage(doc=normalize_doc(cpf_cnpj), cnj=None, service_key=service_key, endpoint=endpoint, http_status=status, items_count=items, cost_brl=cost, notes=f"limit={limit} page={p}")

    def listar_callbacks(self, limit: int = 100, page: Optional[int] = None) -> Dict[str, Any]:
        params: Dict[str, Any] = {"limit": limit}
        if page is not None:
            params["page"] = page
        return self.get("/callbacks", params=params)


# ----------------------------
# SQLite
# ----------------------------
def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def db_init() -> None:
    conn = db_connect()
    cur = conn.cursor()

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS watchlist (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doc TEXT NOT NULL UNIQUE,
        tipo_doc TEXT NOT NULL,
        created_at TEXT NOT NULL
    )"""
    )

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS processos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cnj TEXT NOT NULL UNIQUE,
        created_at TEXT NOT NULL,
        last_sync_at TEXT,
        last_event_at TEXT
    )"""
    )

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS eventos_mov (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cnj TEXT NOT NULL,
        event_hash TEXT NOT NULL,
        data TEXT,
        tipo TEXT,
        tipo_inferido TEXT,
        texto TEXT,
        raw_json TEXT,
        created_at TEXT NOT NULL,
        UNIQUE(cnj, event_hash)
    )"""
    )

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS callback_inbox (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source TEXT NOT NULL,
        payload_hash TEXT NOT NULL UNIQUE,
        payload_json TEXT NOT NULL,
        received_at TEXT NOT NULL,
        processed_at TEXT,
        status TEXT NOT NULL DEFAULT 'PENDING',
        error TEXT
    )"""
    )

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS doc_process (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doc TEXT NOT NULL,
        cnj TEXT NOT NULL,
        created_at TEXT NOT NULL,
        UNIQUE(doc, cnj)
    )"""
    )

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS alert_state (
        doc TEXT PRIMARY KEY,
        last_event_id INTEGER NOT NULL DEFAULT 0,
        updated_at TEXT NOT NULL
    )"""
    )

    
    # Cache de capa (JSON bruto) por CNJ
    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS capa_cache (
        cnj TEXT PRIMARY KEY,
        payload TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )"""
    )
    
    # Uso da API (para métricas de custo)
    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS api_usage (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ts TEXT NOT NULL,
        doc TEXT,
        cnj TEXT,
        service_key TEXT NOT NULL,
        endpoint TEXT NOT NULL,
        http_status INTEGER,
        items_count INTEGER NOT NULL DEFAULT 0,
        cost_brl REAL NOT NULL DEFAULT 0.0,
        notes TEXT
    )"""
    )
    
    # Cache/controle de atualização + documentos (documentos públicos e autos)
    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS processo_updates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cnj TEXT NOT NULL,
        tipo TEXT NOT NULL, -- 'documentos_publicos' | 'autos'
        escavador_update_id INTEGER,
        status TEXT NOT NULL DEFAULT 'PENDENTE', -- PENDENTE|SUCESSO|ERRO|NAO_ENCONTRADO
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        last_error TEXT
    )"""
    )

    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS documentos_cache (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cnj TEXT NOT NULL,
        tipo TEXT NOT NULL, -- 'publicos' | 'autos'
        doc_key TEXT NOT NULL,
        titulo TEXT,
        data TEXT,
        mime TEXT,
        meta_json TEXT,
        download_url TEXT,
        updated_at TEXT NOT NULL,
        UNIQUE(cnj, tipo, doc_key)
    )"""
    )

    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_usage_ts ON api_usage (ts)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_usage_doc_ts ON api_usage (doc, ts)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_usage_cnj_ts ON api_usage (cnj, ts)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_usage_service_ts ON api_usage (service_key, ts)")
    # Extrato real (importado) da API (custos cobrados pelo Escavador)
    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS api_usage_real (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ts TEXT NOT NULL,
        doc TEXT,
        cnj TEXT,
        method TEXT,
        endpoint TEXT NOT NULL,
        cost_brl REAL NOT NULL,
        raw_line TEXT
    )"""
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_usage_real_ts ON api_usage_real (ts)")
    # Dedup (fingerprint)
    try:
        cur.execute("PRAGMA table_info(api_usage_real)")
        cols = [r[1] for r in cur.fetchall()]
        if "fingerprint" not in cols:
            cur.execute("ALTER TABLE api_usage_real ADD COLUMN fingerprint TEXT")
    except Exception:
        pass
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_api_usage_real_fp ON api_usage_real (fingerprint)")
    # Import batch metadata
    try:
        cur.execute("PRAGMA table_info(api_usage_real)")
        cols = [r[1] for r in cur.fetchall()]
        if "imported_at" not in cols:
            cur.execute("ALTER TABLE api_usage_real ADD COLUMN imported_at TEXT")
        if "source" not in cols:
            cur.execute("ALTER TABLE api_usage_real ADD COLUMN source TEXT")
    except Exception:
        pass
    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_usage_real_imported_at ON api_usage_real (imported_at)")
    

    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_usage_real_doc_ts ON api_usage_real (doc, ts)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_api_usage_real_cnj_ts ON api_usage_real (cnj, ts)")


    # Tabela opcional de preços (override dos defaults em código)
    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS pricing (
        service_key TEXT PRIMARY KEY,
        model TEXT NOT NULL,              -- FIXED / ITEMS200
        base_cost_brl REAL NOT NULL,
        step_cost_brl REAL NOT NULL DEFAULT 0.0,
        step_items INTEGER NOT NULL DEFAULT 0,
        updated_at TEXT NOT NULL
    )"""
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_eventos_cnj_id ON eventos_mov (cnj, id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_doc_process_doc ON doc_process (doc)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_docs_cache_cnj_tipo ON documentos_cache (cnj, tipo)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_updates_cnj_status ON processo_updates (cnj, status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_doc_process_cnj ON doc_process (cnj)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_inbox_status ON callback_inbox (status, id)")

    conn.commit()
    conn.close()



# ----------------------------
# Custos (precificação + uso)
# ----------------------------
PRICING_DEFAULTS = {
    # Escavador API V2 (baseado na tabela enviada)
    "v2_capa_processo": {"model": "FIXED", "base": 0.04},
    "v2_movimentacoes_processo": {"model": "FIXED", "base": 0.04},
    # Processos do envolvido: R$ 2,90 até 200 itens + R$ 0,05 a cada 200
    "v2_processos_envolvido": {"model": "ITEMS200", "base": 2.90, "step": 0.05, "step_items": 200},
    # Outros (se você quiser ativar depois)
    "v2_resumo_envolvido": {"model": "FIXED", "base": 0.35},
    "v2_resumo_processo_ia": {"model": "FIXED", "base": 0.04},
}

def _pricing_for(service_key: str) -> dict:
    """Retorna precificação. Primeiro tenta tabela pricing, senão defaults."""
    try:
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("SELECT model, base_cost_brl, step_cost_brl, step_items FROM pricing WHERE service_key=?", (service_key,))
        row = cur.fetchone()
        conn.close()
        if row:
            return {
                "model": (row["model"] if isinstance(row, sqlite3.Row) else row[0]),
                "base": float(row["base_cost_brl"] if isinstance(row, sqlite3.Row) else row[1]),
                "step": float(row["step_cost_brl"] if isinstance(row, sqlite3.Row) else row[2]),
                "step_items": int(row["step_items"] if isinstance(row, sqlite3.Row) else row[3]),
            }
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
    return PRICING_DEFAULTS.get(service_key, {"model": "FIXED", "base": 0.0})

def _estimate_cost_brl(service_key: str, *, items_upto: int = 0) -> float:
    p = _pricing_for(service_key)
    model = (p.get("model") or "FIXED").upper()
    base = float(p.get("base") or 0.0)
    if model == "ITEMS200":
        step = float(p.get("step") or 0.0)
        step_items = int(p.get("step_items") or 200) or 200
        if items_upto <= 0:
            return base
        batches = (items_upto + step_items - 1) // step_items
        return base + step * max(0, batches - 1)
    return base

def record_api_usage(*, doc: Optional[str], cnj: Optional[str], service_key: str, endpoint: str, http_status: Optional[int], items_count: int, cost_brl: float, notes: str = "") -> None:
    try:
        conn = db_connect()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO api_usage (ts, doc, cnj, service_key, endpoint, http_status, items_count, cost_brl, notes) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (utcnow_iso(), doc, cnj, service_key, endpoint, http_status, int(items_count or 0), float(cost_brl or 0.0), notes[:500]),
        )
        conn.commit()
        conn.close()
    except Exception:
        logger.exception("Falha ao registrar api_usage")
        try:
            conn.close()
        except Exception:
            pass

# ----------------------------
# Custos reais (importados do extrato do Escavador)
# ----------------------------
EXTRATO_LINE_RE = re.compile(r"^(GET|POST|PUT|DELETE)\s+(\S+)\s+R\$\s*([+-]?[\d\.,]+)\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}:\d{2})\s*$")

def _parse_brl_value(s: str) -> float:
    s = s.strip().replace("R$", "").strip()
    # "0,04" ou "-0,04" (o extrato vem negativo)
    s = s.replace(".", "").replace(",", ".")
    return float(s)

def _extract_cnj_from_endpoint(ep: str) -> Optional[str]:
    m = re.search(r"/processos/numero_cnj/([^/?]+)", ep)
    return m.group(1) if m else None

def _extract_doc_from_endpoint(ep: str) -> Optional[str]:
    """Extrai CPF/CNPJ (doc) de um endpoint/querystring, de forma tolerante.
    Aceita valor URL-encoded e com pontuação. Retorna apenas se tiver 11 (CPF) ou 14 (CNPJ) dígitos.
    """
    m = re.search(r"(?:\?|&)cpf_cnpj=([^&]+)", ep)
    if not m:
        return None
    raw = m.group(1)
    try:
        raw = unquote_plus(str(raw))
    except Exception:
        raw = str(raw)
    digits = re.sub(r"\D", "", raw)
    if len(digits) in (11, 14):
        return digits
    return None

def _fingerprint_usage(ts_iso: str, method: str, endpoint: str, cost_brl: float) -> str:
    s = f"{ts_iso}|{method}|{endpoint}|{float(cost_brl):.6f}"
    return hashlib.sha1(s.encode("utf-8", errors="ignore")).hexdigest()

def record_api_usage_real(*, ts_iso: str, doc: Optional[str], cnj: Optional[str], method: str, endpoint: str,
                         cost_brl: float, raw_line: str, source: str = "xlsx", imported_at: Optional[str] = None) -> None:
    """Registra consumo REAL importado do extrato (com dedupe via fingerprint)."""
    try:
        fp = _fingerprint_usage(ts_iso, method, endpoint, float(cost_brl))
        imported_at = imported_at or utcnow_iso()
        conn = db_connect()
        cur = conn.cursor()
        cur.execute(
            "INSERT OR IGNORE INTO api_usage_real (ts, doc, cnj, method, endpoint, cost_brl, raw_line, fingerprint, imported_at, source) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (ts_iso, doc, cnj, method, endpoint, float(cost_brl), (raw_line or "")[:800], fp, imported_at, source),
        )
        conn.commit()
        conn.close()
    except Exception:
        logger.exception("Falha ao registrar api_usage_real")
        try:
            conn.close()
        except Exception:
            pass


db_init()

# ----------------------------
# Tipo inferido (fallback)
# ----------------------------
KEYWORDS = [
    ("CITACAO", [r"\bcita", r"\bcitação", r"\bcitacao"]),
    ("INTIMACAO", [r"\bintim", r"\bintimação", r"\bintimacao"]),
    ("SENTENCA", [r"\bsenten", r"\bsentença", r"\bsentenca"]),
    ("DECISAO", [r"\bdecis", r"\bdecisão", r"\bdecisao"]),
    ("DESPACHO", [r"\bdespach"]),
    ("AUDIENCIA", [r"\baudi", r"\baudiência", r"\baudiencia"]),
    ("JUNTADA", [r"\bjuntad"]),
    ("DISTRIBUICAO", [r"\bdistribu", r"\bdistribuição", r"\bdistribuicao"]),
    ("PENHORA", [r"\bpenhor"]),
    ("TRANSITO", [r"\btr[âa]nsit", r"\btransit"]),
]


def infer_tipo(texto: str) -> Optional[str]:
    if not texto:
        return None
    for label, patterns in KEYWORDS:
        for p in patterns:
            if re.search(p, texto, flags=re.IGNORECASE):
                return label
    return None


def mov_to_hash(mov: Dict[str, Any]) -> str:
    for key in ("id", "codigo", "uuid", "hash"):
        if mov.get(key):
            return f"id:{mov[key]}"
    core = {
        "data": mov.get("data") or mov.get("data_hora") or mov.get("dataHora") or mov.get("dataHoraCadastro"),
        "texto": mov.get("texto") or mov.get("descricao") or mov.get("conteudo"),
        "tipo": mov.get("tipo") or mov.get("tipo_movimentacao") or mov.get("tipoMovimentacao"),
    }
    return stable_hash(core)


def extract_list(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, dict):
        for k in ("items", "data", "movimentacoes", "results", "callbacks", "processos"):
            if isinstance(data.get(k), list):
                return data[k]
    if isinstance(data, list):
        return data
    return []


# ----------------------------
# Callback parsing
# ----------------------------
def parse_cnj_candidates(payload: Dict[str, Any]) -> List[str]:
    cnjs: List[str] = []

    def scan(o: Any):
        if isinstance(o, dict):
            for _, v in o.items():
                if isinstance(v, (dict, list)):
                    scan(v)
                elif isinstance(v, str):
                    m = CNJ_REGEX.search(v)
                    if m:
                        cnjs.append(m.group(0))
        elif isinstance(o, list):
            for x in o:
                scan(x)

    scan(payload)
    out, seen = [], set()
    for c in cnjs:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out


def parse_doc_candidates(payload: Dict[str, Any]) -> List[str]:
    docs: List[str] = []

    def add_doc(s: str):
        m = DOC_REGEX.search(s or "")
        if not m:
            return
        raw = m.group(1)
        digits = re.sub(r"\D", "", raw)
        if len(digits) in (11, 14):
            docs.append(raw)

    def scan(o: Any):
        if isinstance(o, dict):
            for _, v in o.items():
                if isinstance(v, (dict, list)):
                    scan(v)
                elif isinstance(v, str):
                    add_doc(v)
        elif isinstance(o, list):
            for x in o:
                scan(x)

    scan(payload)
    out, seen = [], set()
    for d in docs:
        key = re.sub(r"\D", "", d)
        if key not in seen:
            seen.add(key)
            out.append(d)
    return out


# ----------------------------
# Core: sync + persistence
# ----------------------------
@dataclass
class ProcessResult:
    cnj: str
    new_events: int


def ensure_process_registered(conn: sqlite3.Connection, cnj: str) -> None:
    cur = conn.cursor()
    cur.execute("INSERT OR IGNORE INTO processos (cnj, created_at) VALUES (?, ?)", (cnj, utcnow_iso()))
    conn.commit()


def link_doc_process(conn: sqlite3.Connection, doc: str, cnj: str) -> bool:
    """Link doc->cnj. Returns True if a new link was created."""
    cur = conn.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO doc_process (doc, cnj, created_at) VALUES (?, ?, ?)",
        (doc, cnj, utcnow_iso()),
    )
    conn.commit()
    return cur.rowcount == 1


def unlink_doc_process(conn: sqlite3.Connection, doc: str, cnj: str) -> int:
    cur = conn.cursor()
    cur.execute("DELETE FROM doc_process WHERE doc=? AND cnj=?", (doc, cnj))
    conn.commit()
    return cur.rowcount

def upsert_processo(conn: sqlite3.Connection, cnj: str) -> bool:
    cur = conn.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO processos (cnj, created_at) VALUES (?, ?)",
        (cnj, utcnow_iso()),
    )
    conn.commit()
    return cur.rowcount > 0



def save_mov_events(conn: sqlite3.Connection, cnj: str, movs: List[Dict[str, Any]]) -> int:
    cur = conn.cursor()
    new_count = 0
    for mov in movs:
        event_hash = mov_to_hash(mov)
        texto = mov.get("texto") or mov.get("descricao") or mov.get("conteudo") or ""
        tipo = mov.get("tipo") or mov.get("tipo_movimentacao") or mov.get("tipoMovimentacao")
        tipo_inf = infer_tipo(texto)
        data = mov.get("data") or mov.get("data_hora") or mov.get("dataHora") or mov.get("dataHoraCadastro")
        raw = json.dumps(mov, ensure_ascii=False)

        try:
            cur.execute(
                """INSERT INTO eventos_mov
                   (cnj, event_hash, data, tipo, tipo_inferido, texto, raw_json, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (cnj, event_hash, data, tipo, tipo_inf, texto, raw, utcnow_iso()),
            )
            new_count += 1
        except sqlite3.IntegrityError:
            continue

    if new_count > 0:
        cur.execute("UPDATE processos SET last_event_at=?, last_sync_at=? WHERE cnj=?", (utcnow_iso(), utcnow_iso(), cnj))
    else:
        cur.execute("UPDATE processos SET last_sync_at=? WHERE cnj=?", (utcnow_iso(), cnj))
    conn.commit()
    return new_count


def sync_process_movements(client: EscavadorClient, cnj: str, limit: int = 300) -> ProcessResult:
    data = client.listar_movimentacoes(cnj, limit=limit)
    movs = extract_list(data)

    conn = db_connect()
    ensure_process_registered(conn, cnj)
    new_events = save_mov_events(conn, cnj, movs)
    conn.close()
    return ProcessResult(cnj=cnj, new_events=new_events)


def ingest_callback(source: str, payload: Dict[str, Any]) -> Tuple[bool, str]:
    h = stable_hash(payload)
    conn = db_connect()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO callback_inbox (source, payload_hash, payload_json, received_at) VALUES (?, ?, ?, ?)",
            (source, h, json.dumps(payload, ensure_ascii=False), utcnow_iso()),
        )
        conn.commit()
        return True, h
    except sqlite3.IntegrityError:
        return False, h
    finally:
        conn.close()


def process_inbox_once(client: EscavadorClient, max_items: int = 25) -> Dict[str, Any]:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, payload_json FROM callback_inbox WHERE status='PENDING' ORDER BY id ASC LIMIT ?",
        (max_items,),
    )
    rows = cur.fetchall()
    conn.close()

    processed, errors = 0, 0
    cnj_synced: Dict[str, int] = {}
    linked: List[Dict[str, str]] = []

    for row in rows:
        inbox_id = row["id"]
        payload = json.loads(row["payload_json"])
        try:
            cnjs = parse_cnj_candidates(payload)
            docs = parse_doc_candidates(payload)

            for cnj in cnjs:
                try:
                    client.criar_monitor_processo(cnj)
                except EscavadorAlreadyMonitored:
                    logger.info("Monitoramento já existia para CNJ %s", cnj)
                except requests.HTTPError as e:
                    logger.warning("criar_monitor_processo(%s) warning: %s", cnj, str(e))

                res = sync_process_movements(client, cnj, limit=400)
                cnj_synced[cnj] = cnj_synced.get(cnj, 0) + res.new_events

                if docs:
                    conn3 = db_connect()
                    for d in docs:
                        try:
                            dnorm = normalize_doc(d)
                        except Exception:
                            continue
                        link_doc_process(conn3, dnorm, cnj)
                        linked.append({"doc": dnorm, "cnj": cnj})
                    conn3.close()

            conn2 = db_connect()
            cur2 = conn2.cursor()
            cur2.execute("UPDATE callback_inbox SET status='PROCESSED', processed_at=? WHERE id=?", (utcnow_iso(), inbox_id))
            conn2.commit()
            conn2.close()
            processed += 1

        except Exception as ex:
            logger.exception("Failed processing inbox id=%s", inbox_id)
            conn2 = db_connect()
            cur2 = conn2.cursor()
            cur2.execute(
                "UPDATE callback_inbox SET status='ERROR', processed_at=?, error=? WHERE id=?",
                (utcnow_iso(), str(ex)[:1000], inbox_id),
            )
            conn2.commit()
            conn2.close()
            errors += 1

    return {"processed": processed, "errors": errors, "cnj_new_events": cnj_synced, "linked": linked}


# ----------------------------
# Polling thread
# ----------------------------
stop_flag = threading.Event()


def poll_callbacks_loop(client: EscavadorClient):
    logger.info("Polling loop started (interval=%ss)", POLL_INTERVAL_SECONDS)
    while not stop_flag.is_set():
        _set_poll_state(running=True, last_started=utcnow_iso(), last_error=None)
        try:
            data = client.listar_callbacks(limit=100, page=1)
            callbacks = extract_list(data)

            inserted = 0
            for cb in callbacks:
                ins, _ = ingest_callback("poll", cb)
                if ins:
                    inserted += 1

            pr = process_inbox_once(client, max_items=50)
            logger.info(
                "Poll: inserted=%s processed=%s errors=%s cnj_events=%s linked=%s",
                inserted,
                pr["processed"],
                pr["errors"],
                pr["cnj_new_events"],
                len(pr.get("linked", [])),
            )
            _set_poll_state(last_totals={"inserted": inserted, "processed": pr["processed"], "errors": pr["errors"], "linked": len(pr.get("linked", []))}, last_finished=utcnow_iso())
        except EscavadorUnauthorized as e:
            logger.warning("Polling pausado por token não autorizado em /callbacks: %s", str(e)[:300])
            _set_poll_state(last_error=str(e), last_finished=utcnow_iso())
        except Exception as e:
            logger.exception("Polling failed")
            _set_poll_state(last_error=str(e), last_finished=utcnow_iso())
        finally:
            _set_poll_state(running=False)

        # também verifica solicitações pendentes de atualização (docs/autos)
        try:
            upd = process_updates_once(client, max_items=10)
            if upd.get("checked") or upd.get("errors"):
                logger.info("Updates: checked=%s completed=%s errors=%s", upd.get("checked"), upd.get("completed"), upd.get("errors"))
        except Exception:
            logger.exception("Updates polling failed")

        stop_flag.wait(POLL_INTERVAL_SECONDS)



# ----------------------------
# Auto-discover: watchlist -> processes linking
# ----------------------------
def _get_watchlist_docs(conn) -> list[str]:
    cur = conn.cursor()
    cur.execute("SELECT doc FROM watchlist ORDER BY id ASC")
    rows = cur.fetchall()
    docs: list[str] = []
    for r in rows:
        # sqlite3.Row supports mapping-style access, but has no .get()
        v = r["doc"]
        if v:
            docs.append(v)
    return docs


def _doc_has_links(conn, doc: str) -> bool:
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM doc_process WHERE doc=? LIMIT 1", (doc,))
    return cur.fetchone() is not None


def _discover_link_for_doc(client: EscavadorClient, doc: str, limit: int) -> dict:
    """Discover CNJs for a doc (CPF/CNPJ) and link them. Returns stats."""
    docn = normalize_doc(doc)
    resp = client.listar_processos_envolvido(docn, limit=limit, page=1)  # type: ignore
    processos = extract_list(resp)
    if not processos and isinstance(resp, dict) and isinstance(resp.get("processos"), list):
        processos = resp["processos"]

    discovered = []
    inserted_processos = 0
    linked = 0

    conn = db_connect()
    try:
        for p in processos:
            if not isinstance(p, dict):
                continue
            cnj = (p.get("numero_cnj") or p.get("numero") or p.get("cnj") or "").strip()
            if not cnj or not CNJ_REGEX.fullmatch(cnj):
                continue
            discovered.append(cnj)
            if upsert_processo(conn, cnj):
                inserted_processos += 1
            if link_doc_process(conn, docn, cnj):
                linked += 1
    finally:
        conn.close()

    return {
        "doc": docn,
        "discovered": len(discovered),
        "inserted_processos": inserted_processos,
        "linked": linked,
    }


def auto_discover_loop(client: EscavadorClient):
    if DISCOVER_INTERVAL_SECONDS <= 0:
        logger.info("Auto-discover desabilitado (DISCOVER_INTERVAL_SECONDS<=0).")
        return
    logger.info(
        "Auto-discover loop started (enabled=%s interval=%ss)",
        AUTO_DISCOVER_ENABLED,
        DISCOVER_INTERVAL_SECONDS,
    )

    while not stop_flag.is_set():
        if not AUTO_DISCOVER_ENABLED:
            stop_flag.wait(min(DISCOVER_INTERVAL_SECONDS, 10))
            continue

        if not ESCAVADOR_TOKEN:
            logger.info("Auto-discover pausado: ESCAVADOR_TOKEN ausente.")
            stop_flag.wait(DISCOVER_INTERVAL_SECONDS)
            continue

        try:
            run_discover_cycle(client, trigger="auto")  # type: ignore
        except Exception:
            logger.exception("Auto-discover cycle failed")

        stop_flag.wait(DISCOVER_INTERVAL_SECONDS)

# ----------------------------
# Flask app
# ----------------------------

def esc(v: object) -> str:
    """HTML-escape helper for UI rendering."""
    return html.escape("" if v is None else str(v), quote=True)


app = Flask(__name__)
client = EscavadorClient(ESCAVADOR_BASE, ESCAVADOR_TOKEN) if ESCAVADOR_TOKEN else None


@app.get("/")
def home():
    return redirect("/ui")


@app.get("/favicon.ico")
def favicon():
    # Silence browser favicon 404s
    return ("", 204)

@app.get("/.well-known/appspecific/com.chrome.devtools.json")
def chrome_devtools_well_known():
    # Silence Chrome DevTools well-known probe
    return ("", 204)
# ============================================================
# API JSON: health
# ============================================================
@app.get("/health")
def health():
    return jsonify(
        {
            "ok": True,
            "time": utcnow_iso(),
            "escavador_base": ESCAVADOR_BASE,
            "token_configured": bool(ESCAVADOR_TOKEN),
            "webhook_auth_configured": bool(WEBHOOK_AUTH_TOKEN),
            "db_path": DB_PATH,
            "db_path_abs": DB_PATH_ABS,
            "cwd": os.getcwd(),
            "poll_interval_seconds": POLL_INTERVAL_SECONDS,
            "auto_discover_enabled": bool(AUTO_DISCOVER_ENABLED),
            "discover_interval_seconds": DISCOVER_INTERVAL_SECONDS,
            "discover_limit_per_doc": DISCOVER_LIMIT_PER_DOC,
            "discover_only_if_no_links": bool(DISCOVER_ONLY_IF_NO_LINKS),
            "discover_max_docs_per_cycle": DISCOVER_MAX_DOCS_PER_CYCLE,
            "discover_state": _get_discover_state(),
        }
    )


# ============================================================
# API JSON: Watchlist CRUD
# ============================================================
@app.get("/watchlist")
def list_watchlist():
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT id, doc, tipo_doc, created_at FROM watchlist ORDER BY id DESC LIMIT 500")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "count": len(rows), "items": rows})


@app.post("/watchlist")
def create_watchlist():
    require_token_configured()
    payload = request.get_json(force=True, silent=False) or {}
    doc_in = (payload.get("doc") or "").strip()
    if not doc_in:
        abort(400, description="Campo 'doc' é obrigatório.")
    try:
        doc = normalize_doc(doc_in)
    except ValueError as e:
        abort(400, description=str(e))
    tipo = doc_type(doc)

    api_resp: Dict[str, Any]
    try:
        api_resp = client.criar_monitor_novos_processos(doc)  # type: ignore
    except requests.HTTPError as e:
        api_resp = {"warning": "monitoramento pode já existir", "detail": str(e)}

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("INSERT OR IGNORE INTO watchlist (doc, tipo_doc, created_at) VALUES (?, ?, ?)", (doc, tipo, utcnow_iso()))
    conn.commit()
    cur.execute("SELECT id FROM watchlist WHERE doc=?", (doc,))
    row = cur.fetchone()
    conn.close()

    return jsonify({"ok": True, "id": row["id"] if row else None, "doc": doc, "tipo_doc": tipo, "escavador": api_resp})


@app.delete("/watchlist/<int:watch_id>")
def delete_watchlist_by_id(watch_id: int):
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("DELETE FROM watchlist WHERE id=?", (watch_id,))
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return jsonify({"ok": True, "deleted": deleted})


# ============================================================
# API JSON: doc_process CRUD
# ============================================================
@app.get("/processos/<path:cnj>/docs")
def list_docs_for_process(cnj: str):
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT doc, created_at FROM doc_process WHERE cnj=? ORDER BY created_at DESC", (cnj,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "cnj": cnj, "count": len(rows), "items": rows})


@app.post("/docs/link")
def create_doc_link():
    payload = request.get_json(force=True, silent=False) or {}
    doc = (payload.get("doc") or "").strip()
    cnj = (payload.get("cnj") or "").strip()
    if not doc or not cnj:
        abort(400, description="Campos 'doc' e 'cnj' são obrigatórios.")
    try:
        docn = normalize_doc(doc)
    except Exception:
        abort(400, description="Doc inválido.")
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")

    conn = db_connect()
    ensure_process_registered(conn, cnj)
    link_doc_process(conn, docn, cnj)
    conn.close()
    return jsonify({"ok": True, "doc": docn, "cnj": cnj})


@app.get("/docs/<path:doc>/processos")
def list_processes_for_doc(doc: str):
    try:
        docn = normalize_doc(doc)
    except Exception:
        abort(400, description="Doc inválido.")
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "SELECT cnj, created_at FROM doc_process WHERE doc=? ORDER BY created_at DESC LIMIT 500",
        (docn,),
    )
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "doc": docn, "count": len(items), "items": items})


@app.post("/docs/<path:doc>/discover")
def discover_and_link_processes(doc: str):
    require_token_configured()
    try:
        docn = normalize_doc(doc)
    except Exception:
        abort(400, description="Doc inválido.")
    limit = int((request.args.get("limit") or "50").strip())
    if limit <= 0:
        limit = 50

    discovered: List[str] = []
    linked: List[Dict[str, str]] = []
    inserted_processos = 0

    # Endpoint: /api/v2/envolvido/processos?cpf_cnpj=...
    try:
        resp = client.listar_processos_envolvido(docn, limit=limit, page=1)  # type: ignore
    except requests.exceptions.RequestException as e:
        logger.error("Discover failed (network/API) doc=%s: %s", docn, repr(e)[:400])
        return jsonify({"ok": False, "doc": docn, "error": "Falha ao consultar Escavador (conexão/API). Tente novamente."}), 502
    processos = extract_list(resp)
    # fallback: algumas respostas vêm como {"processos":[...]}
    if not processos and isinstance(resp, dict) and isinstance(resp.get("processos"), list):
        processos = resp["processos"]

    conn = db_connect()
    for p in processos:
        if not isinstance(p, dict):
            continue
        cnj = (p.get("numero_cnj") or p.get("numero") or p.get("cnj") or "").strip()
        if not cnj or not CNJ_REGEX.fullmatch(cnj):
            continue
        discovered.append(cnj)
        if upsert_processo(conn, cnj):
            inserted_processos += 1
        link_doc_process(conn, docn, cnj)
        linked.append({"doc": docn, "cnj": cnj})
    conn.close()

    return jsonify(
        {
            "ok": True,
            "doc": docn,
            "discovered": len(discovered),
            "inserted_processos": inserted_processos,
            "linked": linked[:500],
        }
    )


@app.get("/docs/<path:doc>/alerts")
def get_alerts_for_doc(doc: str):
    try:
        docn = normalize_doc(doc)
    except Exception:
        abort(400, description="Doc inválido.")

    conn = db_connect()
    cur = conn.cursor()

    cur.execute("SELECT last_event_id FROM alert_state WHERE doc=?", (docn,))
    row = cur.fetchone()
    last_id = int(row["last_event_id"]) if row else 0

    cur.execute("SELECT cnj FROM doc_process WHERE doc=?", (docn,))
    cnjs = [r["cnj"] for r in cur.fetchall()]

    max_id = 0
    new_count = 0
    if cnjs:
        placeholders = ",".join(["?"] * len(cnjs))
        cur.execute(f"SELECT COALESCE(MAX(id),0) AS mid FROM eventos_mov WHERE cnj IN ({placeholders})", cnjs)
        max_id = int(cur.fetchone()["mid"] or 0)

        cur.execute(f"SELECT COUNT(*) AS c FROM eventos_mov WHERE id>? AND cnj IN ({placeholders})", [last_id, *cnjs])
        new_count = int(cur.fetchone()["c"] or 0)

    conn.close()
    return jsonify({"ok": True, "doc": docn, "new_events": new_count, "last_event_id": last_id, "max_event_id": max_id})


@app.post("/docs/<path:doc>/alerts/ack")
def ack_alerts_for_doc(doc: str):
    try:
        docn = normalize_doc(doc)
    except Exception:
        abort(400, description="Doc inválido.")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT cnj FROM doc_process WHERE doc=?", (docn,))
    cnjs = [r["cnj"] for r in cur.fetchall()]

    max_id = 0
    if cnjs:
        placeholders = ",".join(["?"] * len(cnjs))
        cur.execute(f"SELECT COALESCE(MAX(id),0) AS mid FROM eventos_mov WHERE cnj IN ({placeholders})", cnjs)
        max_id = int(cur.fetchone()["mid"] or 0)

    cur.execute(
        "INSERT INTO alert_state (doc, last_event_id, updated_at) VALUES (?, ?, ?) "
        "ON CONFLICT(doc) DO UPDATE SET last_event_id=excluded.last_event_id, updated_at=excluded.updated_at",
        (docn, max_id, utcnow_iso()),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "doc": docn, "acked_to_event_id": max_id})

@app.delete("/docs/link")
def delete_doc_link():
    payload = request.get_json(force=True, silent=False) or {}
    doc = (payload.get("doc") or "").strip()
    cnj = (payload.get("cnj") or "").strip()
    if not doc or not cnj:
        abort(400, description="Campos 'doc' e 'cnj' são obrigatórios.")
    try:
        docn = normalize_doc(doc)
    except Exception:
        abort(400, description="Doc inválido.")
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")

    conn = db_connect()
    deleted = unlink_doc_process(conn, docn, cnj)
    conn.close()
    return jsonify({"ok": True, "doc": docn, "cnj": cnj, "deleted": deleted})


# ============================================================
# API JSON: Processos + Eventos
# ============================================================
@app.post("/processos/<path:cnj>/sync")
def sync_process(cnj: str):
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")
    try:
        res = sync_process_movements(client, cnj, limit=400)  # type: ignore
        return jsonify({"ok": True, "cnj": cnj, "new_events": res.new_events})
    except requests.HTTPError as e:
        status = getattr(e.response, "status_code", 500)
        esc_msg = None
        try:
            j = e.response.json() if e.response is not None else {}
            esc_msg = j.get("error") or j.get("message")
        except Exception:
            esc_msg = str(e)

        if status == 402:
            return (
                jsonify(
                    {
                        "ok": False,
                        "error": "SEM_CREDITO_API",
                        "message": esc_msg or "Você não possui saldo em crédito da API do Escavador.",
                        "cnj": cnj,
                    }
                ),
                402,
            )

        return (
            jsonify(
                {
                    "ok": False,
                    "error": "ESCAVADOR_HTTP_ERROR",
                    "status": status,
                    "message": esc_msg or "Erro ao consultar o Escavador.",
                    "cnj": cnj,
                }
            ),
            status,
        )

    except requests.RequestException as e:
        # Conexão/timeout/etc (não é HTTPError com response)
        logger.error("Sync failed (network): %s", str(e))
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "ESCAVADOR_CONEXAO",
                    "status": 502,
                    "message": str(e) or "Falha de conexão ao consultar o Escavador.",
                    "cnj": cnj,
                }
            ),
            502,
        )



# ============================================================
# API JSON: Atualização (tribunal) + Documentos (públicos/autos)
# ============================================================

def _docs_cache_upsert(cnj: str, tipo: str, item: Dict[str, Any]) -> None:
    """Armazena metadados de documento no cache local (SQLite)."""
    key = (item.get("key") or item.get("chave") or item.get("id") or "").strip()
    if not key:
        return
    titulo = item.get("titulo") or item.get("nome") or item.get("descricao") or None
    data_doc = item.get("data") or item.get("data_documento") or item.get("data_protocolo") or None
    mime = item.get("mime") or item.get("mime_type") or item.get("content_type") or None
    links = item.get("links") or {}
    download_url = None
    if isinstance(links, dict):
        download_url = links.get("download") or links.get("url") or links.get("api")
    meta_json = safe_json_dumps(item)

    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO documentos_cache (cnj, tipo, doc_key, titulo, data, mime, meta_json, download_url, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?)
        ON CONFLICT(cnj, tipo, doc_key) DO UPDATE SET
          titulo=excluded.titulo,
          data=excluded.data,
          mime=excluded.mime,
          meta_json=excluded.meta_json,
          download_url=excluded.download_url,
          updated_at=excluded.updated_at
        """,
        (cnj, tipo, key, titulo, data_doc, mime, meta_json, download_url, utcnow_iso()),
    )
    conn.commit()
    conn.close()


def _docs_cache_list(cnj: str, tipo: str) -> List[Dict[str, Any]]:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "SELECT doc_key, titulo, data, mime, meta_json, download_url, updated_at FROM documentos_cache WHERE cnj=? AND tipo=? ORDER BY COALESCE(data,'') DESC, doc_key",
        (cnj, tipo),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    # meta_json volta como string; manter também um 'meta' parsed opcional
    for r in rows:
        try:
            r["meta"] = json.loads(r.get("meta_json") or "{}")
        except Exception:
            r["meta"] = {}
    return rows


def _update_row_upsert(cnj: str, tipo: str, *, status: str, escavador_update_id: Optional[int] = None, last_error: Optional[str] = None) -> None:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO processo_updates (cnj, tipo, escavador_update_id, status, created_at, updated_at, last_error)
        VALUES (?,?,?,?,?,?,?)
        ON CONFLICT DO NOTHING
        """,
        (cnj, tipo, escavador_update_id, status, utcnow_iso(), utcnow_iso(), last_error),
    )
    # sempre atualizar o registro mais recente do tipo/cnj
    cur.execute(
        """
        UPDATE processo_updates
           SET escavador_update_id = COALESCE(?, escavador_update_id),
               status=?,
               updated_at=?,
               last_error=?
         WHERE id = (
            SELECT id FROM processo_updates WHERE cnj=? AND tipo=? ORDER BY id DESC LIMIT 1
         )
        """,
        (escavador_update_id, status, utcnow_iso(), last_error, cnj, tipo),
    )
    conn.commit()
    conn.close()


def _update_row_latest(cnj: str, tipo: str) -> Optional[Dict[str, Any]]:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM processo_updates WHERE cnj=? AND tipo=? ORDER BY id DESC LIMIT 1",
        (cnj, tipo),
    )
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None


def _extract_status_value(status_payload: Dict[str, Any]) -> Optional[str]:
    """Normaliza status da API em algo como PENDENTE/SUCESSO/ERRO."""
    if not isinstance(status_payload, dict):
        return None
    # a doc mostra "status" e outros campos, mas vamos ser permissivos
    for k in ("status", "estado", "situacao"):
        v = status_payload.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip().upper()
    return None


def process_updates_once(client: EscavadorClient, max_items: int = 10) -> Dict[str, Any]:
    """Verifica status de solicitações pendentes e, ao concluir, puxa docs/autos e grava no cache."""
    conn = db_connect()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, cnj, tipo, status FROM processo_updates WHERE status IN ('PENDENTE','ERRO') ORDER BY updated_at ASC LIMIT ?",
        (max_items,),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    checked = 0
    completed = 0
    errors = 0

    for r in rows:
        cnj = r["cnj"]
        tipo = r["tipo"]
        try:
            st = client.status_atualizacao_processo(cnj)
            checked += 1
            st_val = _extract_status_value(st) or "PENDENTE"
            if st_val == "SUCESSO":
                _update_row_upsert(cnj, tipo, status="SUCESSO", last_error=None)
                # puxa lista e cacheia
                if tipo == "documentos_publicos":
                    data = client.listar_documentos_publicos(cnj, limit=100)
                    items = extract_list(data)
                    for it in items:
                        _docs_cache_upsert(cnj, "publicos", it)
                elif tipo == "autos":
                    data = client.listar_autos(cnj, limit=50)
                    items = extract_list(data)
                    for it in items:
                        _docs_cache_upsert(cnj, "autos", it)
                completed += 1
            elif st_val in ("ERRO", "FALHA", "FAILED"):
                _update_row_upsert(cnj, tipo, status="ERRO", last_error=safe_json_dumps(st)[:1000])
                errors += 1
            else:
                _update_row_upsert(cnj, tipo, status="PENDENTE", last_error=None)
        except Exception as ex:
            _update_row_upsert(cnj, tipo, status="ERRO", last_error=str(ex)[:1000])
            errors += 1

    return {"checked": checked, "completed": completed, "errors": errors}


@app.post("/processos/<path:cnj>/solicitar-atualizacao")
def solicitar_atualizacao(cnj: str):
    """Solicita atualização no tribunal, com escolha entre documentos públicos ou autos."""
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")
    payload = request.get_json(force=True, silent=True) or {}
    tipo = (payload.get("tipo") or "").strip().lower()
    if tipo not in ("documentos_publicos", "autos"):
        abort(400, description="tipo deve ser 'documentos_publicos' ou 'autos'.")
    try:
        if tipo == "documentos_publicos":
            res = client.solicitar_atualizacao_processo(cnj, documentos_publicos=True, autos=False, enviar_callback=False)  # type: ignore
        else:
            # autos exigem autenticação (certificado ou user/senha). Aqui só repassamos o que vier.
            res = client.solicitar_atualizacao_processo(
                cnj,
                documentos_publicos=False,
                autos=True,
                enviar_callback=False,
                utilizar_certificado=payload.get("utilizar_certificado"),
                certificado_id=payload.get("certificado_id"),
                usuario=payload.get("usuario"),
                senha=payload.get("senha"),
                documentos_especificos=payload.get("documentos_especificos"),
            )  # type: ignore

        _update_row_upsert(cnj, tipo, status="PENDENTE", last_error=None)
        return jsonify({"ok": True, "cnj": cnj, "tipo": tipo, "result": res, "pending": True, "message": "Solicitação enviada ao Escavador. Atualização em andamento."})
    except EscavadorUpdateAlreadyRunning as e:
        payload_remote = e.payload if isinstance(e.payload, dict) else {}
        ultima = (payload_remote.get("appends") or {}).get("ultima_verificacao") or {}
        remote_status = ((ultima.get("status") or payload_remote.get("status") or "PENDENTE").strip().upper() if isinstance((ultima.get("status") or payload_remote.get("status") or "PENDENTE"), str) else "PENDENTE")
        _update_row_upsert(cnj, tipo, status="PENDENTE", last_error=None, escavador_update_id=ultima.get("id"))
        return jsonify({
            "ok": True,
            "cnj": cnj,
            "tipo": tipo,
            "pending": True,
            "already_running": True,
            "message": str(e),
            "status": remote_status,
            "ultima_verificacao": ultima,
        }), 200
    except requests.HTTPError as e:
        status = getattr(e.response, "status_code", 500)
        msg = None
        try:
            msg = (e.response.json() or {}).get("error") or (e.response.json() or {}).get("message")
        except Exception:
            msg = str(e)
        _update_row_upsert(cnj, tipo, status="ERRO", last_error=(msg or str(e))[:1000])
        return jsonify({"ok": False, "error": "ESCAVADOR_HTTP_ERROR", "status": status, "message": msg, "cnj": cnj}), status
    except Exception as ex:
        _update_row_upsert(cnj, tipo, status="ERRO", last_error=str(ex)[:1000])
        return jsonify({"ok": False, "error": "ERRO_INTERNO", "message": str(ex), "cnj": cnj}), 500


@app.get("/processos/<path:cnj>/status-atualizacao")
def status_atualizacao(cnj: str):
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")
    try:
        st = client.status_atualizacao_processo(cnj)  # type: ignore
        return jsonify({"ok": True, "cnj": cnj, "status": st, "local": {
            "documentos_publicos": _update_row_latest(cnj, "documentos_publicos"),
            "autos": _update_row_latest(cnj, "autos"),
        }})
    except requests.HTTPError as e:
        status = getattr(e.response, "status_code", 500)
        msg = None
        try:
            msg = (e.response.json() or {}).get("error") or (e.response.json() or {}).get("message")
        except Exception:
            msg = str(e)
        return jsonify({"ok": False, "error": "ESCAVADOR_HTTP_ERROR", "status": status, "message": msg, "cnj": cnj}), status


@app.get("/processos/<path:cnj>/documentos-publicos")
def api_list_documentos_publicos(cnj: str):
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")
    # primeiro tenta cache
    cached = _docs_cache_list(cnj, "publicos")
    if cached:
        return jsonify({"ok": True, "cnj": cnj, "tipo": "publicos", "source": "cache", "items": cached})

    # fallback: tentar listar direto
    try:
        data = client.listar_documentos_publicos(cnj, limit=100)  # type: ignore
        items = extract_list(data)
        for it in items:
            _docs_cache_upsert(cnj, "publicos", it)
        return jsonify({"ok": True, "cnj": cnj, "tipo": "publicos", "source": "api", "items": _docs_cache_list(cnj, "publicos")})
    except Exception as ex:
        return jsonify({"ok": True, "cnj": cnj, "tipo": "publicos", "source": "empty", "items": [], "warning": str(ex)})


@app.get("/processos/<path:cnj>/autos")
def api_list_autos(cnj: str):
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")
    cached = _docs_cache_list(cnj, "autos")
    if cached:
        return jsonify({"ok": True, "cnj": cnj, "tipo": "autos", "source": "cache", "items": cached})

    try:
        data = client.listar_autos(cnj, limit=50)  # type: ignore
        items = extract_list(data)
        for it in items:
            _docs_cache_upsert(cnj, "autos", it)
        return jsonify({"ok": True, "cnj": cnj, "tipo": "autos", "source": "api", "items": _docs_cache_list(cnj, "autos")})
    except Exception as ex:
        return jsonify({"ok": True, "cnj": cnj, "tipo": "autos", "source": "empty", "items": [], "warning": str(ex)})


@app.get("/processos/<path:cnj>/documentos/<path:key>")
def api_get_documento_key(cnj: str, key: str):
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")
    data = client.obter_documento_por_key(cnj, key)  # type: ignore
    return jsonify({"ok": True, "cnj": cnj, "key": key, "data": data})


@app.get("/processos/<path:cnj>/documentos/<path:key>/download")
def api_download_documento_key(cnj: str, key: str):
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        abort(400, description="CNJ inválido.")
    data = client.obter_documento_por_key(cnj, key)  # type: ignore
    links = (data or {}).get("links") or {}
    url = None
    if isinstance(links, dict):
        url = links.get("download") or links.get("url")
    if not url:
        url = (data or {}).get("url")
    if not url:
        abort(404, description="Link de download não disponível para este documento.")
    return redirect(url)


# ============================================================
# UI API: Documentos (públicos/autos)
# ============================================================
@app.get("/ui/api/processo/<path:cnj>/documentos")
def ui_api_documentos(cnj: str):
    if not CNJ_REGEX.fullmatch(cnj):
        return jsonify({"ok": False, "error": "CNJ_INVALIDO", "message": "CNJ inválido."}), 400

    tipo = (request.args.get("tipo") or "publicos").strip().lower()
    if tipo not in ("publicos", "autos"):
        return jsonify({"ok": False, "error": "TIPO_INVALIDO", "message": "tipo deve ser publicos|autos"}), 400

    # cache sempre primeiro
    cached = _docs_cache_list(cnj, tipo)
    if cached:
        return jsonify({"ok": True, "cnj": cnj, "tipo": tipo, "source": "cache", "items": cached})

    # tentativa via API (pode falhar caso ainda não tenha solicitado atualização)
    try:
        require_token_configured()
        if tipo == "publicos":
            data = client.listar_documentos_publicos(cnj, limit=100)  # type: ignore
        else:
            data = client.listar_autos(cnj, limit=50)  # type: ignore
        items = extract_list(data)
        for it in items:
            _docs_cache_upsert(cnj, tipo, it)
        return jsonify({"ok": True, "cnj": cnj, "tipo": tipo, "source": "api", "items": _docs_cache_list(cnj, tipo)})
    except Exception as ex:
        return jsonify({"ok": True, "cnj": cnj, "tipo": tipo, "source": "empty", "items": [], "warning": str(ex)})


@app.post("/ui/api/processo/<path:cnj>/solicitar-atualizacao")
def ui_api_solicitar_atualizacao(cnj: str):
    if not CNJ_REGEX.fullmatch(cnj):
        return jsonify({"ok": False, "error": "CNJ_INVALIDO", "message": "CNJ inválido."}), 400
    payload = request.get_json(force=True, silent=True) or {}
    tipo = (payload.get("tipo") or "").strip().lower()
    if tipo not in ("documentos_publicos", "autos"):
        return jsonify({"ok": False, "error": "TIPO_INVALIDO", "message": "tipo deve ser documentos_publicos|autos"}), 400
    # reaproveita rota API
    return solicitar_atualizacao(cnj)


@app.get("/ui/api/processo/<path:cnj>/status-atualizacao")
def ui_api_status_atualizacao(cnj: str):
    if not CNJ_REGEX.fullmatch(cnj):
        return jsonify({"ok": False, "error": "CNJ_INVALIDO", "message": "CNJ inválido."}), 400
    return status_atualizacao(cnj)


@app.get("/ui/api/processo/<path:cnj>/documentos/<path:key>/download")
def ui_api_download_documento(cnj: str, key: str):
    # Proxy/redirect para download
    return api_download_documento_key(cnj, key)

@app.get("/processos/<path:cnj>/movimentacoes")
def get_movs(cnj: str):
    limit = int(request.args.get("limit", "50"))
    offset = int(request.args.get("offset", "0"))
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as c FROM eventos_mov WHERE cnj=?", (cnj,))
    total = int(cur.fetchone()["c"])
    cur.execute(
        """SELECT id, data, tipo, tipo_inferido, texto, created_at
           FROM eventos_mov
           WHERE cnj=?
           ORDER BY COALESCE(data, created_at) DESC
           LIMIT ? OFFSET ?""",
        (cnj, limit, offset),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "cnj": cnj, "total": total, "limit": limit, "offset": offset, "items": rows})


@app.get("/processos/<path:cnj>/movimentacoes/busca")
def search_movs(cnj: str):
    q = (request.args.get("q") or "").strip()
    tipo = (request.args.get("tipo") or "").strip().upper()
    limit = int(request.args.get("limit", "50"))
    offset = int(request.args.get("offset", "0"))

    where = ["cnj=?"]
    params: List[Any] = [cnj]

    if q:
        where.append("texto LIKE ?")
        params.append(f"%{q}%")
    if tipo:
        where.append("(UPPER(COALESCE(tipo,''))=? OR UPPER(COALESCE(tipo_inferido,''))=?)")
        params.extend([tipo, tipo])

    where_sql = " AND ".join(where)

    conn = db_connect()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) as c FROM eventos_mov WHERE {where_sql}", params)
    total = int(cur.fetchone()["c"])

    cur.execute(
        f"""SELECT id, data, tipo, tipo_inferido, texto, created_at
            FROM eventos_mov
            WHERE {where_sql}
            ORDER BY COALESCE(data, created_at) DESC
            LIMIT ? OFFSET ?""",
        params + [limit, offset],
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify({"ok": True, "cnj": cnj, "total": total, "limit": limit, "offset": offset, "items": rows})


# ============================================================
# Webhook + Inbox
# ============================================================
@app.post("/webhook/escavador")
def webhook_escavador():
    if WEBHOOK_AUTH_TOKEN:
        incoming = (request.headers.get("Authorization") or "").strip()
        if incoming != WEBHOOK_AUTH_TOKEN:
            abort(401, description="Authorization inválido no webhook.")
    payload = request.get_json(force=True, silent=False) or {}
    inserted, h = ingest_callback("webhook", payload)
    return jsonify({"ok": True, "inserted": inserted, "payload_hash": h})


@app.route("/poll/run-once", methods=["GET","POST"])
def poll_run_once():
    require_token_configured()
    data = client.listar_callbacks(limit=100, page=1)  # type: ignore
    callbacks = extract_list(data)

    inserted = 0
    for cb in callbacks:
        ins, _ = ingest_callback("poll", cb)
        if ins:
            inserted += 1

    pr = process_inbox_once(client, max_items=50)  # type: ignore
    return jsonify({"ok": True, "inserted": inserted, **pr})


# ============================================================
# UI (web) - Blue Service skin
# ============================================================

UI_BASE = """
<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Specter</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    :root{
      --bs-primary: #3050A0;
      --bs-primary-rgb: 48,80,160;
      --ink: #0f172a;
      --muted: #5b6b86;
      --bg: #F4F7FF;
      --card: #ffffff;
      --border: rgba(30,43,90,.14);
      --shadow: 0 10px 26px rgba(30,43,90,.12);
      --accent: #F01040;
      --accent2: #7C3AED;
      --chip: #EEF3FF;
      --chipText: #28407e;
      --navbg: rgba(48,80,160,.06);
      --ok: #16a34a;
      --warn: #f59e0b;
    }
    body { background: var(--bg); color: var(--ink); }
    .navbar { background: linear-gradient(90deg, rgba(48,80,160,.98), rgba(30,43,90,.98)); }
    .navbar a { color: #fff !important; text-decoration:none; }
    .pill {
      border: 1px solid rgba(255,255,255,.25);
      background: rgba(255,255,255,.10);
      border-radius: 999px;
      padding: .35rem .75rem;
      text-decoration: none;
    }
    .pill:hover { background: rgba(255,255,255,.16); }
    .card { background: var(--card); border: 1px solid var(--border); border-radius: 18px; box-shadow: var(--shadow); }
    .muted { color: var(--muted); }
    a { color: var(--bs-primary); text-decoration: none; }
    a:hover { text-decoration: underline; }
    .mono { font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace; }
    .wrap { white-space: pre-wrap; }
    .btn-brand {
      background: linear-gradient(90deg, var(--bs-primary), #2b63c9);
      border: 0; color: #fff; font-weight: 650;
      border-radius: 12px;
      box-shadow: 0 10px 20px rgba(48,80,160,.22);
    }
    .btn-brand:hover { filter: brightness(1.04); }
    .btn-accent {
      background: linear-gradient(90deg, var(--accent), var(--accent2));
      border: 0; color: #fff; font-weight: 650;
      border-radius: 12px;
      box-shadow: 0 10px 20px rgba(240,16,64,.18);
    }
    .btn-accent:hover { filter: brightness(1.04); }
    .btn-mini { padding: .28rem .55rem; font-size: .85rem; border-radius: 10px; }
    .chip {
      display: inline-flex; align-items: center; gap: .4rem;
      padding: .35rem .6rem; border-radius: 999px;
      background: var(--chip); border: 1px solid rgba(48,80,160,.18);
      color: var(--chipText); font-size: .88rem;
    }

    .chip { max-width: 100%; min-width: 0; overflow: hidden; }
    .chip .chip-val { min-width: 0; max-width: 100%; display: inline-block; }
    .truncate { overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .wrap-any { overflow-wrap: anywhere; word-break: break-word; }
    .badge-soft {
      background: rgba(48,80,160,.08);
      border: 1px solid rgba(48,80,160,.18);
      color: #27407d;
      border-radius: 14px;
    }
    .nav-tabs { border-bottom: 1px solid var(--border); }
    .nav-tabs .nav-link { border-radius: 14px 14px 0 0; color: #2a3a62; background: transparent; border: 1px solid transparent; }
    .nav-tabs .nav-link.active { background: #fff; border-color: var(--border); color: var(--bs-primary); }
    .nav-pills .nav-link { border-radius: 999px; background: var(--navbg); color: #2a3a62; }
    .nav-pills .nav-link.active { background: var(--bs-primary); color: #fff; }
    .table td, .table th { border-color: var(--border); }
    .kpi {
      border-radius: 16px;
      border: 1px solid rgba(48,80,160,.16);
      background: linear-gradient(180deg, rgba(48,80,160,.08), rgba(48,80,160,.03));
      padding: 12px;
      height: 100%;
      display: flex;
      flex-direction: column;
      gap: 4px;
      overflow: hidden;
    }
.kpi-click{ cursor:pointer; }
.kpi-click:hover{ filter: brightness(0.98); box-shadow: 0 6px 18px rgba(0,0,0,.08); }
.kpi-click:active{ transform: translateY(1px); }

    .kpi .label { color: var(--muted); font-size: .86rem; }
    .kpi .value { font-weight: 750; font-size: 1.0rem; line-height: 1.15; overflow: hidden; text-overflow: ellipsis; }
    .kpi .value.wrap { white-space: normal; overflow-wrap: anywhere; word-break: break-word; font-size: .95rem; }
    details > summary { cursor: pointer; }

    .sticky-side{
      position: sticky;
      top: 16px;
      z-index: 3;
    }
    .side-card{
      border-radius: 18px;
      border: 1px solid rgba(48,80,160,.18);
      background: linear-gradient(180deg, rgba(255,255,255,1), rgba(255,255,255,.92));
      box-shadow: var(--shadow);
    }
    .side-title{ font-weight: 800; }
    .summary-head{ display:flex; flex-wrap:wrap; align-items:flex-start; justify-content:space-between; gap:10px; }
    .summary-main{ min-width:0; flex:1 1 180px; }
    #badge-hoje{ display:flex; justify-content:flex-end; flex:0 0 auto; max-width:100%; }
    .summary-metrics > [class*='col-']{ display:flex; min-width:0; }
.summary-metrics .kpi{ width:100%; min-height:128px; }
.summary-metrics .kpi .value{ white-space:normal; overflow-wrap:anywhere; word-break:break-word; }
    .docs-note{ border:1px solid rgba(48,80,160,.16); border-radius:14px; background:rgba(48,80,160,.04); padding:10px 12px; }
    @media (max-width: 991.98px){
      .sticky-side{ position:static; }
    }
    @media (max-width: 575.98px){
      .summary-head{ flex-direction:column; }
      #badge-hoje{ justify-content:flex-start; }
    }
    .divider{ border-color: var(--border); }
    .chip-click{
      cursor: pointer;
      user-select: none;
      transition: transform .05s ease-in-out, filter .15s ease-in-out;
    }
    .chip-click:hover{ filter: brightness(0.98); }
    .chip-click:active{ transform: scale(0.98); }
    .chip-active{
      background: rgba(48,80,160,.12);
      border-color: rgba(48,80,160,.35);
      color: #1f3f8b;
      font-weight: 750;
    }

    .timeline-day{
      margin-top: 14px;
      padding: 10px 12px;
      border-radius: 14px;
      border: 1px solid rgba(48,80,160,.18);
      background: rgba(48,80,160,.05);
      display:flex;
      justify-content: space-between;
      align-items:center;
      gap: 8px;
    }
    .tl-item{
      border: 1px solid var(--border);
      border-radius: 16px;
      padding: 12px;
      background: #fff;
      box-shadow: 0 8px 18px rgba(30,43,90,.06);
    }
    .tl-top{
      display:flex;
      justify-content: space-between;
      align-items:flex-start;
      gap: 12px;
      margin-bottom: 6px;
    }
    .tl-icon{
      width: 34px;
      height: 34px;
      border-radius: 12px;
      display:flex;
      align-items:center;
      justify-content:center;
      background: rgba(48,80,160,.08);
      border: 1px solid rgba(48,80,160,.18);
      flex: 0 0 auto;
    }
    .tl-title{
      font-weight: 800;
      color: #1f2a55;
    }
    .tl-meta{ color: var(--muted); font-size: .86rem; }
    .badge-today{
      background: rgba(22,163,74,.12);
      border: 1px solid rgba(22,163,74,.30);
      color: #166534;
      border-radius: 999px;
      padding: .18rem .55rem;
      font-weight: 750;
      font-size: .78rem;
      white-space: nowrap;
    }
    .badge-new{
      background: rgba(245,158,11,.14);
      border: 1px solid rgba(245,158,11,.34);
      color: #92400e;
      border-radius: 999px;
      padding: .18rem .55rem;
      font-weight: 750;
      font-size: .78rem;
      white-space: nowrap;
    }
/* Toasts */
#toast-host{
  position: fixed;
  top: 14px;
  right: 14px;
  z-index: 1080;
  display:flex;
  flex-direction:column;
  gap: 10px;
  pointer-events: none;
}
.toast{
  pointer-events: auto;
  border-radius: 16px;
  border: 1px solid var(--border);
  box-shadow: var(--shadow);
}
.toast .toast-header{
  border-top-left-radius: 16px;
  border-top-right-radius: 16px;
  background: rgba(48,80,160,.06);
  border-bottom: 1px solid var(--border);
}

/* Inline banners */
.inline-banner .alert{
  border-radius: 16px;
  box-shadow: var(--shadow);
  border: 1px solid var(--border);
}

  </style>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
</head>
<body>
<script>
(function(){
  // Helpers need to exist BEFORE any page scripts (body content may call them immediately)
  window.clearInlineBanner = window.clearInlineBanner || function(id){
    try{ var host=document.getElementById(id); if(host) host.innerHTML=""; }catch(e){}
  };
  window.showInlineBanner = window.showInlineBanner || function(id, kind, title, msg){
    try{
      var host=document.getElementById(id); if(!host) return;
      var k=(kind||"info");
      var t=title ? ("<b>"+title+"</b>") : "";
      var m=msg ? ("<div class='mt-1'>"+msg+"</div>") : "";
      host.innerHTML = "<div class='alert alert-"+k+" mb-2'>"+t+m+"</div>";
    }catch(e){}
  };

  window.showToast = window.showToast || function(kind, title, msg, delayMs){
    try{
      var host=document.getElementById("toastHost");
      if(!host){ console.log("[toast]", kind, title, msg); return; }
      var k=(kind||"info");
      var icon = (k==="success")?"✅":(k==="danger")?"⛔":(k==="warning")?"⚠️":"ℹ️";
      var id="t"+Math.random().toString(16).slice(2);
      var html = `
        <div id="${id}" class="toast align-items-center text-bg-${k} border-0 mb-2" role="alert" aria-live="assertive" aria-atomic="true">
          <div class="d-flex">
            <div class="toast-body">
              <div><span class="me-2">${icon}</span><b>${title||""}</b></div>
              <div class="small mt-1">${msg||""}</div>
            </div>
            <button type="button" class="btn-close btn-close-white me-2 m-auto" data-bs-dismiss="toast" aria-label="Close"></button>
          </div>
        </div>`;
      host.insertAdjacentHTML("beforeend", html);
      var el=document.getElementById(id);
      var t=new bootstrap.Toast(el, { delay: (delayMs||4500) });
      el.addEventListener("hidden.bs.toast", ()=>{ try{ el.remove(); }catch(e){} });
      t.show();
    }catch(e){
      console.log("[toast-fallback]", kind, title, msg);
    }
  };

  // Navbar badge (global pending alerts)
  async function updateNavBadge(){
    try{
      const r = await fetch("/ui/api/dashboard/metrics?scope=global");
      const j = await r.json();
      const n = (j && j.ok !== false && (j.alertas_global ?? j.alertas ?? 0)) || 0;
      const b = document.getElementById("navBadge");
      if(!b) return;
      if(n > 0){
        b.style.display = "inline-block";
        b.textContent = String(n);
      } else {
        b.style.display = "none";
        b.textContent = "";
      }
    }catch(e){}
  }
  document.addEventListener("DOMContentLoaded", function(){
    updateNavBadge();
    setInterval(updateNavBadge, 15000);
  });


})();
</script>
<div class="toast-container position-fixed top-0 end-0 p-3" style="z-index:1080" id="toastHost"></div>
<nav class="navbar navbar-dark">
  <div class="container py-2 d-flex justify-content-between align-items-center">
    <div>
      <div class="h4 mb-0">Specter</div>
      <div class="small" style="opacity:.9">Monitor jurídico local • Flask • SQLite • Movimentações em timeline</div>
    </div>
    <div class="d-flex gap-2">
      <a class="pill" href="/ui">Dashboard</a>
      <a class="pill" href="/ui/watchlist">Watch-List</a>
      <a class="pill" href="/ui/admin/monitoramentos">Monitoramentos</a>
      <a class="pill" href="/ui/financeiro">Financeiro</a>
      <a class="pill" href="/ui/admin">Admin <span id="navBadge" class="badge bg-danger ms-1" style="display:none"></span></a>
      <a class="pill" href="/health">Health</a>
    </div>
  </div>
</nav>
<div id="toast-host" aria-live="polite" aria-atomic="true"></div>
<div class="container py-4">
  {{ body|safe }}
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script>
(function(){
  let _toastSeq = 0;
  window.showToast = function(kind, title, msg, delayMs){
    const host = document.getElementById("toast-host");
    if(!host) return;
    const id = "t" + (++_toastSeq);
    const delay = (typeof delayMs === "number" ? delayMs : 4500);
    const icon = (kind==="success" ? "✅" : kind==="warning" ? "⚠️" : kind==="danger" ? "⛔" : "ℹ️");
    const html = `
    <div id="${id}" class="toast" role="status" aria-live="polite" aria-atomic="true">
      <div class="toast-header">
        <span class="me-2">${icon}</span>
        <strong class="me-auto">${title || "Aviso"}</strong>
        <small class="text-muted">agora</small>
        <button type="button" class="btn-close ms-2 mb-1" data-bs-dismiss="toast" aria-label="Close"></button>
      </div>
      <div class="toast-body">${msg || ""}</div>
    </div>`;
    host.insertAdjacentHTML("afterbegin", html);
    const el = document.getElementById(id);
    const t = bootstrap.Toast.getOrCreateInstance(el, { delay: delay });
    el.addEventListener("hidden.bs.toast", ()=>{ try{ el.remove(); }catch(e){} });
    t.show();
  }

  window.clearInlineBanner = function(id){
    var host = document.getElementById(id);
    if(host) host.innerHTML = "";
  };
  window.showInlineBanner = function(id, kind, title, msg){
    var host = document.getElementById(id);
    if(!host) return;
    var k = (kind||"info");
    var t = title ? ("<b>"+title+"</b>") : "";
    var m = msg ? ("<div class='mt-1'>"+msg+"</div>") : "";
    host.innerHTML = "<div class='alert alert-"+k+" mb-2'>"+t+m+"</div>";
  };
})();
</script>

</body>
</html>
"""


def ui_alert(kind: str, title: str, msg: str) -> str:
    return (
        "<div class='alert alert-" + kind + "' style='border-radius:16px; box-shadow: var(--shadow);'>"
        "<b>" + title + "</b><div class='mt-1'>" + msg + "</div></div>"
    )
# ----------------------------
# Runtime config: auto-discover toggle
# ----------------------------
@app.get("/admin/discover")
def get_discover_config():
    return jsonify(
        {
            "ok": True,
            "auto_discover_enabled": bool(AUTO_DISCOVER_ENABLED),
            "discover_interval_seconds": DISCOVER_INTERVAL_SECONDS,
            "discover_limit_per_doc": DISCOVER_LIMIT_PER_DOC,
            "discover_only_if_no_links": bool(DISCOVER_ONLY_IF_NO_LINKS),
            "discover_max_docs_per_cycle": DISCOVER_MAX_DOCS_PER_CYCLE,
            "discover_state": _get_discover_state(),
        }
    )


@app.post("/admin/discover")
def set_discover_config():
    """
    Atualiza configurações do auto-discover em runtime (sem reiniciar).
    Aceita JSON com qualquer combinação de:
      - enabled: bool
      - discover_interval_seconds: int (>=10 recomendado)
      - discover_limit_per_doc: int (>=1)
      - discover_only_if_no_links: bool
      - discover_max_docs_per_cycle: int (>=1)
    """
    global AUTO_DISCOVER_ENABLED, DISCOVER_INTERVAL_SECONDS, DISCOVER_LIMIT_PER_DOC, DISCOVER_ONLY_IF_NO_LINKS, DISCOVER_MAX_DOCS_PER_CYCLE

    payload = request.get_json(force=True, silent=True) or {}
    if not isinstance(payload, dict):
        abort(400, description="Payload inválido (JSON objeto esperado).")

    def _as_int(name: str, v: Any, min_v: int) -> int:
        try:
            iv = int(v)
        except Exception:
            abort(400, description=f"Campo '{name}' deve ser inteiro.")
        if iv < min_v:
            abort(400, description=f"Campo '{name}' deve ser >= {min_v}.")
        return iv

    changed: Dict[str, Any] = {}

    if "enabled" in payload:
        AUTO_DISCOVER_ENABLED = bool(payload.get("enabled"))
        changed["auto_discover_enabled"] = bool(AUTO_DISCOVER_ENABLED)

    if "discover_interval_seconds" in payload:
        DISCOVER_INTERVAL_SECONDS = _as_int("discover_interval_seconds", payload.get("discover_interval_seconds"), 1)
        changed["discover_interval_seconds"] = DISCOVER_INTERVAL_SECONDS

    if "discover_limit_per_doc" in payload:
        DISCOVER_LIMIT_PER_DOC = _as_int("discover_limit_per_doc", payload.get("discover_limit_per_doc"), 1)
        changed["discover_limit_per_doc"] = DISCOVER_LIMIT_PER_DOC

    if "discover_max_docs_per_cycle" in payload:
        DISCOVER_MAX_DOCS_PER_CYCLE = _as_int("discover_max_docs_per_cycle", payload.get("discover_max_docs_per_cycle"), 1)
        changed["discover_max_docs_per_cycle"] = DISCOVER_MAX_DOCS_PER_CYCLE

    if "discover_only_if_no_links" in payload:
        DISCOVER_ONLY_IF_NO_LINKS = bool(payload.get("discover_only_if_no_links"))
        changed["discover_only_if_no_links"] = bool(DISCOVER_ONLY_IF_NO_LINKS)

    if not changed:
        abort(400, description="Informe ao menos um campo para atualizar.")

    logger.info("Auto-discover config updated: %s", changed)

    return jsonify(
        {
            "ok": True,
            "auto_discover_enabled": bool(AUTO_DISCOVER_ENABLED),
            "discover_interval_seconds": DISCOVER_INTERVAL_SECONDS,
            "discover_limit_per_doc": DISCOVER_LIMIT_PER_DOC,
            "discover_only_if_no_links": bool(DISCOVER_ONLY_IF_NO_LINKS),
            "discover_max_docs_per_cycle": DISCOVER_MAX_DOCS_PER_CYCLE,
            "discover_state": _get_discover_state(),
            "changed": changed,
        }
    )





# ----------------------------
# Auto-discover: status + run once (manual trigger)
# ----------------------------
@app.get("/admin/discover/status")
def discover_status():
    st = _get_discover_state()
    return jsonify({"ok": True, "state": st})

@app.post("/admin/discover/run-once")
def discover_run_once():
    require_token_configured()
    if client is None:
        return jsonify({"ok": False, "error": "CLIENT_NOT_READY"}), 500

    # dispara em background para não travar a UI
    st = _get_discover_state()
    if st.get("running"):
        return jsonify({"ok": False, "error": "ALREADY_RUNNING", "state": st}), 409

    def _bg():
        try:
            run_discover_cycle(client, trigger="manual")  # type: ignore
        except Exception:
            logger.exception("Manual discover cycle crashed")

    threading.Thread(target=_bg, daemon=True).start()
    return jsonify({"ok": True, "started": True, "state": _get_discover_state()})


@app.post("/admin/costs/import")
def admin_costs_import():
    """
    Importa extrato do Escavador (linhas como):
    GET /api/v2/processos/numero_cnj/XXXX  R$ -0,04  16/02/2026 19:03
    """
    data = request.get_json(silent=True) or {}
    text = (data.get("text") or "").strip()
    if not text:
        return jsonify({"ok": False, "error": "Texto vazio"}), 400

    inserted = 0
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        m = EXTRATO_LINE_RE.match(line)
        if not m:
            continue
        method, endpoint, val_s, d_s, t_s = m.groups()
        try:
            cost = abs(_parse_brl_value(val_s))
        except Exception:
            continue
        try:
            dt = datetime.strptime(d_s + " " + t_s, "%d/%m/%Y %H:%M").replace(tzinfo=timezone.utc).isoformat()
        except Exception:
            dt = utcnow_iso()

        cnj = _extract_cnj_from_endpoint(endpoint)
        doc = _extract_doc_from_endpoint(endpoint)

        record_api_usage_real(ts_iso=dt, doc=doc, cnj=cnj, method=method, endpoint=endpoint, cost_brl=cost, raw_line=line)
        inserted += 1

    return jsonify({"ok": True, "inserted": inserted})

@app.post("/admin/costs/import-xlsx")
def admin_costs_import_xlsx():
    """Importa XLSX exportado do painel do Escavador (Relatório de Consumo da API).
    Espera colunas como: URL, Método HTTP, Saldo utilizado, Data de Utilização.
    Faz dedupe via fingerprint (UNIQUE) para não importar o mesmo dado mais de uma vez.
    """
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Arquivo não enviado (campo 'file')"}), 400
    f = request.files["file"]
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "Arquivo inválido"}), 400

    imported_at = utcnow_iso()
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # map headers (case-insensitive, tolerant)
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is None:
                continue
            headers[str(v).strip().lower()] = c

        def h(*names):
            for n in names:
                c = headers.get(n.lower())
                if c:
                    return c
            return None

        col_url = h("url")
        col_method = h("método http", "metodo http", "método", "metodo")
        col_cost = h("saldo utilizado", "saldo Utilizado")
        col_ts = h("data de utilização", "data de utilizacao", "data de utilização", "data de utilização")

        if not (col_url and col_method and col_cost and col_ts):
            return jsonify({
                "ok": False,
                "error": "XLSX sem colunas esperadas. Preciso de URL, Método HTTP, Saldo utilizado, Data de Utilização.",
                "found_headers": list(headers.keys())
            }), 400

        inserted = 0
        skipped = 0
        errors = 0

        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            # garante coluna/índice (para DBs antigos)
            try:
                cur.execute("PRAGMA table_info(api_usage_real)")
                cols = [r[1] for r in cur.fetchall()]
                if "fingerprint" not in cols:
                    cur.execute("ALTER TABLE api_usage_real ADD COLUMN fingerprint TEXT")
                if "imported_at" not in cols:
                    cur.execute("ALTER TABLE api_usage_real ADD COLUMN imported_at TEXT")
                if "source" not in cols:
                    cur.execute("ALTER TABLE api_usage_real ADD COLUMN source TEXT")
            except Exception:
                pass
            cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_api_usage_real_fp ON api_usage_real (fingerprint)")

            for r in range(2, ws.max_row + 1):
                endpoint = ws.cell(r, col_url).value
                method = ws.cell(r, col_method).value
                cost = ws.cell(r, col_cost).value
                ts = ws.cell(r, col_ts).value

                if not endpoint or not method:
                    continue

                method_s = str(method).strip().upper()
                if method_s.lower().startswith("total"):
                    continue

                try:
                    cost_val = abs(float(cost))
                except Exception:
                    errors += 1
                    continue

                # ts pode vir como datetime ou string 'YYYY-MM-DD HH:MM:SS'
                try:
                    if hasattr(ts, "isoformat"):
                        ts_iso = ts.replace(tzinfo=timezone.utc).isoformat()  # type: ignore
                    else:
                        ts_s = str(ts).strip()
                        dt = datetime.strptime(ts_s, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                        ts_iso = dt.isoformat()
                except Exception:
                    ts_iso = utcnow_iso()

                endpoint_s = str(endpoint).strip()
                cnj = _extract_cnj_from_endpoint(endpoint_s)
                doc = _extract_doc_from_endpoint(endpoint_s)

                raw_line = f"{method_s} {endpoint_s} R$ -{cost_val:.2f} {ts_iso}"
                fp = _fingerprint_usage(ts_iso, method_s, endpoint_s, float(cost_val))

                try:
                    cur.execute(
                        "INSERT OR IGNORE INTO api_usage_real (ts, doc, cnj, method, endpoint, cost_brl, raw_line, fingerprint, imported_at, source) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (ts_iso, doc, cnj, method_s, endpoint_s, float(cost_val), raw_line[:800], fp, imported_at, "xlsx"),
                    )
                    if cur.rowcount == 1:
                        inserted += 1
                    else:
                        skipped += 1
                except Exception:
                    errors += 1

            conn.commit()

        return jsonify({"ok": True, "inserted": inserted, "skipped": skipped, "errors": errors, "imported_at": imported_at})
    except Exception as e:
        logger.exception("Falha ao importar XLSX")
        return jsonify({"ok": False, "error": str(e)}), 500




@app.post("/admin/costs/clear-real")
def admin_costs_clear_real():
    """Limpa todo o histórico REAL importado (api_usage_real)."""
    try:
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM api_usage_real")
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logger.exception("Falha ao limpar api_usage_real")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.get("/ui")
def ui_home():
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT id, doc, tipo_doc, created_at FROM watchlist ORDER BY id DESC LIMIT 200")
    docs = [dict(r) for r in cur.fetchall()]
    conn.close()

    cards = ""
    if docs:
        for d in docs:
            doc = d["doc"]
            cards += """
            <div class="col-lg-6">
              <div class="card p-3">
                <div class="d-flex justify-content-between align-items-start">
                  <div>
                    <div class="muted small">""" + d["tipo_doc"] + """</div>
                    <div class="h5 mb-1 mono">""" + doc + """</div>
                    <div class="muted small">Criado em """ + d["created_at"] + """</div>
                  </div>
                  <div class="d-flex flex-column gap-2">
                    <a class="btn btn-outline-primary btn-mini" href="/ui/admin">Admin</a>
                    <button class="btn btn-outline-danger btn-mini" onclick="deleteDoc(""" + str(d["id"]) + """, '""" + doc + """')">Remover</button>
                  </div>
                </div>
              </div>
            </div>
            """
    else:
        cards = """
        <div class="col-12">
          <div class="card p-4">
            <div class="muted">Nenhum CPF/CNPJ na watchlist ainda. Cadastre no formulário ao lado.</div>
          </div>
        </div>
        """

    body = """
    
    <div id="dashboard" class="mb-4">
      <div class="d-flex align-items-center justify-content-between mb-2">
        <div>
          <div class="muted small">Visão geral</div>
          <h3 class="m-0">Dashboard</h3>
        </div>
        <div class="d-flex gap-2">
          <a class="btn btn-outline-primary btn-mini" href="/ui/admin">Admin</a>
                    <a class="btn btn-outline-secondary btn-mini" href="/health">Health</a>
        </div>
      </div>

      <div id="bn-dash"></div>

      <div class="d-flex flex-wrap align-items-center justify-content-between gap-2 mb-2">
  <div class="muted">Escopo das métricas</div>
  <div class="d-flex align-items-center gap-2">
    <select id="dash-scope" class="form-select form-select-sm" style="min-width:260px; max-width:420px;"></select>
    <button class="btn btn-sm btn-outline-primary" id="btn-dash-refresh">Atualizar</button>
  </div>
</div>
<div class="row g-3" id="dash-kpis">
        <div class="col-md-3"><div class="card p-3"><div class="muted">Docs monitorados</div><div class="h3 m-0" id="kpi-docs">-</div></div></div>
        <div class="col-md-3"><div class="card p-3"><div class="muted">Processos vinculados</div><div class="h3 m-0" id="kpi-processos">-</div></div></div>
        <div class="col-md-3"><div class="card p-3"><div class="muted">Movimentações salvas</div><div class="h3 m-0" id="kpi-movs">-</div></div></div>
        <div class="col-md-3"><div class="card p-3"><div class="muted">Alertas pendentes</div><div class="h3 m-0" id="kpi-alertas">-</div></div></div>
      </div>

      
      <div class="row g-3 mt-1">
        <div class="col-lg-7">
          <div class="card p-3">
            <div class="d-flex align-items-center justify-content-between">
              <div>
                <div class="fw-semibold">Movimentações salvas por dia</div>
                <div class="muted small">Janela padrão: 14 dias (filtra quando você escolhe um CPF/CNPJ).</div>
              </div>
              <div class="d-flex align-items-center gap-2">
                <select id="dash-days" class="form-select form-select-sm" style="width:110px;">
                  <option value="7">7d</option>
                  <option value="14" selected>14d</option>
                  <option value="30">30d</option>
                </select>
              </div>
            </div>
            <div class="mt-2">
              <canvas id="dash-chart" height="120"></canvas>
            </div>
            <div class="muted small mt-2" id="dash-chart-note"></div>
          </div>
        </div>
        <div class="col-lg-5">
          <div class="row g-3">
            <div class="col-12">
              <div class="card p-3">
                <div class="fw-semibold">Operação</div>
                <div class="muted small">Poll, Discover e último erro de API</div>
                <div class="mt-2 d-flex flex-column gap-2">
                  <div class="d-flex justify-content-between"><span class="muted">Poll</span><span id="ops-poll" class="mono small">-</span></div>
                  <div class="d-flex justify-content-between"><span class="muted">Discover</span><span id="ops-discover" class="mono small">-</span></div>
                  <div class="d-flex justify-content-between"><span class="muted">Último erro API</span><span id="ops-apierr" class="mono small">-</span></div>
                </div>
              </div>
            </div>

            <div class="col-12">
              <div class="card p-3">
                <div class="fw-semibold">Top CNJs com mais alertas</div>
                <div class="muted small">Ranking rápido (pendências).</div>
                <div class="mt-2 small" id="dash-top-cnj">-</div>
              </div>
            </div>
            <div class="col-12">
              <div class="card p-3">
                <div class="fw-semibold">Top CPF/CNPJ com mais alertas</div>
                <div class="muted small">Docs com mais pendências.</div>
                <div class="mt-2 small" id="dash-top-docs">-</div>
              </div>
            </div>
            <div class="col-12">
              <div class="card p-3">
                <div class="fw-semibold">Atalhos</div>
                <div class="d-flex gap-2 flex-wrap mt-2">
                  <a class="btn btn-outline-secondary btn-mini" href="#dashboard">Topo</a>
                  <a class="btn btn-outline-primary btn-mini" href="/ui/admin">Admin</a>
                                    <a class="btn btn-outline-secondary btn-mini" href="/health">Health</a>
                </div>
              </div>
            </div>
          </div>
        </div>
      </div>

      <div class="muted small mt-2">Atualiza a cada 5s (sem travar a tela).</div>
    </div>

    
      <div class="card p-3 mt-3">
        <div class="d-flex align-items-center justify-content-between">
          <div>
            <div class="fw-semibold">Resumo por CPF/CNPJ</div>
            <div class="muted">Alertas pendentes e quantidade de processos vinculados (top 50).</div>
          </div>
        </div>
        <div class="table-responsive mt-2">
          <table class="table table-sm align-middle mb-0">
            <thead><tr><th>Doc</th><th class="text-end">Processos</th><th class="text-end">Alertas</th></tr></thead>
            <tbody id="dash-doc-rows"><tr><td colspan="3" class="muted">Carregando...</td></tr></tbody>
          </table>
        </div>
      
<div class="card p-3 mt-3" id="dash-alerts-panel">
  <div class="d-flex align-items-start justify-content-between flex-wrap gap-2">
    <div>
      <div class="fw-semibold">Painel de alertas</div>
      <div class="muted small">Mostra movimentações <b>pendentes</b> (não zeradas). Você pode filtrar e exportar.</div>
    </div>
    <div class="d-flex gap-2 flex-wrap">
      <button class="btn btn-outline-secondary btn-mini" id="btnAlertsRefresh" type="button">Atualizar</button>
      <a class="btn btn-outline-secondary btn-mini" id="btnAlertsExport" href="#" target="_blank" rel="noopener">Exportar CSV</a>
    </div>
  </div>

  <div class="row g-2 mt-2 align-items-end">
    <div class="col-md-3">
      <label class="form-label small muted">Auto-refresh (segundos)</label>
      <input class="form-control form-control-sm" id="alerts-interval" type="number" min="0" step="1" placeholder="Ex: 10" />
      <div class="muted small">0 desliga.</div>
    </div>
    <div class="col-md-4">
      <label class="form-label small muted">Tipos (separado por vírgula)</label>
      <input class="form-control form-control-sm" id="alerts-types" type="text" placeholder="Ex: SENTENCA,DECISAO,PENHORA" />
    </div>
    <div class="col-md-3">
      <label class="form-label small muted">Busca livre</label>
      <input class="form-control form-control-sm" id="alerts-q" type="text" placeholder="Ex: penhora, SisbaJud..." />
    </div>
    <div class="col-md-2">
      <div class="form-check mt-4">
        <input class="form-check-input" type="checkbox" id="alerts-must-vp">
        <label class="form-check-label small" for="alerts-must-vp">Só valor/penhora</label>
      </div>
    </div>
  </div>

  <div id="bn-alerts" class="mt-2"></div>

  <div class="table-responsive mt-2" style="max-height: 420px; overflow:auto;">
    <table class="table table-sm align-middle mb-0">
      <thead>
        <tr>
          <th class="mono">ID</th>
          <th>Quando</th>
          <th>Doc</th>
          <th>CNJ</th>
          <th>Tipo</th>
          <th>Texto</th>
        </tr>
      </thead>
      <tbody id="dash-alert-rows"><tr><td colspan="6" class="muted">Carregando...</td></tr></tbody>
    </table>
  </div>
</div>

</div>
<script>
    let dashChart = null;

function _fmtOps(st){
  if(!st) return '-';
  const run = st.running ? 'rodando' : 'idle';
  const fin = st.last_finished ? ('fim ' + st.last_finished.replace('T',' ').slice(0,19)) : 'n/a';
  const err = st.last_error ? (' erro: ' + st.last_error) : '';
  return run + ' • ' + fin + err;
}

async function refreshSeries(doc){
  try{
    const daysSel = document.getElementById('dash-days');
    const days = daysSel ? (daysSel.value||'14') : '14';
    const url = '/ui/api/dashboard/timeseries?days=' + encodeURIComponent(days) + (doc ? ('&doc=' + encodeURIComponent(doc)) : '');
    const r = await fetch(url);
    const j = await r.json();
    if(!j.ok){
      const note = document.getElementById('dash-chart-note');
      if(note) note.textContent = 'Série indisponível: ' + (j.error||'');
      return;
    }
    const labels = (j.labels||[]).map(d=>{
      // d = YYYY-MM-DD
      const parts = String(d).split('-');
      return (parts.length===3) ? (parts[2]+'/'+parts[1]) : d;
    });
    const data = j.movs||[];
    const ctx = document.getElementById('dash-chart');
    if(!ctx) return;
    if(dashChart){ dashChart.destroy(); dashChart = null; }
    dashChart = new Chart(ctx, {
      type: 'line',
      data: { labels, datasets: [{ label: 'Movimentações', data, tension: 0.25 }] },
      options: { responsive: true, plugins: { legend: { display: false } }, scales: { y: { beginAtZero: true } } }
    });
    const note = document.getElementById('dash-chart-note');
    if(note) note.textContent = (doc?('Doc '+doc+' • '):'Global • ') + 'últimos ' + j.days + ' dias.';
  }catch(e){
    const note = document.getElementById('dash-chart-note');
    if(note) note.textContent = 'Erro ao carregar série: ' + e.toString();
  }
}

async function refreshDashboard(force=false){
      try{
        clearInlineBanner("bn-dash");

        const sel = document.getElementById("dash-scope");
        const doc = (sel && sel.value && sel.value !== "__global__") ? sel.value : "";
        const url = doc ? ("/ui/api/dashboard/metrics?doc=" + encodeURIComponent(doc)) : "/ui/api/dashboard/metrics";

        const r = await fetch(url);
        const j = await r.json();
        if(!j.ok){
          showInlineBanner("bn-dash","warning","Métricas indisponíveis", esc(j.error || j.message || "Falha ao calcular métricas"));
          return;
        }

        // Inicializa o seletor na primeira carga (usa a lista global se vier com docs_list)
        if(sel && (!sel.options || sel.options.length === 0)){
          const optG = document.createElement("option");
          optG.value = "__global__";
          optG.textContent = "Global (todos os docs)";
          sel.appendChild(optG);

          const list = Array.isArray(j.docs_list) ? j.docs_list : [];
          for(const d of list){
            const o = document.createElement("option");
            o.value = d;
            o.textContent = d;
            sel.appendChild(o);
          }
          sel.value = "__global__";
          sel.addEventListener("change", ()=>{ refreshDashboard(true); const doc2 = (sel.value && sel.value !== "__global__") ? sel.value : ""; setupAlertsAutoRefresh(doc2); });
        }

        // KPIs: se estiver filtrado por doc, usa per_doc; senão usa global
        const scope = (doc && j.per_doc) ? j.per_doc : j;
        document.getElementById("kpi-docs").textContent = doc ? (doc || "-") : String(j.docs ?? "-");
        document.getElementById("kpi-processos").textContent = String(scope.processos ?? "-");
        document.getElementById("kpi-movs").textContent = String(scope.movs ?? "-");
        document.getElementById("kpi-alertas").textContent = String(scope.alertas ?? "-");
        const costTodayEl = document.getElementById("kpi-cost-today"); if(costTodayEl) costTodayEl.textContent = (scope.cost_today_brl != null) ? ("R$ " + Number(scope.cost_today_brl).toFixed(2).replace(".", ",")) : "-";
        const costMonthEl = document.getElementById("kpi-cost-month"); if(costMonthEl) costMonthEl.textContent = (scope.cost_month_brl != null) ? ("R$ " + Number(scope.cost_month_brl).toFixed(2).replace(".", ",")) : "-";

        const realTodayEl = document.getElementById("kpi-cost-real-today");
        const realMonthEl = document.getElementById("kpi-cost-real-month");
        const deltaTodayEl = document.getElementById("kpi-cost-delta-today");
        const deltaMonthEl = document.getElementById("kpi-cost-delta-month");

        if(realTodayEl) realTodayEl.textContent = (scope.cost_real_today_brl != null) ? ("R$ " + Number(scope.cost_real_today_brl).toFixed(2).replace(".", ",")) : "-";
        if(realMonthEl) realMonthEl.textContent = (scope.cost_real_month_brl != null) ? ("R$ " + Number(scope.cost_real_month_brl).toFixed(2).replace(".", ",")) : "-";

        function _deltaFmt(v){
          if(v == null || isNaN(Number(v))) return "-";
          const n = Number(v);
          const s = "R$ " + n.toFixed(2).replace(".", ",");
          return (n > 0 ? ("+"+s) : s);
        }
        if(deltaTodayEl) deltaTodayEl.textContent = _deltaFmt(scope.cost_delta_today_brl);
        if(deltaMonthEl) deltaMonthEl.textContent = _deltaFmt(scope.cost_delta_month_brl);

        // Top CNJs
        const topCnjEl = document.getElementById("dash-top-cnj");
        if(topCnjEl){
          const items = (j.top_cnj_alerts || []);
          topCnjEl.innerHTML = items.length ? items.map(x=>{
            const cnj = x.cnj || x[0] || "";
            const n = (x.alertas !== undefined ? x.alertas : (x.count !== undefined ? x.count : (x[1]||0)));
            return `<a class="chip me-1 mb-1 d-inline-flex" href="/ui/processo/${encodeURIComponent(cnj)}">${esc(cnj)}<span class="ms-2 badge text-bg-light">${n}</span></a>`;
          }).join(" ") : `<span class="muted">Sem dados</span>`;
        }
        // Top Docs
        const topDocsEl = document.getElementById("dash-top-docs");
        if(topDocsEl){
          const items = (j.top_docs_alerts || []);
          topDocsEl.innerHTML = items.length ? items.map(x=>{
            const doc = x.doc || x[0] || "";
            const n = (x.alertas !== undefined ? x.alertas : (x.count !== undefined ? x.count : (x[1]||0)));
            return `<a class="chip me-1 mb-1 d-inline-flex" href="/ui/watchlist#${encodeURIComponent(doc)}">${esc(doc)}<span class="ms-2 badge text-bg-light">${n}</span></a>`;
          }).join(" ") : `<span class="muted">Sem dados</span>`;
        }

                // Operação
        const ps = j.poll_state;
        const ds = j.discover_state;
        const ae = j.last_api_error;
        const elP = document.getElementById('ops-poll'); if(elP) elP.textContent = _fmtOps(ps);
        const elD = document.getElementById('ops-discover'); if(elD) elD.textContent = _fmtOps(ds);
        const elA = document.getElementById('ops-apierr');
        if(elA){
          if(ae && ae.at){
            elA.textContent = (ae.at.replace('T',' ').slice(0,19)) + ' • ' + (ae.method||'') + ' ' + (ae.path||'') + (ae.status?(' ('+ae.status+')'):'') ;
          }else{
            elA.textContent = '-';
          }
        }

        // Série temporal
        refreshSeries(doc);
        // Painel de alertas
        refreshAlerts(doc);

        // Tabela por doc (sempre global)
        const rows = document.getElementById("dash-doc-rows");
        if(rows){
          const summary = Array.isArray(j.docs_summary) ? j.docs_summary : [];
          if(summary.length === 0){
            rows.innerHTML = '<tr><td colspan="3" class="muted">Sem dados ainda. Adicione um CPF/CNPJ no Admin e descubra processos.</td></tr>';
          }else{
            rows.innerHTML = summary.map(x => `
              <tr>
                <td><code class="small">${esc(x.doc||"")}</code></td>
                <td class="text-end">${esc(String(x.processos ?? 0))}</td>
                <td class="text-end"><span class="badge rounded-pill ${Number(x.alertas||0)>0?'text-bg-danger':'text-bg-secondary'}">${esc(String(x.alertas ?? 0))}</span></td>
              </tr>
            `).join("");
          }
        }

      }catch(e){
        showInlineBanner("bn-dash","danger","Erro de rede ao carregar métricas", esc(e.toString()));
      }
    }

    const btnR = document.getElementById("btn-dash-refresh");
    if(btnR){ btnR.addEventListener("click", ()=>refreshDashboard(true)); }

    const daysSel = document.getElementById('dash-days');
    if(daysSel){ daysSel.addEventListener('change', ()=>{
      const sel = document.getElementById('dash-scope');
      const doc = (sel && sel.value && sel.value !== '__global__') ? sel.value : '';
      refreshSeries(doc);
    }); }


// -----------------------------
// Painel de alertas (auto-refresh, filtros, export CSV)
// -----------------------------
let alertsTimer = null;

function _alertsRead(){
  try{
    const raw = localStorage.getItem("dash_alerts_cfg");
    return raw ? JSON.parse(raw) : {};
  }catch(e){ return {}; }
}
function _alertsWrite(cfg){
  try{ localStorage.setItem("dash_alerts_cfg", JSON.stringify(cfg||{})); }catch(e){}
}

function _alertsCfgFromUI(){
  const intervalEl = document.getElementById("alerts-interval");
  const typesEl = document.getElementById("alerts-types");
  const qEl = document.getElementById("alerts-q");
  const vpEl = document.getElementById("alerts-must-vp");
  return {
    interval: intervalEl ? Number(intervalEl.value||0) : 0,
    types: typesEl ? String(typesEl.value||"").trim() : "",
    q: qEl ? String(qEl.value||"").trim() : "",
    must_vp: vpEl ? !!vpEl.checked : false,
  };
}

function _alertsApplyCfgToUI(cfg){
  const intervalEl = document.getElementById("alerts-interval");
  const typesEl = document.getElementById("alerts-types");
  const qEl = document.getElementById("alerts-q");
  const vpEl = document.getElementById("alerts-must-vp");
  if(intervalEl && cfg.interval != null) intervalEl.value = String(cfg.interval);
  if(typesEl && cfg.types != null) typesEl.value = String(cfg.types);
  if(qEl && cfg.q != null) qEl.value = String(cfg.q);
  if(vpEl && cfg.must_vp != null) vpEl.checked = !!cfg.must_vp;
}

function _alertsBuildUrl(doc){
  const cfg = _alertsCfgFromUI();
  const params = new URLSearchParams();
  params.set("limit","50");
  if(doc) params.set("doc", doc);
  if(cfg.types) params.set("types", cfg.types);
  if(cfg.q) params.set("q", cfg.q);
  if(cfg.must_vp) params.set("must_value_penhora","1");
  return "/ui/api/alerts?" + params.toString();
}

function _alertsBuildExportHref(doc){
  const cfg = _alertsCfgFromUI();
  const params = new URLSearchParams();
  params.set("limit","500");
  if(doc) params.set("doc", doc);
  if(cfg.types) params.set("types", cfg.types);
  if(cfg.q) params.set("q", cfg.q);
  if(cfg.must_vp) params.set("must_value_penhora","1");
  return "/ui/api/alerts/export.csv?" + params.toString();
}

async function refreshAlerts(doc){
  try{
    clearInlineBanner("bn-alerts");
    const rows = document.getElementById("dash-alert-rows");
    const exportBtn = document.getElementById("btnAlertsExport");
    if(exportBtn) exportBtn.href = _alertsBuildExportHref(doc||"");
    if(rows) rows.innerHTML = '<tr><td colspan="6" class="muted">Carregando...</td></tr>';

    const url = _alertsBuildUrl(doc||"");
    const r = await fetch(url);
    const j = await r.json();
    if(!j.ok){
      showInlineBanner("bn-alerts","warning","Alertas indisponíveis", esc(j.error||j.message||"Falha"));
      if(rows) rows.innerHTML = '<tr><td colspan="6" class="muted">Sem dados.</td></tr>';
      return;
    }

    const items = Array.isArray(j.items) ? j.items : [];
    if(!items.length){
      if(rows) rows.innerHTML = '<tr><td colspan="6" class="muted">Sem alertas pendentes (com esses filtros).</td></tr>';
      return;
    }

    function _when(it){
      const d = it.data || it.created_at || "";
      return d ? esc(String(d).replace('T',' ').slice(0,19)) : "-";
    }
    function _tipo(it){
      const t = (it.tipo_inferido || it.tipo || "-");
      return esc(String(t));
    }
    function _short(s, n){
      const x = String(s||"");
      return x.length > n ? (x.slice(0,n-1) + "…") : x;
    }

    if(rows){
      rows.innerHTML = items.map(it => `
        <tr>
          <td class="mono small">${esc(String(it.event_id||""))}</td>
          <td class="small">${_when(it)}</td>
          <td class="mono small"><code>${esc(String(it.doc||""))}</code></td>
          <td class="mono small"><a href="/ui/processo/${encodeURIComponent(String(it.cnj||""))}">${esc(String(it.cnj||""))}</a></td>
          <td class="small"><span class="badge text-bg-light">${_tipo(it)}</span></td>
          <td class="small wrap" title="${esc(String(it.texto||""))}">${esc(_short(it.texto, 220))}</td>
        </tr>
      `).join("");
    }
  }catch(e){
    showInlineBanner("bn-alerts","warning","Falha ao carregar alertas", esc(e.message||String(e)));
  }
}

function setupAlertsAutoRefresh(doc){
  const cfg = _alertsCfgFromUI();
  _alertsWrite(cfg);

  if(alertsTimer){
    clearInterval(alertsTimer);
    alertsTimer = null;
  }
  const sec = Number(cfg.interval||0);
  if(sec > 0){
    alertsTimer = setInterval(()=>refreshAlerts(doc||""), Math.max(1, sec) * 1000);
  }
}

// Inicializa filtros do painel de alertas
(function initAlertsPanel(){
  const sel = document.getElementById("dash-scope");
  const currentDoc = (sel && sel.value && sel.value !== "__global__") ? sel.value : "";
  const cfg = _alertsRead();
  if(cfg && Object.keys(cfg).length){
    _alertsApplyCfgToUI(cfg);
  }else{
    // defaults
    _alertsApplyCfgToUI({interval: 10, types: "SENTENCA,DECISAO", q: "", must_vp: false});
  }

  const btn = document.getElementById("btnAlertsRefresh");
  if(btn) btn.addEventListener("click", ()=>refreshAlerts(currentDoc));

  const intervalEl = document.getElementById("alerts-interval");
  const typesEl = document.getElementById("alerts-types");
  const qEl = document.getElementById("alerts-q");
  const vpEl = document.getElementById("alerts-must-vp");
  const onChange = ()=>{
    const sel2 = document.getElementById("dash-scope");
    const doc2 = (sel2 && sel2.value && sel2.value !== "__global__") ? sel2.value : "";
    refreshAlerts(doc2);
    setupAlertsAutoRefresh(doc2);
  };
  if(intervalEl) intervalEl.addEventListener("change", onChange);
  if(typesEl) typesEl.addEventListener("change", onChange);
  if(qEl) qEl.addEventListener("change", onChange);
  if(vpEl) vpEl.addEventListener("change", onChange);
})();

refreshDashboard();
setInterval(()=>refreshDashboard(false), 5000);

</script>

<div class="row g-3">
      <div class="col-lg-5">
        <div class="card p-4">
          <h5 class="mb-1">Adicionar CPF/CNPJ</h5>
          <div class="muted mb-3">Cria monitoramento de <b>novos processos</b> no Escavador e salva localmente.</div>

          <form id="frmAdd" class="row g-2">
            <div class="col-12">
              <input name="doc" class="form-control form-control-lg mono" placeholder="08.840.686/0001-30 ou 123.456.789-00" required>
            </div>
            <div class="col-12 d-grid">
              <button class="btn btn-brand btn-lg">Adicionar</button>
            </div>
          </form>

          <div id="addResult" class="mt-3"></div>

          <hr class="divider">

          <div class="muted small">
            Para abrir um processo direto: <span class="mono">/ui/processo/SEU_CNJ</span>
          </div>
        </div>
      </div>

      <div class="col-lg-7">
        <div class="d-flex justify-content-between align-items-center mb-2">
          <div>
            <h5 class="mb-0">Watchlist</h5>
            <div class="muted small">Adicionar, ver, remover</div>
          </div>
        </div>

        <div class="row g-3">""" + cards + """</div>
      </div>
    </div>

<script>
function esc(s){ return (s||"").toString().replaceAll("&","&amp;").replaceAll("<","&lt;").replaceAll(">","&gt;"); }

document.getElementById("frmAdd").addEventListener("submit", async function(ev){
  ev.preventDefault();
  var doc = ev.target.doc.value.trim();
  var box = document.getElementById("addResult");
  box.innerHTML = '<div class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Enviando…</div>';
  try{
    var r = await fetch("/watchlist", { method:"POST", headers:{ "Content-Type":"application/json" }, body: JSON.stringify({ doc: doc }) });
    var j = await r.json();
    if(j.ok){
      box.innerHTML = '<div class="alert alert-success mb-0"><b>OK!</b> Doc cadastrado. Recarregando…</div>';
      setTimeout(function(){ location.reload(); }, 800);
    } else {
      box.innerHTML = '<div class="alert alert-warning mb-0"><b>Falhou:</b> ' + esc(j.message || j.error || "erro") + '</div>';
    }
  } catch(e){
    var _e = esc(e.toString());
    showInlineBanner("bn-movs","danger","Erro ao carregar movimentações", _e);
    box.innerHTML = '<div class="muted">Sem movimentações para exibir.</div>';
  }
});

async function deleteDoc(id, doc){
  if(!confirm("Remover da watchlist?\\n" + doc)) return;
  var r = await fetch("/watchlist/" + id, { method:"DELETE" });
  var j = await r.json();
  if(j.ok) location.reload();
  else showToast("danger","Falhou",(j.error || j.message || "erro"));
}
</script>
    """
    return render_template_string(UI_BASE, body=body)


# ---------------- UI: processo "produto" (Sidebar + Abas) ----------------
@app.get("/ui/processo/<path:cnj>")
def ui_processo(cnj: str):
    if not CNJ_REGEX.fullmatch(cnj):
        return render_template_string(UI_BASE, body=ui_alert("danger", "CNJ inválido", "Use o formato 0000000-00.0000.0.00.0000."))

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT doc, tipo_doc FROM watchlist ORDER BY id DESC LIMIT 500")
    wl = [dict(r) for r in cur.fetchall()]
    conn.close()

    opts = "<option value=''>Selecionar doc</option>"
    for it in wl:
        opts += "<option value='" + it["doc"] + "'>" + it["tipo_doc"] + ": " + it["doc"] + "</option>"

    cnj_json = json.dumps(cnj)

    body = """
    <div class="d-flex justify-content-between align-items-center mb-3">
      <div>
        <div class="muted small">Manutenção do processo</div>
        <div class="h4 mb-0 mono">__CNJ__</div>
      </div>
      <div class="d-flex gap-2 flex-wrap justify-content-end">
        <a class="btn btn-outline-primary" href="/ui">Dashboard</a>
        <button class="btn btn-brand" id="btn-sync">Sync agora</button>
      </div>
    </div>

    <div class="row g-3">
      <!-- Sidebar -->
      <div class="col-lg-4 col-xl-3">
        <div class="sticky-side">
          <div class="side-card p-3 mb-3" id="side-summary">
            <div id="bn-summary" class="inline-banner"></div>
            <div id="side-summary-content" class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando resumo…</div>
          </div>

          <div class="side-card p-3 mb-3">
            <div class="side-title mb-1">Filtros rápidos</div>
            <div class="muted small mb-2">Clique em um tipo para filtrar a timeline.</div>

            <div class="d-flex flex-wrap gap-2" id="quick-chips">
              <span class="chip chip-click" data-t="SENTENCA">⚖️ SENTENCA</span>
              <span class="chip chip-click" data-t="DECISAO">🧑‍⚖️ DECISAO</span>
              <span class="chip chip-click" data-t="INTIMACAO">📣 INTIMACAO</span>
              <span class="chip chip-click" data-t="CITACAO">📨 CITACAO</span>
              <span class="chip chip-click" data-t="AUDIENCIA">🗓️ AUDIENCIA</span>
              <span class="chip chip-click" data-t="DESPACHO">📝 DESPACHO</span>
              <span class="chip chip-click" data-t="JUNTADA">📎 JUNTADA</span>
              <span class="chip chip-click" data-t="DISTRIBUICAO">🧭 DISTRIBUICAO</span>
              <span class="chip chip-click" data-t="TRANSITO">🏁 TRANSITO</span>
              <span class="chip chip-click" data-t="">🧽 LIMPAR</span>
            </div>

            <hr class="divider">

            <div class="muted small mb-1">Busca livre:</div>
            <input class="form-control" id="q_side" placeholder="Ex: liminar, valor, petição">
            <div class="d-grid mt-2">
              <button class="btn btn-outline-primary" id="btn-side-search">Buscar</button>
            </div>
          </div>

          <div class="side-card p-3">
            <div class="side-title mb-1">Vincular doc (CRUD)</div>
            <div class="muted small">Vincula este CNJ a CPF/CNPJ para agregação e navegação.</div>

            <div class="mt-2">
              <label class="form-label muted small">Escolher da watchlist</label>
              <select class="form-select mono" id="docSelect">__OPTS__</select>
            </div>

            <div class="mt-2">
              <label class="form-label muted small">Ou digitar</label>
              <input class="form-control mono" id="docManual" placeholder="08.840.686/0001-30 ou 123.456.789-00">
            </div>

            <div class="d-grid mt-2">
              <button class="btn btn-brand" id="btnLinkDoc">Vincular</button>
            </div>

            <div class="mt-2">
              <span class="muted small" id="link-status"></span>
            </div>

            <hr class="divider">

            <div class="muted small mb-1">Docs vinculados:</div>
            <div id="bn-linked" class="inline-banner"></div>
            <div id="linkedDocs" class="muted small">Carregando…</div>
          </div>
        </div>
      </div>

      <!-- Main -->
      <div class="col-lg-8 col-xl-9">
        <ul class="nav nav-tabs mb-3" role="tablist">
          <li class="nav-item" role="presentation">
            <button class="nav-link active" data-bs-toggle="tab" data-bs-target="#pane-capa" type="button" role="tab">Capa</button>
          </li>
          <li class="nav-item" role="presentation">
            <button class="nav-link" data-bs-toggle="tab" data-bs-target="#pane-partes" type="button" role="tab">Partes</button>
          </li>
          <li class="nav-item" role="presentation">
            <button class="nav-link" data-bs-toggle="tab" data-bs-target="#pane-pedmult" type="button" role="tab">Pedidos &amp; Multas</button>
          </li>
          <li class="nav-item" role="presentation">
            <button class="nav-link" data-bs-toggle="tab" data-bs-target="#pane-docs" type="button" role="tab">Documentos</button>
          </li>
          <li class="nav-item" role="presentation">
            <button class="nav-link" data-bs-toggle="tab" data-bs-target="#pane-movs" type="button" role="tab">Movimentações</button>
          </li>
        </ul>

        <div class="tab-content">
          <div class="tab-pane fade show active" id="pane-capa" role="tabpanel">
            <div id="bn-capa" class="inline-banner mb-2"></div>
            <div class="card p-3" id="capa-box"><span class="muted">Carregando capa…</span></div>
          </div>

          <div class="tab-pane fade" id="pane-partes" role="tabpanel">
            <div id="bn-partes" class="inline-banner mb-2"></div>
            <div class="card p-3" id="partes-box"><span class="muted">Carregando partes…</span></div>
          </div>

          <div class="tab-pane fade" id="pane-pedmult" role="tabpanel">
            <div id="bn-pedmult" class="inline-banner mb-2"></div>

            <div class="d-flex justify-content-between align-items-center flex-wrap gap-2 mb-3">
              <div class="muted small">Extraído da capa, movimentações e documentos (quando disponíveis) e apresentado de forma consolidada.</div>
              <button class="btn btn-outline-primary btn-mini" id="btn-pedmult-refresh">Recarregar</button>
            </div>

            <ul class="nav nav-pills gap-2 mb-3" id="pedmult-subtabs" role="tablist">
              <li class="nav-item" role="presentation">
                <button class="nav-link active" id="pm-tab-valores" data-bs-toggle="pill" data-bs-target="#pm-pane-valores" type="button" role="tab">Valores &amp; Custas</button>
              </li>
              <li class="nav-item" role="presentation">
                <button class="nav-link" id="pm-tab-pedidos" data-bs-toggle="pill" data-bs-target="#pm-pane-pedidos" type="button" role="tab">Pedidos</button>
              </li>
              <li class="nav-item" role="presentation">
                <button class="nav-link" id="pm-tab-multas" data-bs-toggle="pill" data-bs-target="#pm-pane-multas" type="button" role="tab">Multas</button>
              </li>
            </ul>

            <div class="tab-content">
              <div class="tab-pane fade show active" id="pm-pane-valores" role="tabpanel">
                <div class="card p-3" id="valores-box" style="min-height:420px;"><span class="muted">Carregando valores…</span></div>
              </div>
              <div class="tab-pane fade" id="pm-pane-pedidos" role="tabpanel">
                <div class="card p-3" id="pedidos-box" style="min-height:420px;"><span class="muted">Carregando pedidos…</span></div>
              </div>
              <div class="tab-pane fade" id="pm-pane-multas" role="tabpanel">
                <div class="card p-3" id="multas-box" style="min-height:420px;"><span class="muted">Carregando multas…</span></div>
              </div>
            </div>
          </div>
          <div class="tab-pane fade" id="pane-docs" role="tabpanel">
            <div id="bn-docs" class="inline-banner mb-2"></div>
            <div class="card p-3">
              <div class="d-flex justify-content-between align-items-center flex-wrap gap-2 mb-2">
                <div class="side-title mb-0">Documentos</div>
                <div class="d-flex gap-2 flex-wrap">
                  <div class="btn-group" role="group" aria-label="tipo docs">
                    <input type="radio" class="btn-check" name="docsTipo" id="docsTipoPub" autocomplete="off" checked>
                    <label class="btn btn-outline-secondary btn-sm" for="docsTipoPub">Públicos</label>

                    <input type="radio" class="btn-check" name="docsTipo" id="docsTipoAutos" autocomplete="off">
                    <label class="btn btn-outline-secondary btn-sm" for="docsTipoAutos">Autos</label>
                  </div>
                  <button class="btn btn-outline-primary btn-sm" id="btn-docs-refresh">Recarregar</button>
                  <button class="btn btn-brand btn-sm" id="btn-docs-update">Atualizar no tribunal</button>
                </div>
              </div>

              <div class="muted small mb-2">
                Público: baixa documentos públicos. Autos: exige autenticação/certificado na conta do Escavador.
              </div>

              <div id="docs-box"><span class="muted">Selecione e carregue…</span></div>
            </div>
          </div>


          <div class="tab-pane fade" id="pane-movs" role="tabpanel">
            <div id="bn-movs" class="inline-banner mb-2"></div>
            <div class="card p-3 mb-2">
              <div class="row g-2 align-items-end">
                <div class="col-md-6">
                  <label class="form-label muted small">Buscar texto</label>
                  <input class="form-control" id="q" placeholder="Ex: sentença, citação, audiência">
                </div>
                <div class="col-md-3">
                  <label class="form-label muted small">Tipo</label>
                  <input class="form-control mono" id="tipo" placeholder="Ex: INTIMACAO">
                </div>
                <div class="col-md-3 d-grid">
                  <button class="btn btn-outline-primary" id="btn-filtrar">Filtrar</button>
                </div>
              </div>
            </div>

            <div class="card p-3" id="movs-box"><span class="muted">Carregando timeline…</span></div>

            <div class="d-flex justify-content-between align-items-center mt-2">
              <button class="btn btn-outline-primary btn-mini" id="prev">Anterior</button>
              <span class="muted small" id="pageinfo"></span>
              <button class="btn btn-outline-primary btn-mini" id="next">Próxima</button>
            </div>
          </div>
        </div>
      </div>
    </div>

<script>
var CNJ = __CNJ_JSON__;
var limit = 25;
var offset = 0;
var q = "";
var tipo = "";
var capaData = null;

function esc(s){
  return (s||"").toString().replaceAll("&","&amp;").replaceAll("<","&lt;").replaceAll(">","&gt;");
}

// Converte valores vindos da API (que às vezes são objetos) em texto amigável
function safeMoney(v){
  if(v === null || v === undefined || v === "") return "";
  if(typeof v === "number"){
    try{ return new Intl.NumberFormat("pt-BR", {style:"currency", currency:"BRL"}).format(v); }catch(e){ return String(v); }
  }
  var s = safeText(v);
  if(!s) return "";
  var m = s.match(new RegExp("R\\$\\s*\\d{1,3}(?:\\.\\d{3})*,\\d{2}|\\d{1,3}(?:\\.\\d{3})*,\\d{2}"));
  return m ? m[0].trim() : s;
}

function safeText(v){
  if(v === null || v === undefined) return "";
  if(typeof v === "number") return String(v);
  if(typeof v === "string"){
    const s = v.trim();
    if((s.startsWith("{") && s.endsWith("}")) || (s.startsWith("[") && s.endsWith("]"))){
      try{ return safeText(JSON.parse(s)); }catch(e){}
    }
    return v;
  }
  if(typeof v === "boolean") return v ? "Sim" : "Não";
  if(Array.isArray(v)) return v.map(safeText).filter(Boolean).join(", ");
  if(typeof v === "object"){
    // casos comuns de objetos retornados pela API
    if(v.valor_formatado) return String(v.valor_formatado);
    if(v.valor !== undefined && (typeof v.valor === "number" || typeof v.valor === "string")) return String(v.valor);
    if(v.nome) return String(v.nome);
    if(v.descricao) return String(v.descricao);
    if(v.sigla) return String(v.sigla);

    // unidade_origem costuma vir com cidade/estado/tribunal_sigla
    var cidade = v.cidade || "";
    var estado = "";
    try{
      if(v.estado){
        if(typeof v.estado === "string") estado = v.estado;
        else if(v.estado.sigla) estado = v.estado.sigla;
        else if(v.estado.nome) estado = v.estado.nome;
      }
    }catch(e){}
    var trib = v.tribunal_sigla || v.tribunal || "";
    var parts = [];
    var a = [cidade, estado].filter(Boolean).join(" - ");
    if(a) parts.push(a);
    if(trib) parts.push(String(trib));
    if(parts.length) return parts.join(" • ");

    // fallback: JSON curto para não estourar layout
    try{
      var s = JSON.stringify(v);
      if(s.length > 80) s = s.slice(0, 77) + "...";
      return s;
    }catch(e){
      return "[obj]";
    }
  }
  return String(v);
}

function formatDateBR(dt){
  if(!dt) return "";
  try{
    var d = new Date(dt);
    if(isNaN(d.getTime())) return String(dt);
    // formato compacto pt-BR sem vírgula
    var s = d.toLocaleString("pt-BR", { year:"numeric", month:"2-digit", day:"2-digit", hour:"2-digit", minute:"2-digit", second:"2-digit" });
    return s.replace(",", "");
  } catch(e){
    return String(dt);
  }
}

function formatMoneyBR(v){
  if(v === null || v === undefined || v === "") return "";
  if(typeof v === "object"){
    // tenta extrair de estruturas comuns
    if(v.valor_formatado) return safeText(v.valor_formatado);
    if(v.valor !== undefined) v = v.valor;
  }
  var s = safeText(v).trim();
  // já vem formatado?
  if(/^R[$]/.test(s) || /^[0-9]{1,3}([.][0-9]{3})*,[0-9]{2}$/.test(s)) return s;
  // tenta número
  var n = Number(String(s).replace(/[.]/g, "").replace(",", "."));
  if(!isFinite(n)) return s;
  try{
    return n.toLocaleString("pt-BR", { style:"currency", currency:"BRL" });
  }catch(e){
    return "R$ " + n.toFixed(2);
  }
}
function chip(label, value){
  var txt = safeText(value);
  if(!txt) return "";
  return '<span class="chip"><b>' + esc(label) + ':</b> <span class="chip-val truncate">' + esc(txt) + '</span></span>';
}
function kpi(label, value, action){
  var txt = safeText(value);
  var act = action ? String(action) : "";
  var cls = act ? "kpi kpi-click" : "kpi";
  var attrs = act ? (' role="button" tabindex="0" data-action="' + esc(act) + '"') : "";
  var wrap = (txt && txt.length > 12) ? ' wrap' : '';
  return '<div class="col-6 col-lg-3"><div class="' + cls + '"' + attrs + '><div class="label">' + esc(label) + '</div><div class="value' + wrap + '">' + esc(txt || "") + '</div></div></div>';
}
function kpiSummary(label, value, action){
  var txt = safeText(value);
  var act = action ? String(action) : "";
  var cls = act ? "kpi kpi-click" : "kpi";
  var attrs = act ? (' role="button" tabindex="0" data-action="' + esc(act) + '"') : "";
  var wrap = (txt && txt.length > 12) ? ' wrap' : '';
  return '<div class="col-6"><div class="' + cls + '"' + attrs + '><div class="label">' + esc(label) + '</div><div class="value' + wrap + '">' + esc(txt || "") + '</div></div></div>';
}
function activateTab(target){
  try{
    var btn = document.querySelector('button[data-bs-target="' + target + '"]');
    if(btn) btn.click();
  }catch(e){}
}
function scrollToEl(id){
  try{
    var el = document.getElementById(id);
    if(el) el.scrollIntoView({ behavior:"smooth", block:"start" });
  }catch(e){}
}
function bindKpiClicks(summary, fontes){
  var host = document.getElementById("side-summary-content");
  if(!host) return;
  host.querySelectorAll(".kpi[data-action]").forEach(function(el){
    if(el.__bound) return;
    el.__bound = true;
    function run(){
      var a = el.getAttribute("data-action") || "";
      if(a === "movs"){
        activateTab("#pane-movs");
        setTimeout(function(){ scrollToEl("movs-box"); }, 60);
      } else if(a === "ultima"){
        activateTab("#pane-movs");
        setTimeout(function(){ scrollToEl("movs-box"); }, 60);
      } else if(a === "fontes"){
        activateTab("#pane-capa");
        setTimeout(function(){ scrollToEl("capa-box"); }, 60);
        try{ showToast("info", "Fontes", (fontes && fontes.length ? (fontes.length + " fonte(s) na capa") : "Sem fontes") , 2200); }catch(e){}
      } else if(a === "verificado"){
        try{
          var v = summary && summary.data_ultima_verificacao ? formatDateBR(summary.data_ultima_verificacao) : "";
          showToast("info", "Verificado", v ? ("Última verificação: " + v) : "Sem data de verificação", 3000);
        }catch(e){}
      }
    }
    el.addEventListener("click", run);
    el.addEventListener("keydown", function(ev){ if(ev.key==="Enter"||ev.key===" "){ ev.preventDefault(); run(); }});
  });
}

function readComplement(capa, tipo){
  var arr = (capa && capa.informacoes_complementares) ? capa.informacoes_complementares : [];
  var hit = arr.find(function(x){ return (x.tipo||"").toString().toLowerCase() === tipo.toLowerCase(); });
  return hit ? hit.valor : null;
}
function safeDateStr(s){
  if(!s) return "";
  var m = ("" + s).match(/^(\\d{4}-\\d{2}-\\d{2})/);
  return m ? m[1] : "";
}
function todayStrLocal(){
  var d = new Date();
  var y = d.getFullYear();
  var m = String(d.getMonth()+1).padStart(2,"0");
  var dd = String(d.getDate()).padStart(2,"0");
  return y + "-" + m + "-" + dd;
}

function iconForType(t){
  var x = (t||"").toUpperCase();
  if(x.indexOf("SENTEN")>=0) return "⚖️";
  if(x.indexOf("DECIS")>=0) return "🧑‍⚖️";
  if(x.indexOf("INTIM")>=0) return "📣";
  if(x.indexOf("CIT")>=0) return "📨";
  if(x.indexOf("AUDI")>=0) return "🗓️";
  if(x.indexOf("DESP")>=0) return "📝";
  if(x.indexOf("JUNT")>=0) return "📎";
  if(x.indexOf("DISTR")>=0) return "🧭";
  if(x.indexOf("TRANS")>=0) return "🏁";
  return "📄";
}

function renderPessoa(p){
  var nome = p.nome || "";
  var doc = p.cnpj || p.cpf || "";
  var tipoN = p.tipo_normalizado || p.tipo || "";
  var advs = Array.isArray(p.advogados) ? p.advogados : [];
  var advHtml = "";
  if(advs.length > 0){
    advHtml = "<div class='mt-2 muted small'><b>Advogados:</b><ul class='mb-0'>";
    for(var i=0;i<advs.length;i++){
      var a = advs[i];
      var oabs = Array.isArray(a.oabs) ? a.oabs : [];
      var oabStr = oabs.map(function(o){ return (o.uf||"") + (o.numero||""); }).join(", ");
      advHtml += "<li>" + esc(a.nome || "") + " <span class='mono'>(" + esc(oabStr) + ")</span></li>";
    }
    advHtml += "</ul></div>";
  }
  return (
    "<div class='p-2' style='border:1px solid var(--border); border-radius:14px; background:rgba(48,80,160,.03);'>" +
      "<div class='d-flex justify-content-between align-items-start gap-2'>" +
        "<div>" +
          "<div style='font-weight:750;'>" + esc(nome) + "</div>" +
          "<div class='muted small'>" + esc(tipoN) + (doc ? " • " + esc(doc) : "") + "</div>" +
        "</div>" +
        "<span class='badge-soft p-2'>" + esc((p.polo||"").toString().toUpperCase()||"") + "</span>" +
      "</div>" +
      advHtml +
    "</div>"
  );
}

function renderEnvolvidos(envolvidos){
  var arr = Array.isArray(envolvidos) ? envolvidos : [];
  if(arr.length===0) return "<div class='muted'>Sem envolvidos nesta fonte.</div>";

  var ativos = [], passivos = [], outros = [];
  for(var i=0;i<arr.length;i++){
    var p = arr[i];
    var polo = (p.polo||"").toString().toUpperCase();
    if(polo==="ATIVO") ativos.push(p);
    else if(polo==="PASSIVO") passivos.push(p);
    else outros.push(p);
  }

  function section(title, items){
    if(!items || items.length===0) return "";
    var html = "<div class='mt-3'><div class='h6 mb-2'>" + esc(title) + "</div><div class='d-grid gap-2'>";
    for(var j=0;j<items.length;j++) html += renderPessoa(items[j]);
    html += "</div></div>";
    return html;
  }

  return section("Polo Ativo", ativos) + section("Polo Passivo", passivos) + section("Outros", outros);
}

function aggregatePartes(fontes){
  var map = {};
  for(var i=0;i<fontes.length;i++){
    var f = fontes[i] || {};
    var env = Array.isArray(f.envolvidos) ? f.envolvidos : [];
    for(var j=0;j<env.length;j++){
      var p = env[j] || {};
      var key = (p.nome||"") + "|" + (p.cpf||p.cnpj||"") + "|" + (p.tipo_normalizado||p.tipo||"") + "|" + (p.polo||"");
      if(!map[key]) map[key] = p;
    }
  }
  return Object.keys(map).map(function(k){ return map[k]; });
}

async function loadCapaPalatavel(){
  clearInlineBanner("bn-summary");
  clearInlineBanner("bn-capa");
  clearInlineBanner("bn-partes");

  var side = document.getElementById("side-summary-content");
  var capaBox = document.getElementById("capa-box");
  var partesBox = document.getElementById("partes-box");

  side.innerHTML = '<div class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando resumo…</div>';
  capaBox.innerHTML = '<div class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando…</div>';
  partesBox.innerHTML = '<div class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando…</div>';

  var r = null;
  var j = null;
  try{
    r = await fetch("/ui/api/processo/" + encodeURIComponent(CNJ) + "/capa");
    try { j = await r.json(); } catch(_e){ j = { ok:false, error:"RESPOSTA_INVALIDA", message:"Resposta não-JSON do servidor." }; }
  } catch(e){
    showInlineBanner("bn-capa","danger","Falha ao consultar capa", esc(e.toString()));
    showInlineBanner("bn-summary","danger","Falha ao carregar", esc(e.toString()));
    side.innerHTML = '<div class="muted">Sem resumo disponível.</div>';
    capaBox.innerHTML = '<div class="muted">Sem capa disponível.</div>';
    partesBox.innerHTML = '<div class="muted">Sem partes disponíveis.</div>';
    return;
  }

  if(!j.ok){
    var msg = esc(j.message || j.error || ("HTTP " + (r ? r.status : "")) || "erro");
    showInlineBanner("bn-summary","warning","Capa indisponível", msg);
    showInlineBanner("bn-capa","warning","Capa indisponível", msg);
    showInlineBanner("bn-partes","warning","Partes indisponíveis", msg);
    side.innerHTML = '<div class="muted">Sem resumo disponível.</div>';
    capaBox.innerHTML = '<div class="muted">Sem capa disponível.</div>';
    partesBox.innerHTML = '<div class="muted">Sem partes disponíveis.</div>';
    return;
  }

  capaData = j.data || {};
  // A API pode devolver {data:{...}} ou direto {...}. Vamos aceitar ambos.
  var d = (capaData && (capaData.data || capaData)) || {};
  // 'fontes' pode estar em d.fontes ou capaData.fontes dependendo do formato.
  var fontes = [];
  if(Array.isArray(d.fontes)) fontes = d.fontes;
  else if(Array.isArray(capaData.fontes)) fontes = capaData.fontes;

  // Sidebar summary
  var resumoHtml = ""
    + "<div class='summary-head'>"
    + "  <div class='summary-main'>"
    + "    <div class='muted small'>Resumo</div>"
    + "    <div class='side-title mono'>" + esc(d.numero_cnj || CNJ) + "</div>"
    + "  </div>"
    + "  <div id='badge-hoje'></div>"
    + "</div>"
    + "<div class='d-flex flex-wrap gap-2 mt-2'>"
    + chip("Ativo", d.titulo_polo_ativo)
    + chip("Passivo", d.titulo_polo_passivo)
    + chip("Unidade", d.unidade_origem)
    + chip("UF", d.estado_origem)
    + "</div>"
    + "<hr class='divider'>"
    + "<div class='row g-2 summary-metrics'>"
    + kpiSummary("Movimentações", d.quantidade_movimentacoes || "", "movs")
    + kpiSummary("Última movimentação", formatDateBR(d.data_ultima_movimentacao || ""), "ultima")
    + kpiSummary("Verificação", formatDateBR(d.data_ultima_verificacao || ""), "verificado")
    + kpiSummary("Fontes", fontes.length, "fontes")
    + "</div>"
    + "<div class='muted small mt-2'>Dica: use os chips de tipo para focar a timeline.</div>";

  side.innerHTML = resumoHtml;

bindKpiClicks(d, fontes);

  // Partes agregadas
  if(fontes.length>0){
    var all = aggregatePartes(fontes);
    partesBox.innerHTML = "<div class='muted small mb-2'>Agregado de todas as fontes (deduplicado)</div>" + renderEnvolvidos(all);
  } else {
    partesBox.innerHTML = "<div class='muted'>Sem fontes retornadas.</div>" + "<div class='mt-2 small muted'>Dica: abra o Network e veja a resposta de <code>/ui/api/processo/&lt;cnj&gt;/capa</code>.</div>";
  }

  // Abas por fonte na Capa
  if(fontes.length === 0){
    capaBox.innerHTML = "<div class='muted'>Sem fontes retornadas.</div>" + "<details class='mt-2'><summary class='small'>Ver payload (debug)</summary><pre class='small'>" + esc(JSON.stringify(capaData, null, 2)) + "</pre></details>";
    return;
  }

  var tabs = [];
  for(var i=0;i<fontes.length;i++){
    var f = fontes[i];
    var grau = f.grau_formatado || (f.grau ? ("" + f.grau + "º grau") : "");
    var tipoF = (f.tipo || "").toString().toUpperCase();
    var label = f.descricao || f.nome || grau || tipoF || "Fonte";
    if(label.toLowerCase().includes("diário de justiça")) label = "Diário (DJ)";
    if(grau) label = label + " • " + grau;
    tabs.push({ idx:i, id:"fonte-" + i, label:label });
  }

  var nav = "<ul class='nav nav-pills mb-3 gap-2' role='tablist'>";
  for(var t=0;t<tabs.length;t++){
    var it = tabs[t];
    nav += "<li class='nav-item' role='presentation'>"
        +  "<button class='nav-link " + (t===0 ? "active" : "") + "' data-bs-toggle='tab' data-bs-target='#" + it.id + "' type='button' role='tab'>"
        +    esc(it.label)
        +  "</button>"
        + "</li>";
  }
  nav += "</ul>";

  var panes = "<div class='tab-content'>";
  for(var p=0;p<tabs.length;p++){
    var tab = tabs[p];
    var ff = fontes[tab.idx] || {};
    var c = ff.capa || {};
    var isDiario = (ff.tipo || "").toString().toUpperCase().includes("DIARIO");
    var juiz = readComplement(c, "Juiz") || readComplement(c, "Relator") || "";
    var chips = "<div class='d-flex flex-wrap gap-2 mb-3'>"
      + chip("Tribunal", ff.sigla || (ff.tribunal && ff.tribunal.sigla) || "")
      + chip("Sistema", ff.sistema || "")
      + chip("Status", ff.status_predito || "")
      + chip("Segredo", ff.segredo_justica ? "Sim" : "Não")
      + chip("Físico", ff.fisico ? "Sim" : "Não")
      + chip("Qtd movs", ff.quantidade_movimentacoes || "")
      + "</div>";

    panes += "<div class='tab-pane fade " + (p===0 ? "show active" : "") + "' id='" + tab.id + "' role='tabpanel'>";

    if(isDiario){
      panes += chips
        + "<div class='row g-2'>"
        + kpi("Tipo", ff.tipo || "DIARIO")
        + kpi("Descrição", ff.descricao || "")
        + kpi("Última mov", formatDateBR(ff.data_ultima_movimentacao || ""))
        + kpi("Verificação", formatDateBR(ff.data_ultima_verificacao || ""))
        + "</div>";
      panes += "<details class='mt-3'><summary class='muted'>Ver JSON da fonte</summary><pre class='mono wrap mt-2 mb-0'>"
        + esc(JSON.stringify(ff, null, 2)) + "</pre></details>";
      panes += "</div>";
      continue;
    }

    panes += chips
      + "<div class='row g-2'>"
      + kpi("Classe", c.classe || "")
      + kpi("Assunto", c.assunto || "")
      + kpi("Área", c.area || "")
      + kpi("Órgão julgador", c.orgao_julgador || "")
      + kpi("Distribuição", formatDateBR(c.data_distribuicao || ""))
      + kpi("Situação", c.situacao || "")
      + kpi("Valor causa", formatMoneyBR(c.valor_causa || ""))
      + kpi("Juiz/Relator", juiz || "")
      + "</div>"
      + "<hr class='divider'>"
      + "<div class='h6 mb-2'>Envolvidos (desta fonte)</div>"
      + renderEnvolvidos(ff.envolvidos);

    panes += "<details class='mt-3'><summary class='muted'>Ver JSON da capa</summary><pre class='mono wrap mt-2 mb-0'>"
      + esc(JSON.stringify(c, null, 2)) + "</pre></details>";

    panes += "<details class='mt-3'><summary class='muted'>Ver JSON da fonte</summary><pre class='mono wrap mt-2 mb-0'>"
      + esc(JSON.stringify(ff, null, 2)) + "</pre></details>";

    panes += "</div>";
  }
  panes += "</div>";

  capaBox.innerHTML = nav + panes + "<details class='mt-3'><summary class='muted'>Ver JSON completo (processo)</summary><pre class='mono wrap mt-2 mb-0'>"
    + esc(JSON.stringify(capaData, null, 2)) + "</pre></details>";
}

function renderTimeline(items){
  if(!items || items.length===0) return "<div class='muted'>Sem movimentações neste recorte.</div>";

  var today = todayStrLocal();
  var anyToday = false;

  // agrupa por dia (YYYY-MM-DD)
  var groups = {};
  for(var i=0;i<items.length;i++){
    var it = items[i];
    var dt = safeDateStr(it.data || it.created_at || "");
    if(!dt) dt = "Sem data";
    if(!groups[dt]) groups[dt] = [];
    groups[dt].push(it);
    if(dt === today) anyToday = true;
  }

  // ordena dias: itens já vêm em DESC, mas garantimos
  var days = Object.keys(groups).sort(function(a,b){
    if(a==="Sem data") return 1;
    if(b==="Sem data") return -1;
    return (a<b) ? 1 : (a>b) ? -1 : 0;
  });

  var html = "";
  for(var d=0; d<days.length; d++){
    var day = days[d];
    var label = day;
    var badge = (day === today) ? "<span class='badge-today'>mudou hoje</span>" : "";
    html += "<div class='timeline-day'><div class='mono'><b>" + esc(label) + "</b></div>" + badge + "</div>";
    var arr = groups[day] || [];

    for(var j=0;j<arr.length;j++){
      var it = arr[j];
      var t = (it.tipo || it.tipo_inferido || "SEM_TIPO").toString().toUpperCase();
      var icon = iconForType(t);
      var meta = (it.data || it.created_at || "");
      var texto = it.texto || "";

      var newBadge = (safeDateStr(meta) === today) ? "<span class='badge-new'>novo</span>" : "";

      html += ""
        + "<div class='tl-item mt-2'>"
        + "  <div class='tl-top'>"
        + "    <div class='d-flex gap-2'>"
        + "      <div class='tl-icon'>" + esc(icon) + "</div>"
        + "      <div>"
        + "        <div class='tl-title'>" + esc(t) + " " + newBadge + "</div>"
        + "        <div class='tl-meta mono'>" + esc(meta) + "</div>"
        + "      </div>"
        + "    </div>"
        + "  </div>"
        + "  <div class='wrap'>" + esc(texto) + "</div>"
        + "</div>";
    }
  }

  // Atualiza badge no sidebar (mudou hoje)
  var badgeBox = document.getElementById("badge-hoje");
  if(badgeBox){
    badgeBox.innerHTML = anyToday ? "<span class='badge-today'>mudou hoje</span>" : "<span class='badge-soft p-2'>sem mudanças hoje</span>";
  }
  return html;
}


async function loadDocs(){
  clearInlineBanner("bn-docs");
  var box = document.getElementById("docs-box");
  if(!box) return;
  var tipo = document.getElementById("docsTipoAutos") && document.getElementById("docsTipoAutos").checked ? "autos" : "publicos";
  box.innerHTML = '<div class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando documentos…</div>';

  var r=null, j=null;
  try{
    r = await fetch("/ui/api/processo/" + encodeURIComponent(CNJ) + "/documentos?tipo=" + encodeURIComponent(tipo));
    try { j = await r.json(); } catch(_e){ j = { ok:false, error:"RESPOSTA_INVALIDA", message:"Resposta não-JSON do servidor." }; }
  }catch(e){
    showInlineBanner("bn-docs","danger","Falha ao carregar documentos", esc(e.toString()));
    box.innerHTML = '<div class="muted">Sem documentos.</div>';
    return;
  }

  if(!j.ok){
    showInlineBanner("bn-docs","warning","Documentos indisponíveis", esc(j.message || j.error || ("HTTP " + (r ? r.status : ""))));
    box.innerHTML = '<div class="muted">Sem documentos.</div>';
    return;
  }

  var items = j.items || [];
  if(!Array.isArray(items) || items.length === 0){
    var hint = "Nenhum documento encontrado no cache.";
    if(j.warning) hint += " " + esc(j.warning);
    var extra = tipo==="autos"
      ? "Autos podem exigir login, senha ou certificado configurados no Escavador para o tribunal deste processo."
      : "Se a atualização já foi solicitada, o tribunal ou o Escavador ainda podem estar processando os documentos públicos.";
    box.innerHTML = "<div class='docs-note'>"
      + "<div class='fw-semibold mb-1'>Ainda sem documentos disponíveis</div>"
      + "<div class='small text-muted mb-2'>" + hint + "</div>"
      + "<ul class='small mb-0 ps-3'>"
      + "<li>Use <b>Atualizar no tribunal</b> e aguarde a conclusão do processamento.</li>"
      + "<li>Depois clique em <b>Recarregar</b> para consultar o cache novamente.</li>"
      + "<li>" + extra + "</li>"
      + "</ul>"
      + "</div>";
    return;
  }

  var html = "";
  html += "<div class='d-flex justify-content-between align-items-center mb-2'>";
  html += "  <div class='muted small'>Fonte: " + esc(j.source || "cache") + "</div>";
  html += "  <div class='muted small'>Itens: " + items.length + "</div>";
  html += "</div>";

  html += "<div class='list-group'>";
  for(var i=0;i<items.length;i++){
    var it = items[i] || {};
    var key = it.doc_key || it.key || "";
    var titulo = it.titulo || (it.meta && (it.meta.titulo || it.meta.nome)) || ("Documento " + (i+1));
    var data = it.data || (it.meta && (it.meta.data || it.meta.data_documento)) || "";
    var mime = it.mime || (it.meta && (it.meta.mime || it.meta.mime_type)) || "";
    var dl = "/ui/api/processo/" + encodeURIComponent(CNJ) + "/documentos/" + encodeURIComponent(key) + "/download";
    html += "<div class='list-group-item'>";
    html += "  <div class='d-flex justify-content-between align-items-start gap-2'>";
    html += "    <div>";
    html += "      <div class='fw-semibold'>" + esc(titulo) + "</div>";
    html += "      <div class='muted small mono'>" + esc(key) + (data ? (" • " + esc(data)) : "") + (mime ? (" • " + esc(mime)) : "") + "</div>";
    html += "    </div>";
    html += "    <div class='text-nowrap'>";
    if(key){
      html += "      <a class='btn btn-outline-primary btn-sm' href='" + dl + "' target='_blank'>Download</a>";
    }else{
      html += "      <span class='muted small'>Sem chave</span>";
    }
    html += "    </div>";
    html += "  </div>";
    html += "</div>";
  }
  html += "</div>";
  box.innerHTML = html;
}

var __docsPollTimer = null;
var __docsPollTries = 0;

async function pollDocsStatus(tipo){
  if(__docsPollTimer){ clearInterval(__docsPollTimer); __docsPollTimer = null; }
  __docsPollTries = 0;

  __docsPollTimer = setInterval(async function(){
    __docsPollTries += 1;
    if(__docsPollTries > 30){
      clearInterval(__docsPollTimer); __docsPollTimer = null;
      showInlineBanner("bn-docs","warning","Atualização em andamento", "Ainda não finalizou. Você pode recarregar mais tarde.");
      return;
    }
    try{
      var r = await fetch("/ui/api/processo/" + encodeURIComponent(CNJ) + "/status-atualizacao");
      var j = await r.json();
      if(j && j.ok){
        var st = j.status || {};
        var sv = (st.status || st.estado || st.situacao || "").toString().toUpperCase();
        if(sv === "SUCESSO"){
          clearInterval(__docsPollTimer); __docsPollTimer = null;
          showInlineBanner("bn-docs","success","Atualização concluída", "Documentos disponíveis. Recarregando…");
          await loadDocs();
        }else if(sv === "ERRO" || sv === "FALHA"){
          clearInterval(__docsPollTimer); __docsPollTimer = null;
          showInlineBanner("bn-docs","danger","Atualização falhou", esc(JSON.stringify(st).slice(0,400)));
        }else{
          showInlineBanner("bn-docs","info","Atualizando…", "Status: " + esc(sv || "PENDENTE"));
        }
      }
    }catch(_e){
      // ignore
    }
  }, 2000);
}

async function requestDocsUpdate(){
  clearInlineBanner("bn-docs");
  var tipo_ui = document.getElementById("docsTipoAutos") && document.getElementById("docsTipoAutos").checked ? "autos" : "publicos";
  var tipo = (tipo_ui === "autos") ? "autos" : "documentos_publicos";

  showInlineBanner("bn-docs","info","Solicitando atualização…","Enviando requisição para o tribunal via Escavador.");

  try{
    var r = await fetch("/ui/api/processo/" + encodeURIComponent(CNJ) + "/solicitar-atualizacao", {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify({ tipo: tipo })
    });
    var j = null;
    try{ j = await r.json(); }catch(_e){ j = { ok:false, error:"RESPOSTA_INVALIDA", message:"Resposta não-JSON do servidor." }; }
    if(!j.ok){
      showInlineBanner("bn-docs","danger","Falha ao solicitar", esc(j.message || j.error || ("HTTP " + r.status)));
      return;
    }
    if(j.already_running || j.pending){
      var pendingMsg = j.already_running
        ? "O Escavador informou que a atualização já estava em andamento. Vou acompanhar o status."
        : "Solicitação aceita. Aguardando processamento…";
      if(tipo_ui === "autos"){
        pendingMsg += " Autos podem exigir credenciais ou certificado do tribunal.";
      }
      showInlineBanner("bn-docs","info", j.already_running ? "Atualização já em andamento" : "Atualização solicitada", esc(pendingMsg));
      await pollDocsStatus(tipo_ui);
      return;
    }
    showInlineBanner("bn-docs","info","Atualização solicitada","Aguardando processamento…");
    await pollDocsStatus(tipo_ui);
  }catch(e){
    showInlineBanner("bn-docs","danger","Falha ao solicitar", esc(e.toString()));
  }
}


async function loadMovs(){
  clearInlineBanner("bn-movs");
  var box = document.getElementById("movs-box");
  box.innerHTML = '<span class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando…</span>';

  try{
  var url = "";
  if(q || tipo){
    url = "/processos/" + encodeURIComponent(CNJ) + "/movimentacoes/busca?q=" + encodeURIComponent(q) + "&tipo=" + encodeURIComponent(tipo) + "&limit=" + limit + "&offset=" + offset;
  } else {
    url = "/processos/" + encodeURIComponent(CNJ) + "/movimentacoes?limit=" + limit + "&offset=" + offset;
  }

  var r = await fetch(url);
  var j = await r.json();
  if(!j.ok){
    var _m = esc(j.message || j.error || "erro");
    showInlineBanner("bn-movs","warning","Falha ao carregar movimentações", _m);
    box.innerHTML = '<div class="muted">Sem movimentações para exibir.</div>';
    return;
  }

  var items = j.items || [];
  var total = (j.total === undefined || j.total === null) ? 0 : j.total;

  box.innerHTML = renderTimeline(items);

  var end = offset + items.length;
  document.getElementById("pageinfo").textContent = items.length ? ("Mostrando " + (offset+1) + "–" + end + " (total aprox: " + total + ")") : ("offset=" + offset);
  } catch(e){
    box.innerHTML = '<div class="alert alert-danger mb-0"><b>Erro:</b> ' + esc(e.toString()) + '</div>';
  }
}


// ----------------------------
// Pedidos & Multas (extraído da capa)
// ----------------------------
var pedmultLoadedOnce = false;

function objToPairs(o){
  if(!o || typeof o !== "object") return [];
  var keys = Object.keys(o);
  var preferred = ["descricao","pedido","multa","tipo","valor","valor_formatado","data","data_hora","dataHora","origem","tribunal","sistema","observacao","observacoes"];
  keys.sort(function(a,b){
    var ia = preferred.indexOf(a), ib = preferred.indexOf(b);
    if(ia !== -1 || ib !== -1){
      if(ia === -1) return 1;
      if(ib === -1) return -1;
      return ia - ib;
    }
    return a.localeCompare(b);
  });
  return keys.map(function(k){ return [k, o[k]]; });
}

function cardHeader(title, count){
  return "<div class='d-flex justify-content-between align-items-center mb-2'>"
    + "<div class='h6 mb-0'>" + esc(title) + "</div>"
    + "<span class='badge-soft p-2'>" + esc(String(count)) + "</span>"
    + "</div>";
}

function tagChip(text){
  var t = safeText(text);
  if(!t) return "";
  return "<span style='display:inline-block;padding:4px 8px;border:1px solid var(--border);border-radius:999px;background:rgba(48,80,160,.05);font-size:12px;'>" + esc(t) + "</span>";
}

function excelHeader(title, count, columns){
  var cols = Array.isArray(columns) ? columns : [];
  var html = cardHeader(title, count);
  html += "<div style='overflow:auto;border:1px solid var(--border);border-radius:16px;background:#fff;'>";
  html += "<table class='table table-sm align-middle mb-0' style='min-width:980px;'>";
  html += "<thead style='position:sticky;top:0;background:#f7f9fc;z-index:1;'><tr>";
  for(var i=0;i<cols.length;i++){
    var align = cols[i] === 'Valor' ? "right" : 'left';
    html += "<th style='white-space:nowrap;padding:12px 10px;border-bottom:1px solid var(--border);text-align:" + align + ";font-size:12px;color:#5b6b86;text-transform:uppercase;letter-spacing:.03em;'>" + esc(cols[i]) + "</th>";
  }
  html += "</tr></thead><tbody>";
  return html;
}

function excelFooter(){
  return "</tbody></table></div>";
}

function tdCell(value, opts){
  opts = opts || {};
  var align = opts.align || 'left';
  var nowrap = opts.nowrap ? ';white-space:nowrap;' : '';
  var mono = opts.mono ? 'mono' : '';
  return "<td class='" + mono + "' style='padding:10px;border-bottom:1px solid #eef2f7;text-align:" + align + nowrap + "vertical-align:top;'>" + (value || "<span class='muted'>-</span>") + "</td>";
}

function compactText(s, max){
  var t = safeText(s || '');
  max = max || 180;
  if(!t) return '';
  return t.length > max ? t.slice(0, max-1) + '…' : t;
}

function renderStandardCards(title, items){
  var arr = Array.isArray(items) ? items : [];
  if(arr.length === 0) return cardHeader(title, 0) + "<div class='muted'>Nada encontrado.</div>";
  var html = excelHeader(title, arr.length, ['Tipo', 'Descrição', 'Origem', 'Data', 'Trecho']);
  for(var i=0;i<arr.length;i++){
    var it = arr[i];
    var tipo = '';
    var descricao = '';
    var origem = '';
    var data = '';
    var trecho = '';
    if(typeof it === 'string' || typeof it === 'number' || typeof it === 'boolean'){
      descricao = String(it);
      trecho = descricao;
    } else if(it && typeof it === 'object'){
      tipo = safeText(it.tipo || it.pedido || it.multa || '');
      descricao = safeText(it.descricao || it.texto || '');
      origem = safeText(it.origem || '');
      data = safeText(it.data || '');
      trecho = safeText(it.trecho || it.contexto || it.texto || it.descricao || '');
      if(!tipo && !descricao && !trecho){
        trecho = JSON.stringify(it);
      }
    } else {
      descricao = 'Item inválido';
    }
    html += '<tr>'
      + tdCell(esc(tipo || title.slice(0,-1) || 'Item'), {nowrap:true})
      + tdCell("<div class='wrap'>" + esc(compactText(descricao, 140)) + "</div>")
      + tdCell(origem ? tagChip(origem) : '')
      + tdCell(esc(data), {nowrap:true})
      + tdCell("<div class='wrap small'>" + esc(compactText(trecho, 240)) + "</div>")
      + '</tr>';
  }
  html += excelFooter();
  return html;
}

function renderValueCards(title, items){
  var arr = Array.isArray(items) ? items : [];
  if(arr.length === 0) return cardHeader(title, 0) + "<div class='muted'>Nada encontrado.</div>";
  var html = excelHeader(title, arr.length, ['Tipo', 'Valor', 'Descrição', 'Origem', 'Data', 'Trecho']);
  for(var i=0;i<arr.length;i++){
    var it = arr[i] || {};
    if(typeof it !== 'object'){
      it = { tipo: 'Valor', valor: safeMoney(it), descricao: String(it), trecho: String(it) };
    }
    var tipo = safeText(it.tipo || 'Valor');
    var descricao = safeText(it.descricao || '');
    var valor = safeMoney(it.valor_formatado || it.valor || '');
    var origem = safeText(it.origem || '');
    var data = safeText(it.data || '');
    var trecho = safeText(it.trecho || it.contexto || it.descricao || '');
    html += '<tr>'
      + tdCell("<strong>" + esc(tipo) + "</strong>", {nowrap:true})
      + tdCell(valor ? esc(valor) : '', {align:'right', nowrap:true, mono:true})
      + tdCell("<div class='wrap'>" + esc(compactText(descricao, 140)) + "</div>")
      + tdCell(origem ? tagChip(origem) : '')
      + tdCell(esc(data), {nowrap:true})
      + tdCell("<div class='wrap small'>" + esc(compactText(trecho, 240)) + "</div>")
      + '</tr>';
  }
  html += excelFooter();
  return html;
}

function renderCards(title, items){
  return (title === 'Valores & Custas') ? renderValueCards(title, items) : renderStandardCards(title, items);
}

async function loadPedidosMultas(force){
  clearInlineBanner("bn-pedmult");
  var pedidosBox = document.getElementById("pedidos-box");
  var multasBox = document.getElementById("multas-box");
  var valoresBox = document.getElementById("valores-box");
  pedidosBox.innerHTML = '<span class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando…</span>';
  multasBox.innerHTML  = '<span class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando…</span>';
  valoresBox.innerHTML = '<span class="muted"><span class="spinner-border spinner-border-sm me-2"></span>Carregando…</span>';

  try{
    var url = "/ui/api/processo/" + encodeURIComponent(CNJ) + "/pedidos-multas";
    if(force) url += "?refresh=1";
    var r = await fetch(url);
    var raw = await r.text();
    var j = {};
    try{ j = JSON.parse(raw); } catch(parseErr){
      showInlineBanner("bn-pedmult","danger","Falha ao carregar","Resposta não-JSON do servidor.");
      pedidosBox.innerHTML = "<div class='muted'>Falha.</div>";
      multasBox.innerHTML = "<div class='muted'>Falha.</div>";
      valoresBox.innerHTML = "<div class='muted'>Falha.</div>";
      return;
    }
    if(!j.ok){
      var msg = esc(j.message || j.error || ("HTTP " + r.status));
      showInlineBanner("bn-pedmult","warning","Falha ao carregar", msg);
      pedidosBox.innerHTML = "<div class='muted'>Sem dados.</div>";
      multasBox.innerHTML = "<div class='muted'>Sem dados.</div>";
      valoresBox.innerHTML = "<div class='muted'>Sem dados.</div>";
      return;
    }
    pedidosBox.innerHTML = renderCards("Pedidos", j.pedidos || []);
    multasBox.innerHTML = renderCards("Multas", j.multas || []);
    valoresBox.innerHTML = renderCards("Valores & Custas", j.valores || []);
    if((j.pedidos||[]).length === 0 && (j.multas||[]).length === 0 && (j.valores||[]).length === 0){
      showInlineBanner("bn-pedmult","info","Sem registros","Não localizei pedidos, multas ou valores relevantes na capa, nas movimentações ou nos documentos já disponíveis para este CNJ.");
    } else if(j.cached){
      try{ showToast("info","Pedidos & Multas", j.stale ? "Cache (stale) usado" : "Cache usado", 2200); }catch(e){}
    }
  } catch(e){
    showInlineBanner("bn-pedmult","danger","Erro ao carregar", esc(e.toString()));
    pedidosBox.innerHTML = "<div class='muted'>Falha.</div>";
    multasBox.innerHTML = "<div class='muted'>Falha.</div>";
    valoresBox.innerHTML = "<div class='muted'>Falha.</div>";
  }
}
document.getElementById("prev").addEventListener("click", function(){
  offset = Math.max(0, offset - limit);
  loadMovs();
});
document.getElementById("next").addEventListener("click", function(){
  offset = offset + limit;
  loadMovs();
});
document.getElementById("btn-filtrar").addEventListener("click", function(){
  q = document.getElementById("q").value.trim();
  tipo = document.getElementById("tipo").value.trim().toUpperCase();
  offset = 0;
  loadMovs();
});
document.getElementById("btn-side-search").addEventListener("click", function(){
  q = document.getElementById("q_side").value.trim();
  document.getElementById("q").value = q;
  offset = 0;
  loadMovs();
});

document.getElementById("btn-sync").addEventListener("click", async function(){
  clearInlineBanner("bn-summary");
  clearInlineBanner("bn-capa");
  clearInlineBanner("bn-movs");
  var btn = document.getElementById("btn-sync");
  btn.disabled = true;
  btn.textContent = "Sincronizando…";
  try{
    var r = await fetch("/processos/" + encodeURIComponent(CNJ) + "/sync", { method:"POST" });
    var j = null;
    try { j = await r.json(); } catch(_e){ j = { ok:false, error:"RESPOSTA_INVALIDA", message:"Resposta não-JSON do servidor." }; }
    if(j.ok){
      showToast("success","Sync concluído","Novos eventos: " + j.new_events);
      offset = 0;
      await loadMovs();
      await loadCapaPalatavel();
    } else {
      var msg = (j.message || j.error || ("HTTP " + r.status));
      showToast("danger","Sync falhou", msg);
      showInlineBanner("bn-summary","danger","Sync falhou", esc(msg));
    }
  } catch(e){
    showToast("danger","Erro ao sincronizar", e.toString());
    showInlineBanner("bn-summary","danger","Erro ao sincronizar", esc(e.toString()));
  } finally {
    btn.disabled = false;
    btn.textContent = "Sync agora";
  }
});


// Documentos tab
try{
  var btnDocsRefresh = document.getElementById("btn-docs-refresh");
  if(btnDocsRefresh) btnDocsRefresh.addEventListener("click", function(){ loadDocs(); });

  var btnDocsUpdate = document.getElementById("btn-docs-update");
  if(btnDocsUpdate) btnDocsUpdate.addEventListener("click", function(){ requestDocsUpdate(); });

  var rdPub = document.getElementById("docsTipoPub");
  var rdAutos = document.getElementById("docsTipoAutos");
  if(rdPub) rdPub.addEventListener("change", function(){ loadDocs(); });
  if(rdAutos) rdAutos.addEventListener("change", function(){ loadDocs(); });

  var tabDocsBtn = document.querySelector('button[data-bs-target="#pane-docs"]');
  if(tabDocsBtn){
    tabDocsBtn.addEventListener("click", function(){
      // carrega na primeira abertura
      setTimeout(loadDocs, 50);
    });
  }
}catch(_e){}

async function loadLinkedDocs(){
  clearInlineBanner("bn-linked");
  var box = document.getElementById("linkedDocs");
  box.textContent = "Carregando…";
  try{
    var r = await fetch("/processos/" + encodeURIComponent(CNJ) + "/docs");
    var j = await r.json();
    if(!j.ok){
      showInlineBanner("bn-linked","warning","Falha ao carregar vínculos", esc(j.message || j.error || "erro"));
      box.textContent = "Falha ao carregar vínculos.";
      return;
    }
  var items = j.items || [];
  if(items.length === 0){
    box.innerHTML = "<span class='muted'>Nenhum doc vinculado ainda.</span>";
    return;
  }
  var html = "";
  for(var i=0;i<items.length;i++){
    var x = items[i];
    html += "<span class='badge-soft p-2 me-2 mb-2 d-inline-flex align-items-center gap-2'>"
      + "<span class='mono'>" + esc(x.doc) + "</span>"
      + "<button class='btn btn-outline-danger btn-mini js-unlink' data-doc='" + esc(x.doc) + "'>remover</button>"
      + "</span>";
  }
  box.innerHTML = html;
  // bind unlink handlers (avoid inline onclick quoting issues)
  var bs = box.querySelectorAll(".js-unlink");
  for(var k=0;k<bs.length;k++){
    (function(doc){
      bs[k].addEventListener("click", function(){ unlinkDoc(doc); });
    })(bs[k].getAttribute("data-doc"));
  }
  } catch(e){
    showInlineBanner("bn-linked","danger","Erro ao carregar vínculos", esc(e.toString()));
    box.textContent = "Falha ao carregar vínculos.";
  }
}

async function unlinkDoc(doc){
  if(!confirm("Remover vínculo desse doc?\\n" + doc)) return;
  var r = await fetch("/docs/link", { method:"DELETE", headers:{ "Content-Type":"application/json" }, body: JSON.stringify({ doc: doc, cnj: CNJ }) });
  var j = await r.json();
  if(j.ok) loadLinkedDocs();
  else showToast("danger","Falhou",(j.message || j.error || "erro"));
}

document.getElementById("btnLinkDoc").addEventListener("click", async function(){
  var status = document.getElementById("link-status");
  var fromSelect = (document.getElementById("docSelect").value || "").trim();
  var fromManual = (document.getElementById("docManual").value || "").trim();
  var doc = fromManual || fromSelect;
  if(!doc){
    status.textContent = "Informe um CPF/CNPJ.";
    return;
  }
  status.textContent = "Vinculando…";
  var r = await fetch("/docs/link", { method:"POST", headers:{ "Content-Type":"application/json" }, body: JSON.stringify({ doc: doc, cnj: CNJ }) });
  var j = await r.json();
  if(j.ok){
    status.textContent = "Vínculo criado ✅";
    document.getElementById("docManual").value = "";
    loadLinkedDocs();
  } else {
    status.textContent = "Falhou: " + (j.message || j.error || "erro");
  }
});

// Quick chips (tipo)
function setActiveChip(tipoValue){
  var chips = document.querySelectorAll("#quick-chips .chip-click");
  for(var i=0;i<chips.length;i++){
    var el = chips[i];
    var t = (el.getAttribute("data-t")||"").toUpperCase();
    if(t === (tipoValue||"").toUpperCase() && t !== ""){
      el.classList.add("chip-active");
    } else {
      el.classList.remove("chip-active");
    }
  }
}
document.getElementById("quick-chips").addEventListener("click", function(ev){
  var el = ev.target;
  if(!el || !el.classList.contains("chip-click")) return;
  var t = (el.getAttribute("data-t") || "").toUpperCase();
  tipo = t;
  document.getElementById("tipo").value = tipo;
  offset = 0;
  setActiveChip(tipo);
  loadMovs();
});


// Lazy-load Pedidos & Multas when tab is opened
try{
  var tabPed = document.querySelector('button[data-bs-target="#pane-pedmult"]');
  if(tabPed){
    tabPed.addEventListener("click", function(){
      if(!pedmultLoadedOnce){
        pedmultLoadedOnce = true;
        loadPedidosMultas(false);
      }
    });
  }
  var btnRefresh = document.getElementById("btn-pedmult-refresh");
  if(btnRefresh){
    btnRefresh.addEventListener("click", function(){ loadPedidosMultas(true); });
  }
}catch(e){}
loadCapaPalatavel();
loadMovs();
loadLinkedDocs();
</script>
    """

    body = body.replace("__CNJ__", cnj).replace("__OPTS__", opts).replace("__CNJ_JSON__", cnj_json)
    return render_template_string(UI_BASE, body=body)


# ---------------- UI API proxies ----------------
@app.get("/ui/api/processo/<path:cnj>/capa")
def ui_api_capa(cnj: str):
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        return jsonify({"ok": False, "error": "CNJ_INVALIDO"}), 400

    refresh = (request.args.get("refresh") or "").strip() in ("1", "true", "yes")
    now_iso = datetime.now(timezone.utc).isoformat()

    def _get_cached():
        try:
            with db_connect() as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("SELECT payload, updated_at FROM capa_cache WHERE cnj=?", (cnj,))
                r = cur.fetchone()
                if not r:
                    return None
                updated_at = r["updated_at"]
                try:
                    ts = datetime.fromisoformat(updated_at.replace("Z", "+00:00"))
                except Exception:
                    ts = None
                if ts and (datetime.now(timezone.utc) - ts).total_seconds() <= CAPA_CACHE_TTL_SECONDS:
                    return json.loads(r["payload"])
        except Exception:
            return None
        return None

    def _set_cached(payload: dict):
        try:
            with db_connect() as conn:
                cur = conn.cursor()
                cur.execute(
                    "INSERT INTO capa_cache(cnj,payload,updated_at) VALUES(?,?,?) "
                    "ON CONFLICT(cnj) DO UPDATE SET payload=excluded.payload, updated_at=excluded.updated_at",
                    (cnj, json.dumps(payload, ensure_ascii=False), now_iso),
                )
                conn.commit()
        except Exception:
            pass

    if not refresh:
        cached = _get_cached()
        if cached is not None:
            return jsonify({"ok": True, "cached": True, "data": cached})

    try:
        data = client.obter_capa_processo(cnj)  # type: ignore[attr-defined]
        _set_cached(data)
        return jsonify({"ok": True, "cached": False, "data": data})
    except requests.RequestException as e:
        _set_last_api_error(method="GET", path="/processos/{cnj}/capa", status=None, message=str(e))
        cached = _get_cached()
        if cached is not None:
            return jsonify({"ok": True, "cached": True, "stale": True, "data": cached, "warning": "API_FAIL_USING_CACHE"})
        return jsonify({"ok": False, "error": "ESCAVADOR_UNAVAILABLE", "message": str(e)}), 502


def _collect_key_anywhere(obj: Any, keys: set) -> List[Any]:
    # Percorre recursivamente dict/list e coleta valores de chaves específicas.
    out: List[Any] = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            lk = str(k).strip().lower()
            if lk in keys:
                out.append(v)
            if isinstance(v, (dict, list)):
                out.extend(_collect_key_anywhere(v, keys))
    elif isinstance(obj, list):
        for it in obj:
            if isinstance(it, (dict, list)):
                out.extend(_collect_key_anywhere(it, keys))
    return out


def _normalize_to_list(v: Any) -> List[Any]:
    if v is None:
        return []
    if isinstance(v, list):
        return v
    if isinstance(v, dict):
        return [v]
    return [v]


def _flatten_strings(obj: Any, limit: int = 5000) -> List[str]:
    out: List[str] = []

    def walk(x: Any):
        if len(out) >= limit:
            return
        if x is None:
            return
        if isinstance(x, str):
            s = re.sub(r"\s+", " ", x).strip()
            if s:
                out.append(s)
            return
        if isinstance(x, (int, float, bool)):
            out.append(str(x))
            return
        if isinstance(x, dict):
            for v in x.values():
                walk(v)
            return
        if isinstance(x, list):
            for it in x:
                walk(it)

    walk(obj)
    return out


_PEDIDO_PATTERNS = [
    (re.compile(r"pedido(?:s)?[:\-\s]*(.{0,180})", re.I), "Pedido"),
    (re.compile(r"requer(?:eu|imento|imentos)?[:\-\s]*(.{0,180})", re.I), "Requerimento"),
    (re.compile(r"liminar.{0,180}", re.I), "Liminar"),
    (re.compile(r"tutela(?:\s+de\s+urg[eê]ncia|\s+antecipada)?.{0,180}", re.I), "Tutela"),
    (re.compile(r"indeniza[cç][aã]o.{0,180}", re.I), "Indenização"),
    (re.compile(r"danos?\s+(?:morais|materiais).{0,180}", re.I), "Danos"),
    (re.compile(r"obriga[cç][aã]o\s+de\s+fazer.{0,180}", re.I), "Obrigação de fazer"),
    (re.compile(r"embargos?.{0,180}", re.I), "Embargos"),
    (re.compile(r"execu[cç][aã]o.{0,180}", re.I), "Execução"),
]

_MULTA_PATTERNS = [
    (re.compile(r"multa.{0,180}", re.I), "Multa"),
    (re.compile(r"astreintes.{0,180}", re.I), "Astreintes"),
    (re.compile(r"cl[aá]usula\s+penal.{0,180}", re.I), "Cláusula penal"),
    (re.compile(r"40%\s+do\s+fgts.{0,180}", re.I), "Multa FGTS 40%"),
    (re.compile(r"art(?:igo)?\.?\s*467\s+da\s+clt.{0,180}", re.I), "Multa art. 467 CLT"),
    (re.compile(r"art(?:igo)?\.?\s*477\s+da\s+clt.{0,180}", re.I), "Multa art. 477 CLT"),
]


_VALOR_PATTERNS = [
    (re.compile(r"\bcustas?\b.{0,120}", re.I), "Custas"),
    (re.compile(r"\bhonor[aá]rios?\b.{0,120}", re.I), "Honorários"),
    (re.compile(r"\bvalor\s+da\s+causa\b.{0,120}", re.I), "Valor da causa"),
    (re.compile(r"\bindeniza[cç][aã]o\b.{0,120}", re.I), "Indenização"),
    (re.compile(r"\bcondena[cç][aã]o\b.{0,120}", re.I), "Condenação"),
    (re.compile(r"\bpenhora\b.{0,120}", re.I), "Penhora"),
    (re.compile(r"\bbloqueio\b.{0,120}", re.I), "Bloqueio"),
    (re.compile(r"\bdep[oó]sito\b.{0,120}", re.I), "Depósito"),
]

_MONEY_RX = re.compile(r"(R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}|R\$\s*\d+,\d{2}|\d{1,3}(?:\.\d{3})*,\d{2})", re.I)




def _maybe_contains_money_or_keyword(text: str) -> bool:
    if not text:
        return False
    s = str(text)
    low = s.lower()
    if 'r$' in low:
        return True
    if re.search(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\b", s):
        return True
    keywords = (
        'custa', 'honor', 'multa', 'astreinte', 'indeniza', 'condena',
        'valor da causa', 'penhora', 'bloque', 'sisba', 'bacen', 'depósito', 'deposito'
    )
    return any(k in low for k in keywords)

def _extract_money_mentions(text: str, source: str, data_ref: Optional[str] = None) -> List[Dict[str, Any]]:
    if not text:
        return []
    base = re.sub(r"\s+", " ", text).strip()
    out: List[Dict[str, Any]] = []
    for m in _MONEY_RX.finditer(base):
        amount = m.group(1).strip()
        start = max(0, m.start() - 100)
        end = min(len(base), m.end() + 100)
        trecho = base[start:end].strip(" .;:-")
        tipo = "Valor"
        low = trecho.lower()
        if "custa" in low:
            tipo = "Custas"
        elif "honor" in low:
            tipo = "Honorários"
        elif "multa" in low:
            tipo = "Multa"
        elif "astreinte" in low:
            tipo = "Astreintes"
        elif "indeniza" in low:
            tipo = "Indenização"
        elif "valor da causa" in low:
            tipo = "Valor da causa"
        elif "condena" in low:
            tipo = "Condenação"
        elif "penhor" in low:
            tipo = "Penhora"
        elif "bloque" in low or "sisba" in low or "bacen" in low:
            tipo = "Bloqueio"
        out.append({
            "tipo": tipo,
            "descricao": trecho,
            "valor": amount if amount.upper().startswith("R$") else f"R$ {amount}",
            "origem": source,
            "data": data_ref or "",
        })
    return out


def _extract_keyword_values(text: str, source: str, data_ref: Optional[str] = None) -> List[Dict[str, Any]]:
    if not text:
        return []
    base = re.sub(r"\s+", " ", text).strip()
    out: List[Dict[str, Any]] = []
    for rx, label in _VALOR_PATTERNS:
        for m in rx.finditer(base):
            trecho = m.group(0).strip(" .;:-")
            if not trecho:
                continue
            money = _MONEY_RX.search(trecho)
            out.append({
                "tipo": label,
                "descricao": trecho,
                "valor": (money.group(1).strip() if money else ""),
                "origem": source,
                "data": data_ref or "",
            })
    return out


def _extract_matches_from_text(text: str, source: str, kind: str, data_ref: Optional[str] = None) -> List[Dict[str, Any]]:
    if not text:
        return []
    patterns = _PEDIDO_PATTERNS if kind == "pedido" else _MULTA_PATTERNS
    out: List[Dict[str, Any]] = []
    base = re.sub(r"\s+", " ", text).strip()
    for rx, label in patterns:
        for m in rx.finditer(base):
            trecho = m.group(0).strip(" .;:-")
            if not trecho:
                continue
            out.append({
                "tipo": label,
                "descricao": trecho,
                "origem": source,
                "data": data_ref or "",
            })
    return out


def _dedupe_any(items: List[Any]) -> List[Any]:
    seen = set()
    out = []
    for it in items:
        try:
            if isinstance(it, (dict, list)):
                h = stable_hash(it)
            else:
                h = stable_hash({"v": str(it)})
        except Exception:
            h = str(it)
        if h in seen:
            continue
        seen.add(h)
        out.append(it)
    return out


def _looks_like_money_scalar(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    if not s:
        return False
    if _MONEY_RX.search(s):
        return True
    if re.fullmatch(r"\d+[.,]\d{2,4}", s):
        return True
    return False


def _format_scalar_money(v: Any) -> str:
    s = str(v).strip()
    if not s:
        return ""
    m = _MONEY_RX.search(s)
    if m:
        raw = m.group(1).strip()
        return raw if raw.upper().startswith("R$") else f"R$ {raw}"
    if re.fullmatch(r"\d+[.,]\d{2,4}", s):
        try:
            num = float(s.replace('.', '').replace(',', '.')) if ',' in s and '.' in s else float(s.replace(',', '.'))
            inte = int(num)
            dec = f"{num:.2f}".split('.')[-1]
            inteiro_fmt = f"{inte:,}".replace(',', '.')
            return f"R$ {inteiro_fmt},{dec}"
        except Exception:
            return s
    return s


def _normalize_valor_item(v: Any) -> Optional[Dict[str, Any]]:
    if v is None:
        return None
    if isinstance(v, dict):
        tipo = str(v.get("tipo") or v.get("natureza") or "Valor").strip() or "Valor"
        descricao = str(v.get("descricao") or v.get("titulo") or v.get("texto") or "").strip()
        valor_raw = v.get("valor_formatado") or v.get("valor") or v.get("valor_causa") or v.get("custas") or v.get("honorarios") or v.get("honorários")
        moeda = str(v.get("moeda") or "R$").strip()
        origem = str(v.get("origem") or "").strip()
        data = str(v.get("data") or "").strip()
        if valor_raw is None:
            for k, vv in v.items():
                if any(token in str(k).lower() for token in ["valor", "custa", "honor", "multa", "inden", "conden"]):
                    valor_raw = vv
                    break
        if valor_raw is None and not descricao:
            return None
        valor_fmt = _format_scalar_money(valor_raw) if valor_raw is not None else ""
        if not valor_fmt and not descricao:
            return None
        return {
            "tipo": tipo,
            "descricao": descricao,
            "valor": valor_fmt,
            "origem": origem,
            "data": data,
            "moeda": moeda,
        }
    if isinstance(v, (int, float)):
        if isinstance(v, int):
            return None
        return {"tipo": "Valor", "descricao": "", "valor": _format_scalar_money(v), "origem": "capa", "data": ""}
    s = str(v).strip()
    if not _looks_like_money_scalar(s):
        return None
    return {"tipo": "Valor", "descricao": "", "valor": _format_scalar_money(s), "origem": "capa", "data": ""}



@app.get("/ui/api/processo/<path:cnj>/pedidos-multas")
def ui_api_pedidos_multas(cnj: str):
    require_token_configured()
    if not CNJ_REGEX.fullmatch(cnj):
        return jsonify({"ok": False, "error": "CNJ_INVALIDO"}), 400

    refresh = (request.args.get("refresh") or "").strip() in ("1", "true", "yes")
    now_iso = datetime.now(timezone.utc).isoformat()

    def _get_cached_capa():
        try:
            with db_connect() as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("SELECT payload, updated_at FROM capa_cache WHERE cnj=?", (cnj,))
                r = cur.fetchone()
                if not r:
                    return None, None
                updated_at = r["updated_at"]
                try:
                    ts = datetime.fromisoformat(updated_at.replace("Z", "+00:00"))
                except Exception:
                    ts = None
                payload = json.loads(r["payload"])
                if ts and (datetime.now(timezone.utc) - ts).total_seconds() <= CAPA_CACHE_TTL_SECONDS:
                    return payload, False
                return payload, True
        except Exception:
            return None, None

    def _set_cached_capa(payload: dict):
        try:
            with db_connect() as conn:
                cur = conn.cursor()
                cur.execute(
                    "INSERT INTO capa_cache(cnj,payload,updated_at) VALUES(?,?,?) "
                    "ON CONFLICT(cnj) DO UPDATE SET payload=excluded.payload, updated_at=excluded.updated_at",
                    (cnj, json.dumps(payload, ensure_ascii=False), now_iso),
                )
                conn.commit()
        except Exception:
            pass

    cached_payload, is_stale = (None, None)
    if not refresh:
        cached_payload, is_stale = _get_cached_capa()

    data = None
    used_cache = False
    stale_used = False

    if not refresh:
        if cached_payload is not None and is_stale is False:
            data = cached_payload
            used_cache = True
        else:
            try:
                data = client.obter_capa_processo(cnj)  # type: ignore[attr-defined]
                _set_cached_capa(data)
            except requests.RequestException as e:
                _set_last_api_error(method="GET", path="/processos/{cnj}", status=None, message=str(e))
                if cached_payload is not None:
                    data = cached_payload
                    used_cache = True
                    stale_used = True
                else:
                    return jsonify({"ok": False, "error": "ESCAVADOR_UNAVAILABLE", "message": str(e)}), 502
    else:
        try:
            data = client.obter_capa_processo(cnj)  # type: ignore[attr-defined]
            _set_cached_capa(data)
        except requests.RequestException as e:
            _set_last_api_error(method="GET", path="/processos/{cnj}", status=None, message=str(e))
            cached_payload, _st = _get_cached_capa()
            if cached_payload is not None:
                data = cached_payload
                used_cache = True
                stale_used = True
            else:
                return jsonify({"ok": False, "error": "ESCAVADOR_UNAVAILABLE", "message": str(e)}), 502

    pedidos_vals = _collect_key_anywhere(data, {"pedidos", "pedido"})
    multas_vals = _collect_key_anywhere(data, {"multas", "multa"})

    pedidos: List[Any] = []
    multas: List[Any] = []
    for v in pedidos_vals:
        pedidos.extend(_normalize_to_list(v))
    for v in multas_vals:
        multas.extend(_normalize_to_list(v))

    # Heurística complementar: examina capa, movimentações locais e documentos já baixados/cacheados.
    text_sources: List[Tuple[str, str, Optional[str]]] = []
    for s in _flatten_strings(data):
        if len(s) >= 12:
            text_sources.append(("capa", s, None))

    with db_connect() as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute(
                "SELECT data, tipo, texto, raw_json FROM eventos_mov WHERE cnj=? ORDER BY COALESCE(data,'') DESC LIMIT 1000",
                (cnj,),
            )
            for row in cur.fetchall() or []:
                parts = []
                if row["tipo"]:
                    parts.append(str(row["tipo"]))
                if row["texto"]:
                    parts.append(str(row["texto"]))
                merged = " | ".join([p for p in parts if p])
                if merged:
                    text_sources.append(("movimentação", merged, row["data"]))
                if row["raw_json"]:
                    try:
                        raw_obj = json.loads(row["raw_json"])
                        for s in _flatten_strings(raw_obj, limit=120):
                            if _maybe_contains_money_or_keyword(s):
                                text_sources.append(("movimentação", s, row["data"]))
                    except Exception:
                        raw_s = str(row["raw_json"])
                        if _maybe_contains_money_or_keyword(raw_s):
                            text_sources.append(("movimentação", raw_s, row["data"]))
        except Exception:
            pass
        try:
            cur.execute(
                "SELECT tipo, titulo, data, meta_json FROM documentos_cache WHERE cnj=? ORDER BY COALESCE(data,'') DESC LIMIT 200",
                (cnj,),
            )
            for row in cur.fetchall() or []:
                parts = []
                if row["titulo"]:
                    parts.append(str(row["titulo"]))
                if row["meta_json"]:
                    try:
                        parts.extend(_flatten_strings(json.loads(row["meta_json"]), limit=80))
                    except Exception:
                        parts.append(str(row["meta_json"]))
                merged = " | ".join([p for p in parts if p])
                if merged:
                    origem = "documento" if row["tipo"] == "documentos_publicos" else "autos"
                    text_sources.append((origem, merged, row["data"]))
        except Exception:
            pass

    extracted_pedidos: List[Dict[str, Any]] = []
    extracted_multas: List[Dict[str, Any]] = []
    extracted_valores: List[Dict[str, Any]] = []
    for source, txt, dt in text_sources:
        extracted_pedidos.extend(_extract_matches_from_text(txt, source, "pedido", dt))
        extracted_multas.extend(_extract_matches_from_text(txt, source, "multa", dt))
        extracted_valores.extend(_extract_keyword_values(txt, source, dt))
        extracted_valores.extend(_extract_money_mentions(txt, source, dt))

    pedidos.extend(extracted_pedidos)
    multas.extend(extracted_multas)

    valores: List[Any] = []
    for v in _collect_key_anywhere(data, {"valor", "valor_formatado", "valor_causa", "custas", "honorarios", "honorários"}):
        valores.extend(_normalize_to_list(v))
    valores.extend(extracted_valores)

    valores_norm: List[Dict[str, Any]] = []
    for item in valores:
        norm = _normalize_valor_item(item)
        if norm:
            valores_norm.append(norm)

    def _valor_sort_key(x: Dict[str, Any]):
        origem = (x.get("origem") or "").lower()
        prioridade = {"movimentação": 0, "documento": 1, "autos": 2, "capa": 3}.get(origem, 9)
        data = x.get("data") or ""
        return (prioridade, str(data))

    valores_norm.sort(key=_valor_sort_key)

    pedidos = _dedupe_any(pedidos)
    multas = _dedupe_any(multas)
    valores = _dedupe_any(valores_norm)

    return jsonify({
        "ok": True,
        "cached": bool(used_cache),
        "stale": bool(stale_used),
        "pedidos": pedidos,
        "multas": multas,
        "valores": valores,
        "fontes": sorted(list({src for src, _, _ in text_sources})),
    })


@app.get("/ui/admin")
def ui_admin():
    body = r"""
    <div class="card p-3">
      <div class="h5 mb-1">Admin</div>
      <div class="muted">Ajustes operacionais do monitor (auto-discover, saúde, diagnósticos).</div>

      <div class="mt-3"><a class="btn btn-outline-primary btn-sm" href="/ui/admin/monitoramentos">Abrir Monitoramentos</a></div>

      <hr class="divider">

      <div class="row g-2 align-items-end">
        <div class="col-md-4">
          <label class="form-label small muted">Adicionar doc (CPF/CNPJ)</label>
          <input id="docInput" class="form-control" placeholder="000.000.000-00 ou 00.000.000/0000-00">
        </div>
        <div class="col-md-2">
          <button id="btnAdd" class="btn btn-primary w-100">Adicionar</button>
        </div>
        <div class="col-md-6">
          <div class="muted small">Dica: o botão “Descobrir” usa a rota <code>/envolvido/processos</code> da API v2 do Escavador.</div>
          <div id="status" class="muted small mt-1"></div>
        </div>
      </div>

      <hr class="divider">

      <div class="card p-3 mb-3" style="background: rgba(255,255,255,.02); border: 1px solid rgba(255,255,255,.06);">
        <div class="d-flex justify-content-between align-items-center">
          <div>
            <div class="h6 mb-0">Auto-Discover</div>
            <div class="muted small">Descoberta automática de CNJs para cada doc da watchlist.</div>
          </div>
          <div class="form-check form-switch">
            <input class="form-check-input" type="checkbox" role="switch" id="adEnabled">
            <label class="form-check-label small muted" for="adEnabled">Ativo</label>
          </div>
        </div>

        <div class="row g-2 mt-2 align-items-end">
          <div class="col-md-3">
            <label class="form-label small muted">Intervalo (seg)</label>
            <input id="adInterval" type="number" min="10" class="form-control" placeholder="ex: 900">
          </div>
          <div class="col-md-3">
            <label class="form-label small muted">Limite por doc</label>
            <input id="adLimit" type="number" min="1" class="form-control" placeholder="ex: 50">
          </div>
          <div class="col-md-3">
            <label class="form-label small muted">Máx docs/ciclo</label>
            <input id="adMaxDocs" type="number" min="1" class="form-control" placeholder="ex: 50">
          </div>
          <div class="col-md-3">
            <label class="form-label small muted">Modo econômico</label>
            <div class="form-check">
              <input class="form-check-input" type="checkbox" id="adOnlyNoLinks">
              <label class="form-check-label small" for="adOnlyNoLinks">Só se não houver vínculos</label>
            </div>
          </div>
        </div>

        <div class="d-flex flex-wrap gap-2 mt-3 align-items-center">
          <button id="btnSaveDiscover" class="btn btn-outline-light">Salvar</button>
          <button id="btnReloadDiscover" class="btn btn-outline-secondary">Recarregar</button>
          <button id="btnRunDiscover" class="btn btn-outline-primary">Rodar agora</button>
          <div class="muted small" id="adStatus"></div>
        </div>
        <div class="mt-3">
          <div class="muted small mb-1">Status do último ciclo</div>
          <div id="adCycleStatus" class="small muted">Carregando…</div>
        </div>
      </div>

      <hr class="divider">

      <hr class="divider">
      <div class="muted small">Atalhos</div>
      <ul class="mb-0 small">
        <li><a href="/health">/health</a></li>
        <li><code>/poll/run-once</code> (GET/POST)</li>
      </ul>
    </div>

    <script>
      const $ = (id) => document.getElementById(id);
      const status = (t) => { $("status").textContent = t || ""; };

      function esc(s){ return (""+s).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }

      async function api(path, opts){
        const r = await fetch(path, opts || {});
        const j = await r.json().catch(()=>({ok:false,error:"Resposta inválida"}));
        if(!r.ok){ throw new Error((j && (j.error||j.message)) || ("HTTP "+r.status)); }
        return j;
      }


      async function loadDiscoverConfig(){
        try{
          const j = await api("/admin/discover");
          $("adEnabled").checked = !!j.auto_discover_enabled;
          $("adInterval").value = j.discover_interval_seconds ?? "";
          $("adLimit").value = j.discover_limit_per_doc ?? "";
          $("adMaxDocs").value = j.discover_max_docs_per_cycle ?? "";
          $("adOnlyNoLinks").checked = !!j.discover_only_if_no_links;
          $("adStatus").textContent = "Config carregada.";
        }catch(e){
          $("adStatus").textContent = "Falha ao carregar: " + e.message;
        }
      }

      async function saveDiscoverConfig(){
        const payload = {
          enabled: $("adEnabled").checked,
          discover_interval_seconds: Number($("adInterval").value || 0),
          discover_limit_per_doc: Number($("adLimit").value || 0),
          discover_max_docs_per_cycle: Number($("adMaxDocs").value || 0),
          discover_only_if_no_links: $("adOnlyNoLinks").checked
        };
        $("adStatus").textContent = "Salvando…";
        try{
          const j = await api("/admin/discover", {
            method: "POST",
            headers: {"Content-Type":"application/json"},
            body: JSON.stringify(payload)
          });
          $("adStatus").textContent = "Salvo. Ativo=" + (j.auto_discover_enabled ? "sim" : "não");
        }catch(e){
          $("adStatus").textContent = "Falha ao salvar: " + e.message;
        }
      }


      function fmtTotals(t){
        if(!t) return "";
        return `docs=${t.docs} ran=${t.ran_docs} skipped=${t.skipped} discovered=${t.discovered} inserted=${t.inserted_processos} linked=${t.linked} errors=${t.errors}`;
      }

      async function loadDiscoverStatus(){
        try{
          const j = await api("/admin/discover/status");
          const st = j.state || {};
          let html = "";
          if(st.running){
            html += `<div><span class="spinner-border spinner-border-sm me-2"></span><b>Rodando agora</b> (${esc(st.last_trigger||"")})</div>`;
          }else{
            html += `<div><b>Parado</b></div>`;
          }
          if(st.last_started) html += `<div class="muted small">Início: ${esc(st.last_started)}</div>`;
          if(st.last_finished) html += `<div class="muted small">Fim: ${esc(st.last_finished)}</div>`;
          if(st.last_totals) html += `<div class="mt-1 mono small">${esc(fmtTotals(st.last_totals))}</div>`;
          if(st.last_error) html += `<div class="text-warning small mt-1">Erro: ${esc(st.last_error)}</div>`;
          $("adCycleStatus").innerHTML = html || '<span class="muted">Sem execuções ainda.</span>';
        }catch(e){
          $("adCycleStatus").textContent = "Falha ao carregar status: " + e.message;
        }
      }

      async function runDiscoverNow(){
        $("adStatus").textContent = "Disparando ciclo…";
        try{
          await api("/admin/discover/run-once", { method: "POST" });
          $("adStatus").textContent = "Ciclo iniciado ✅";
          await loadDiscoverStatus();
        }catch(e){
          $("adStatus").textContent = "Não foi possível iniciar: " + e.message;
        }
      }

      async function loadWatchlist(){
        const j = await api("/watchlist");
        const items = j.items || [];
        if(!items.length){
          $("wl").innerHTML = '<div class="muted">Nenhum doc cadastrado.</div>';
          return;
        }

        let html = '';
        for(const it of items){
          const doc = it.doc;
          html += `
            <div class="card p-2 mb-2">
              <div class="d-flex justify-content-between align-items-start">
                <div>
                  <div><b>${esc(doc)}</b> <span class="muted">(${esc(it.tipo_doc)})</span></div>
                  <div class="muted small">Criado em: ${esc(it.created_at || "")}</div>
                </div>
                <div class="d-flex gap-2">
                  <button class="btn btn-sm btn-outline-primary" onclick="discover('${esc(doc)}')">Descobrir</button>
                  <button class="btn btn-sm btn-outline-secondary" onclick="ack('${esc(doc)}')">Zerar alertas</button>
                </div>
              </div>

              <div class="row mt-2">
                <div class="col-md-7">
                  <div class="muted small mb-1">Processos vinculados</div>
                  <div id="p_${esc(doc)}" class="small muted">Carregando…</div>
                </div>
                <div class="col-md-5">
                  <div class="muted small mb-1">Alertas</div>
                  <div id="a_${esc(doc)}" class="small muted">Carregando…</div>
                </div>
              </div>
            </div>
          `;
        }
        $("wl").innerHTML = html;

        // carregar detalhes (processos + alertas)
        for(const it of items){
          await refreshDoc(it.doc);
        }
      }

      async function refreshDoc(doc){
        try{
          const pj = await api(`/docs/${encodeURIComponent(doc)}/processos`);
          const aj = await api(`/docs/${encodeURIComponent(doc)}/alerts`);

          const pEl = document.getElementById(`p_${doc}`);
          const aEl = document.getElementById(`a_${doc}`);

          const procs = pj.items || [];
          if(!procs.length){
            pEl.innerHTML = '<span class="muted">Nenhum vínculo ainda.</span>';
          }else{
            pEl.innerHTML = procs.map(x => `<a href="/ui/processo/${encodeURIComponent(x.cnj)}">${esc(x.cnj)}</a>`).join("<br>");
          }

          aEl.innerHTML = `
            <div><b>${aj.new_events}</b> novos eventos</div>
            <div class="muted small">last_event_id=${aj.last_event_id} | max_event_id=${aj.max_event_id}</div>
          `;
        }catch(e){
          const pEl = document.getElementById(`p_${doc}`);
          const aEl = document.getElementById(`a_${doc}`);
          if(pEl) pEl.textContent = "Erro: " + e.message;
          if(aEl) aEl.textContent = "Erro: " + e.message;
        }
      }

      window.discover = async (doc) => {
        status("Descobrindo processos…");
        try{
          const j = await api(`/docs/${encodeURIComponent(doc)}/discover`, {method:"POST"});
          status(`OK: descobertos=${j.discovered}, novos_processos=${j.inserted_processos}, vinculados=${(j.linked||[]).length}`);
          await refreshDoc(doc);
        }catch(e){
          status("Erro ao descobrir: " + e.message);
        }
      };

      window.ack = async (doc) => {
        status("Zerando alertas…");
        try{
          const j = await api(`/docs/${encodeURIComponent(doc)}/alerts/ack`, {method:"POST"});
          status(`Alertas zerados até event_id=${j.acked_to_event_id}`);
          await refreshDoc(doc);
        }catch(e){
          status("Erro ao zerar: " + e.message);
        }
      };

      $("btnAdd").addEventListener("click", async () => {
        const doc = $("docInput").value.trim();
        if(!doc){ status("Informe um CPF/CNPJ."); return; }
        status("Adicionando…");
        try{
          const r = await api("/watchlist", {method:"POST", headers:{"Content-Type":"application/json"}, body: JSON.stringify({doc})});
          status("Adicionado. Agora clique em “Descobrir”.");
          $("docInput").value = "";
          await loadWatchlist();
        }catch(e){
          status("Erro ao adicionar: " + e.message);
        }
      });

      $("btnSaveDiscover").addEventListener("click", async ()=>{ await saveDiscoverConfig(); });
      $("btnReloadDiscover").addEventListener("click", async ()=>{ await loadDiscoverConfig(); });
$("btnRunDiscover").addEventListener("click", async ()=>{ await runDiscoverNow(); });

      loadDiscoverConfig();
      loadDiscoverStatus();
      setInterval(loadDiscoverStatus, 5000);
      loadWatchlist();
    </script>
    """
    return render_template_string(UI_BASE, body=body)


# ============================================================
# Main
# ===============================

@app.get("/ui/watchlist")
def ui_watchlist():
    body = r"""
    <div class="card p-3">
      <div class="h5 mb-1">Watch-List</div>
      <div class="muted">Gerencie CPF/CNPJ monitorados, descubra CNJs e visualize processos associados.</div>

      <hr class="divider">

      <div class="row g-2 align-items-end">
        <div class="col-md-4">
          <label class="form-label small muted">Adicionar doc (CPF/CNPJ)</label>
          <input id="docInput" class="form-control" placeholder="000.000.000-00 ou 00.000.000/0000-00">
        </div>
        <div class="col-md-2">
          <button id="btnAdd" class="btn btn-primary w-100">Adicionar</button>
        </div>
        <div class="col-md-6">
          <div class="muted small">Dica: o botão “Descobrir” usa a rota <code>/envolvido/processos</code> da API v2 do Escavador.</div>
          <div id="status" class="muted small mt-1"></div>
        </div>
      </div>

      <hr class="divider">

      <div class="muted small mb-1">Watchlist</div>
      <div id="wl" class="small">Carregando…</div>

      <div class="row g-2 mt-3 align-items-end">
        <div class="col-md-6">
          <label class="form-label small muted">Procurar</label>
          <input id="wlSearch" class="form-control" placeholder="digite parte do CPF/CNPJ…">
        </div>
        <div class="col-md-2">
          <button id="btnReloadWl" class="btn btn-outline-primary w-100">Recarregar</button>
        </div>
        <div class="col-md-4 text-md-end">
          <span class="muted small">Clique em um doc para ver processos vinculados.</span>
        </div>
      </div>

      <div class="accordion mt-3" id="wlAcc"></div>

    </div>

    <script>
      async function apiJson(url, opts){
        const r = await fetch(url, opts || {});
        const t = await r.text();
        let j = null;
        try { j = JSON.parse(t); } catch(e) { j = { ok:false, error:"Resposta inválida", raw:t }; }
        if(!r.ok) throw new Error((j && (j.error||j.message)) || ("HTTP "+r.status));
        return j;
      }

      function esc(s){ return String(s||"").replace(/[&<>"']/g, (c)=>({ "&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#039;" }[c])); }

      function normalizeDoc(s){ return String(s||"").trim(); }

      async function loadWatchlist(){
        const wlDiv = document.getElementById("wl");
        const acc = document.getElementById("wlAcc");
        wlDiv.textContent = "Carregando…";
        acc.innerHTML = "";
        try{
          const wl = await apiJson("/watchlist");
          const docs = (wl && wl.items) ? wl.items : (wl || []);
          wlDiv.textContent = `${docs.length} doc(s)`;
          window.__wl_docs = docs;
          renderAcc();
        }catch(e){
          wlDiv.innerHTML = `<span class="text-danger">Falha: ${esc(e.message)}</span>`;
        }
      }

      function renderAcc(){
        const acc = document.getElementById("wlAcc");
        const q = (document.getElementById("wlSearch").value||"").trim();
        const docs = (window.__wl_docs||[]).filter(d => !q || String(d.doc||"").includes(q));
        acc.innerHTML = "";
        docs.forEach((d, idx)=>{
          const doc = d.doc;
          const hid = "wlH"+idx;
          const cid = "wlC"+idx;
          acc.insertAdjacentHTML("beforeend", `
            <div class="accordion-item">
              <h2 class="accordion-header" id="${hid}">
                <button class="accordion-button collapsed" type="button" data-bs-toggle="collapse" data-bs-target="#${cid}" aria-expanded="false" aria-controls="${cid}">
                  <span class="fw-semibold">${esc(doc)}</span>
                  <span class="ms-2 badge text-bg-light">${esc(d.tipo_doc||"")}</span>
                </button>
              </h2>
              <div id="${cid}" class="accordion-collapse collapse" aria-labelledby="${hid}" data-bs-parent="#wlAcc">
                <div class="accordion-body">
                  <div class="d-flex flex-wrap gap-2 mb-2">
                    <button class="btn btn-sm btn-primary" data-action="discover" data-doc="${esc(doc)}">Descobrir</button>
                    <button class="btn btn-sm btn-outline-primary" data-action="refresh" data-doc="${esc(doc)}">Atualizar lista</button>
                  </div>
                  <div class="small muted mb-1">Processos vinculados</div>
                  <div class="small" data-procs="${esc(doc)}">Carregando…</div>
                </div>
              </div>
            </div>
          `);
        });
        // attach events
        acc.querySelectorAll('[data-action="discover"]').forEach(btn=>{
          btn.addEventListener("click", async ()=>{
            const doc = btn.getAttribute("data-doc");
            btn.disabled=true;
            try{
              await apiJson(`/docs/${encodeURIComponent(doc)}/discover`, {method:"POST"});
              if(window.showToast) showToast("success","Discover","Processos descobertos/vinculados.");
              await loadDocProcs(doc);
            }catch(e){
              if(window.showToast) showToast("danger","Discover falhou", e.message);
            }finally{ btn.disabled=false; }
          });
        });
        acc.querySelectorAll('[data-action="refresh"]').forEach(btn=>{
          btn.addEventListener("click", async ()=>{
            const doc = btn.getAttribute("data-doc");
            await loadDocProcs(doc);
          });
        });
        // lazy load procs on expand
        acc.querySelectorAll(".accordion-collapse").forEach(col=>{
          col.addEventListener("shown.bs.collapse", async ()=>{
            const body = col.querySelector("[data-procs]");
            const doc = body.getAttribute("data-procs");
            await loadDocProcs(doc);
          });
        });
      }

      async function loadDocProcs(doc){
        const el = document.querySelector(`[data-procs="${CSS.escape(doc)}"]`);
        if(!el) return;
        el.textContent = "Carregando…";
        try{
          const j = await apiJson(`/docs/${encodeURIComponent(doc)}/processos`);
          const items = (j && j.items) ? j.items : (j || []);
          if(!items.length){
            el.innerHTML = `<span class="muted">Nenhum processo vinculado.</span>`;
            return;
          }
          el.innerHTML = items.map(p=>{
            const cnj = (p.cnj||p);
            return `<a class="chip me-1 mb-1 d-inline-flex" href="/ui/processo/${encodeURIComponent(cnj)}">${esc(cnj)}</a>`;
          }).join(" ");
        }catch(e){
          el.innerHTML = `<span class="text-danger">Falha: ${esc(e.message)}</span>`;
        }
      }

      document.getElementById("btnReloadWl").addEventListener("click", loadWatchlist);
      document.getElementById("wlSearch").addEventListener("input", ()=>renderAcc());

      document.getElementById("btnAdd").addEventListener("click", async ()=>{
        const doc = normalizeDoc(document.getElementById("docInput").value);
        const status = document.getElementById("status");
        status.textContent = "";
        if(!doc){ status.textContent="Informe um CPF/CNPJ."; return; }
        try{
          await apiJson("/watchlist", {method:"POST", headers:{"Content-Type":"application/json"}, body: JSON.stringify({doc})});
          if(window.showToast) showToast("success","Watch-List","Doc adicionado.");
          document.getElementById("docInput").value="";
          await loadWatchlist();
        }catch(e){
          if(window.showToast) showToast("danger","Falha ao adicionar", e.message);
        }
      });

      loadWatchlist();
    </script>


      """
    return render_template_string(UI_BASE, body=body)

def start_background_tasks():
    # Polling loop (callbacks + inbox processing)
    if POLL_INTERVAL_SECONDS > 0 and ESCAVADOR_TOKEN:
        t1 = threading.Thread(target=poll_callbacks_loop, args=(client,), daemon=True)  # type: ignore
        t1.start()
    else:
        if POLL_INTERVAL_SECONDS <= 0:
            logger.info("Polling desabilitado (POLL_INTERVAL_SECONDS<=0).")
        if not ESCAVADOR_TOKEN:
            logger.info("Polling não iniciado: ESCAVADOR_TOKEN ausente.")

    # Auto-discover loop (watchlist -> processes linking)
    if client is not None and DISCOVER_INTERVAL_SECONDS > 0:
        t2 = threading.Thread(target=auto_discover_loop, args=(client,), daemon=True)  # type: ignore
        t2.start()
    else:
        if DISCOVER_INTERVAL_SECONDS <= 0:
            logger.info("Auto-discover desabilitado (DISCOVER_INTERVAL_SECONDS<=0).")
        if client is None:
            logger.info("Auto-discover não iniciado: client indisponível (token ausente).")


# ---------------------------
# Dashboard Executivo (UI)
# ---------------------------

@app.get("/ui/api/dashboard/metrics")
def ui_api_dashboard_metrics():
    """Métricas agregadas (global) e por doc (CPF/CNPJ) para o dashboard.
    Resiliente a bases antigas (tabelas/colunas ainda não criadas).
    Query param opcional: ?doc=<cpf_cnpj>
    """
    doc_filter = (request.args.get("doc") or "").strip()
    scope = (request.args.get("scope") or "").strip().lower()
    if scope == "global":
        doc_filter = ""

    def _safe_err(msg: str):
        logger.exception(msg)

    def _row_count(cur: sqlite3.Cursor, sql: str, params: tuple = ()) -> int:
        cur.execute(sql, params)
        r = cur.fetchone()
        if r is None:
            return 0
        if isinstance(r, sqlite3.Row):
            return int(r[0] or 0)
        return int(r[0] or 0)

    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            def table_exists(name: str) -> bool:
                cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1", (name,))
                return cur.fetchone() is not None

            def col_exists(table: str, col: str) -> bool:
                try:
                    cur.execute(f"PRAGMA table_info({table})")
                    cols = [r[1] for r in cur.fetchall()]
                    return col in cols
                except Exception:
                    return False

            # -----------------------
            # Lista de docs (watchlist)
            # -----------------------
            docs_list: list[str] = []
            if table_exists("watchlist") and col_exists("watchlist", "doc"):
                try:
                    cur.execute("SELECT doc FROM watchlist ORDER BY doc")
                    docs_list = [r[0] for r in cur.fetchall() if r[0]]
                except Exception:
                    docs_list = []

            # -----------------------
            # Métricas base (global)
            # -----------------------
            docs_total = _row_count(cur, "SELECT COUNT(*) FROM watchlist") if table_exists("watchlist") else 0
            processos_total = _row_count(cur, "SELECT COUNT(*) FROM processos") if table_exists("processos") else 0
            movs_total = _row_count(cur, "SELECT COUNT(*) FROM eventos_mov") if table_exists("eventos_mov") else 0

            # Alertas pendentes: esquema legado alert_state(doc,last_event_id)
            alertas_total = 0
            if table_exists("eventos_mov") and table_exists("doc_process") and col_exists("doc_process", "doc") and col_exists("doc_process", "cnj"):
                if table_exists("alert_state") and col_exists("alert_state", "doc") and col_exists("alert_state", "last_event_id"):
                    try:
                        alertas_total = _row_count(cur, """
                            SELECT COUNT(*) 
                            FROM eventos_mov e
                            JOIN doc_process dp ON dp.cnj = e.cnj
                            LEFT JOIN alert_state a ON a.doc = dp.doc
                            WHERE e.id > COALESCE(a.last_event_id, 0)
                        """)
                    except sqlite3.OperationalError:
                        alertas_total = 0
                else:
                    # Sem tabela de ACK: assume tudo como "novo" (mas apenas o que está vinculado a algum doc)
                    try:
                        alertas_total = _row_count(cur, """
                            SELECT COUNT(*)
                            FROM eventos_mov e
                            JOIN doc_process dp ON dp.cnj = e.cnj
                        """)
                    except sqlite3.OperationalError:
                        alertas_total = movs_total
            else:
                alertas_total = movs_total

            
            
            # Custos (estimados) globais
            cost_today_brl = 0.0
            cost_month_brl = 0.0
            if table_exists("api_usage"):
                try:
                    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    cur.execute("SELECT COALESCE(SUM(cost_brl),0) FROM api_usage WHERE substr(ts,1,10)=?", (today,))
                    cost_today_brl = float(cur.fetchone()[0] or 0.0)
                    month = datetime.now(timezone.utc).strftime("%Y-%m")
                    cur.execute("SELECT COALESCE(SUM(cost_brl),0) FROM api_usage WHERE substr(ts,1,7)=?", (month,))
                    cost_month_brl = float(cur.fetchone()[0] or 0.0)
                except Exception:
                    cost_today_brl, cost_month_brl = 0.0, 0.0

            # Custos (reais) globais (importados do extrato)
            cost_real_today_brl = 0.0
            cost_real_month_brl = 0.0
            if table_exists("api_usage_real"):
                try:
                    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                    cur.execute("SELECT COALESCE(SUM(cost_brl),0) FROM api_usage_real WHERE substr(ts,1,10)=?", (today,))
                    cost_real_today_brl = float(cur.fetchone()[0] or 0.0)
                    month = datetime.now(timezone.utc).strftime("%Y-%m")
                    cur.execute("SELECT COALESCE(SUM(cost_brl),0) FROM api_usage_real WHERE substr(ts,1,7)=?", (month,))
                    cost_real_month_brl = float(cur.fetchone()[0] or 0.0)
                except Exception:
                    cost_real_today_brl, cost_real_month_brl = 0.0, 0.0

            cost_delta_today_brl = cost_real_today_brl - cost_today_brl
            cost_delta_month_brl = cost_real_month_brl - cost_month_brl

            # Top CNJs por alertas pendentes (aprox.: pode contar duplicado se CNJ vinculado a múltiplos docs)
            top_cnj_alerts = []
            try:
                if table_exists("eventos_mov") and table_exists("doc_process") and col_exists("doc_process", "doc") and col_exists("doc_process", "cnj") and col_exists("eventos_mov", "cnj"):
                    if doc_filter:
                        # Top CNJs somente do doc
                        if table_exists("alert_state") and col_exists("alert_state", "doc") and col_exists("alert_state", "last_event_id"):
                            cur.execute("""
                                SELECT e.cnj AS cnj, COUNT(*) AS c
                                FROM eventos_mov e
                                JOIN doc_process dp ON dp.cnj = e.cnj
                                LEFT JOIN alert_state a ON a.doc = dp.doc
                                WHERE dp.doc=? AND e.id > COALESCE(a.last_event_id, 0)
                                GROUP BY e.cnj
                                ORDER BY c DESC
                                LIMIT 5
                            """, (doc_filter,))
                        else:
                            cur.execute("""
                                SELECT e.cnj AS cnj, COUNT(*) AS c
                                FROM eventos_mov e
                                JOIN doc_process dp ON dp.cnj = e.cnj
                                WHERE dp.doc=?
                                GROUP BY e.cnj
                                ORDER BY c DESC
                                LIMIT 5
                            """, (doc_filter,))
                    else:
                        if table_exists("alert_state") and col_exists("alert_state", "doc") and col_exists("alert_state", "last_event_id"):
                            cur.execute("""
                                SELECT e.cnj AS cnj, COUNT(*) AS c
                                FROM eventos_mov e
                                JOIN doc_process dp ON dp.cnj = e.cnj
                                LEFT JOIN alert_state a ON a.doc = dp.doc
                                WHERE e.id > COALESCE(a.last_event_id, 0)
                                GROUP BY e.cnj
                                ORDER BY c DESC
                                LIMIT 5
                            """)
                        else:
                            cur.execute("""
                                SELECT e.cnj AS cnj, COUNT(*) AS c
                                FROM eventos_mov e
                                JOIN doc_process dp ON dp.cnj = e.cnj
                                GROUP BY e.cnj
                                ORDER BY c DESC
                                LIMIT 5
                            """)
                    top_cnj_alerts = [{"cnj": r["cnj"], "count": int(r["c"])} for r in cur.fetchall()]
            except Exception:
                top_cnj_alerts = []
# -----------------------
            # Métricas por doc (opcional)
            # -----------------------
            per_doc = None
            if doc_filter:
                doc = doc_filter
                processos_doc = 0
                movs_doc = 0
                alertas_doc = 0

                if table_exists("doc_process") and col_exists("doc_process", "doc") and col_exists("doc_process", "cnj"):
                    try:
                        processos_doc = _row_count(cur, "SELECT COUNT(DISTINCT cnj) FROM doc_process WHERE doc=?", (doc,))
                    except sqlite3.OperationalError:
                        processos_doc = 0

                    if table_exists("eventos_mov") and col_exists("eventos_mov", "cnj"):
                        try:
                            movs_doc = _row_count(cur, """
                                SELECT COUNT(*)
                                FROM eventos_mov e
                                JOIN doc_process dp ON dp.cnj = e.cnj
                                WHERE dp.doc=?
                            """, (doc,))
                        except sqlite3.OperationalError:
                            movs_doc = 0

                        if table_exists("alert_state") and col_exists("alert_state", "doc") and col_exists("alert_state", "last_event_id"):
                            try:
                                alertas_doc = _row_count(cur, """
                                    SELECT COUNT(*)
                                    FROM eventos_mov e
                                    JOIN doc_process dp ON dp.cnj = e.cnj
                                    LEFT JOIN alert_state a ON a.doc = dp.doc
                                    WHERE dp.doc=? AND e.id > COALESCE(a.last_event_id, 0)
                                """, (doc,))
                            except sqlite3.OperationalError:
                                alertas_doc = 0
                        else:
                            alertas_doc = movs_doc

                # Custos (estimados) por doc
                cost_doc_today = 0.0
                cost_doc_month = 0.0
                if table_exists("api_usage"):
                    try:
                        # hoje (UTC)
                        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                        cur.execute("SELECT COALESCE(SUM(cost_brl),0) FROM api_usage WHERE doc=? AND substr(ts,1,10)=?", (doc, today))
                        cost_doc_today = float(cur.fetchone()[0] or 0.0)
                        month = datetime.now(timezone.utc).strftime("%Y-%m")
                        cur.execute("SELECT COALESCE(SUM(cost_brl),0) FROM api_usage WHERE doc=? AND substr(ts,1,7)=?", (doc, month))
                        cost_doc_month = float(cur.fetchone()[0] or 0.0)
                    except Exception:
                        cost_doc_today, cost_doc_month = 0.0, 0.0

                
                # Custos (reais) por doc (importados do extrato)
                cost_doc_real_today = 0.0
                cost_doc_real_month = 0.0
                if table_exists("api_usage_real"):
                    try:
                        today_s = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                        cur.execute("SELECT COALESCE(SUM(cost_brl),0) FROM api_usage_real WHERE doc=? AND substr(ts,1,10)=?", (doc, today_s))
                        cost_doc_real_today = float(cur.fetchone()[0] or 0.0)
                        month_s = datetime.now(timezone.utc).strftime("%Y-%m")
                        cur.execute("SELECT COALESCE(SUM(cost_brl),0) FROM api_usage_real WHERE doc=? AND substr(ts,1,7)=?", (doc, month_s))
                        cost_doc_real_month = float(cur.fetchone()[0] or 0.0)
                    except Exception:
                        cost_doc_real_today, cost_doc_real_month = 0.0, 0.0

                cost_doc_delta_today = cost_doc_real_today - cost_doc_today
                cost_doc_delta_month = cost_doc_real_month - cost_doc_month

                per_doc = {
                    "doc": doc,
                    "processos": processos_doc,
                    "movs": movs_doc,
                    "alertas": alertas_doc,
                    "cost_today_brl": round(cost_doc_today, 2),
                    "cost_month_brl": round(cost_doc_month, 2),
                    "cost_real_today_brl": round(cost_doc_real_today, 2),
                    "cost_real_month_brl": round(cost_doc_real_month, 2),
                    "cost_delta_today_brl": round(cost_doc_delta_today, 2),
                    "cost_delta_month_brl": round(cost_doc_delta_month, 2),
                }

            # -----------------------
            # Tabela resumo por doc (leve)
            # -----------------------
            docs_summary: list[dict] = []
            if docs_list and table_exists("doc_process") and col_exists("doc_process", "doc") and col_exists("doc_process", "cnj"):
                for d in docs_list[:50]:
                    try:
                        pcount = _row_count(cur, "SELECT COUNT(DISTINCT cnj) FROM doc_process WHERE doc=?", (d,))
                    except Exception:
                        pcount = 0
                    acount = 0
                    if table_exists("eventos_mov") and col_exists("eventos_mov", "cnj"):
                        if table_exists("alert_state") and col_exists("alert_state", "doc") and col_exists("alert_state", "last_event_id"):
                            try:
                                acount = _row_count(cur, """
                                    SELECT COUNT(*)
                                    FROM eventos_mov e
                                    JOIN doc_process dp ON dp.cnj = e.cnj
                                    LEFT JOIN alert_state a ON a.doc = dp.doc
                                    WHERE dp.doc=? AND e.id > COALESCE(a.last_event_id, 0)
                                """, (d,))
                            except Exception:
                                acount = 0
                        else:
                            try:
                                acount = _row_count(cur, """
                                    SELECT COUNT(*)
                                    FROM eventos_mov e
                                    JOIN doc_process dp ON dp.cnj = e.cnj
                                    WHERE dp.doc=?
                                """, (d,))
                            except Exception:
                                acount = 0
                    docs_summary.append({"doc": d, "processos": pcount, "alertas": acount})

            
            # Top docs por alertas (para o dashboard)
            try:
                top_docs_alerts = sorted(
                    docs_summary,
                    key=lambda x: int(x.get("alertas") or 0),
                    reverse=True,
                )[:5]
            except Exception:
                top_docs_alerts = []
            return jsonify({
                "ok": True,
                "docs": docs_total,
                "processos": processos_total,
                "movs": movs_total,
                "alertas": alertas_total,
                "alertas_global": alertas_total,
                "cost_today_brl": round(cost_today_brl, 2),
                "cost_month_brl": round(cost_month_brl, 2),
                "cost_real_today_brl": round(cost_real_today_brl, 2),
                "cost_real_month_brl": round(cost_real_month_brl, 2),
                "cost_delta_today_brl": round(cost_delta_today_brl, 2),
                "cost_delta_month_brl": round(cost_delta_month_brl, 2),
                "top_cnj_alerts": top_cnj_alerts,
                "top_docs_alerts": top_docs_alerts,
                "docs_list": docs_list,
                "docs_summary": docs_summary,
                "per_doc": per_doc,
                "poll_state": _get_poll_state(),
                "discover_state": _get_discover_state(),
                "last_api_error": _get_last_api_error(),
            })
    except Exception:
        _safe_err("Dashboard: falha ao calcular métricas")
        return jsonify({"ok": False, "error": "Falha ao calcular métricas"}), 200


@app.get("/ui/api/costs/summary")
def ui_api_costs_summary():
    """Resumo de custos (estimado x real) para dashboard. Opcional: ?doc=<CPF/CNPJ>."""
    doc = request.args.get("doc")
    now = datetime.now(timezone.utc)
    day0 = datetime(now.year, now.month, now.day, tzinfo=timezone.utc).isoformat()
    mon0 = datetime(now.year, now.month, 1, tzinfo=timezone.utc).isoformat()

    def _sum(table: str, col: str) -> tuple[float, float]:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            # tabela pode não existir em bases antigas
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
            if cur.fetchone() is None:
                return 0.0, 0.0
            if doc:
                cur.execute(f"SELECT COALESCE(SUM({col}),0) AS s FROM {table} WHERE ts>=? AND doc=?", (day0, doc))
                today = float(cur.fetchone()["s"] or 0.0)
                cur.execute(f"SELECT COALESCE(SUM({col}),0) AS s FROM {table} WHERE ts>=? AND doc=?", (mon0, doc))
                month = float(cur.fetchone()["s"] or 0.0)
            else:
                cur.execute(f"SELECT COALESCE(SUM({col}),0) AS s FROM {table} WHERE ts>=?", (day0,))
                today = float(cur.fetchone()["s"] or 0.0)
                cur.execute(f"SELECT COALESCE(SUM({col}),0) AS s FROM {table} WHERE ts>=?", (mon0,))
                month = float(cur.fetchone()["s"] or 0.0)
            return today, month

    real_today, real_month = _sum("api_usage_real", "cost_brl")
    est_today, est_month = _sum("api_usage_est", "cost_brl")

    return jsonify({
        "ok": True,
        "doc": doc,
        "real": {"today": real_today, "month": real_month},
        "est": {"today": est_today, "month": est_month},
        "delta": {"today": real_today - est_today, "month": real_month - est_month},
    })


@app.get("/ui/api/costs/timeseries")
def ui_api_costs_timeseries():
    """Série diária de custos (real x est) para gráfico. Params: days=30, doc opcional."""
    doc = request.args.get("doc")
    try:
        days = int(request.args.get("days") or 30)
    except Exception:
        days = 30
    days = max(1, min(days, 365))
    # Agrupa por dia UTC (YYYY-MM-DD)
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        def _load(table: str, col: str) -> dict[str, float]:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
            if cur.fetchone() is None:
                return {}
            if doc:
                cur.execute(
                    f"""SELECT substr(ts,1,10) AS d, COALESCE(SUM({col}),0) AS s
                         FROM {table}
                         WHERE doc=? AND ts >= datetime('now','-{days} days')
                         GROUP BY d ORDER BY d""",
                    (doc,),
                )
            else:
                cur.execute(
                    f"""SELECT substr(ts,1,10) AS d, COALESCE(SUM({col}),0) AS s
                         FROM {table}
                         WHERE ts >= datetime('now','-{days} days')
                         GROUP BY d ORDER BY d"""
                )
            return {r["d"]: float(r["s"] or 0.0) for r in cur.fetchall()}

        real = _load("api_usage_real", "cost_brl")
        est = _load("api_usage_est", "cost_brl")

    # Monta eixo
    # pega união de datas e completa faltantes
    keys = sorted(set(real.keys()) | set(est.keys()))
    labels = keys[-days:] if len(keys) > days else keys
    return jsonify({
        "ok": True,
        "doc": doc,
        "labels": labels,
        "real": [real.get(d, 0.0) for d in labels],
        "est": [est.get(d, 0.0) for d in labels],
    })


@app.get("/ui/api/dashboard/timeseries")
def ui_api_dashboard_timeseries():
    """Retorna série temporal (últimos N dias) de movimentações salvas.
    Params:
      - doc (opcional): filtra por CPF/CNPJ (via doc_process)
      - days (opcional): default 14
    """
    try:
        doc = (request.args.get("doc") or "").strip()
        days = int(request.args.get("days") or "14")
        days = max(3, min(days, 90))

        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            # tabela existe?
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='eventos_mov'")
            if not cur.fetchone():
                return jsonify({"ok": True, "labels": [], "movs": []})

            params: list[Any] = []
            join = ""
            where = "WHERE 1=1"
            if doc:
                # só filtra se tiver doc_process
                cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='doc_process'")
                if cur.fetchone():
                    join = "JOIN doc_process dp ON dp.cnj = e.cnj"
                    where += " AND dp.doc=?"
                    params.append(doc)

            # janela
            where += " AND e.created_at >= datetime('now', ?)"
            params.append(f"-{days} days")

            q = f"""
                SELECT substr(e.created_at, 1, 10) AS d, COUNT(*) AS c
                FROM eventos_mov e
                {join}
                {where}
                GROUP BY substr(e.created_at, 1, 10)
                ORDER BY d ASC
            """
            cur.execute(q, tuple(params))
            rows = cur.fetchall()
            labels = [r["d"] for r in rows]
            movs = [int(r["c"]) for r in rows]

        return jsonify({"ok": True, "labels": labels, "movs": movs, "days": days, "doc": doc or None})
    except Exception as e:
        logger.exception("Dashboard: falha ao calcular timeseries")
        return jsonify({"ok": False, "error": str(e)}), 200


@app.get("/ui/api/alerts")
def ui_api_alerts_list():
    """Lista alertas pendentes (eventos_mov não-ACKed) para o painel do dashboard.
    Params (querystring):
      - doc: opcional (CPF/CNPJ)
      - limit: default 50 (max 500)
      - types: lista separada por vírgula (ex: SENTENCA,DECISAO,PENHORA)
      - must_value_penhora: 0/1 (quando 1, tenta filtrar por textos que indiquem valor/penhora/bloqueio)
      - q: termo livre (contém em texto)
    """
    doc = (request.args.get("doc") or "").strip()
    qterm = (request.args.get("q") or "").strip().lower()
    types_raw = (request.args.get("types") or "").strip()
    must_vp = (request.args.get("must_value_penhora") or "").strip() in ("1", "true", "True", "yes", "on")
    try:
        limit = int(request.args.get("limit") or "50")
    except Exception:
        limit = 50
    limit = max(1, min(limit, 500))

    types: list[str] = []
    if types_raw:
        types = [t.strip().upper() for t in types_raw.split(",") if t.strip()]

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        # valida tabelas mínimas
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='eventos_mov'")
        if not cur.fetchone():
            return jsonify({"ok": True, "items": [], "count": 0})

        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='doc_process'")
        if not cur.fetchone():
            return jsonify({"ok": True, "items": [], "count": 0})

        # alert_state é opcional (base antiga)
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='alert_state'")
        has_ack = cur.fetchone() is not None

        where = "WHERE 1=1"
        params: list[Any] = []

        if doc:
            where += " AND dp.doc=?"
            params.append(doc)

        if has_ack:
            where += " AND e.id > COALESCE(a.last_event_id, 0)"

        if types:
            placeholders = ",".join(["?"] * len(types))
            where += f" AND UPPER(COALESCE(e.tipo_inferido,'')) IN ({placeholders})"
            params.extend(types)

        if qterm:
            where += " AND LOWER(COALESCE(e.texto,'')) LIKE ?"
            params.append(f"%{qterm}%")

        if must_vp:
            # Heurística simples (SQLite sem REGEXP padrão): palavras-chave e presença de R$
            where += " AND (LOWER(COALESCE(e.texto,'')) LIKE '%penhor%' OR LOWER(COALESCE(e.texto,'')) LIKE '%bloque%' OR LOWER(COALESCE(e.texto,'')) LIKE '%bacen%' OR LOWER(COALESCE(e.texto,'')) LIKE '%sisba%' OR LOWER(COALESCE(e.texto,'')) LIKE '%r$%' OR LOWER(COALESCE(e.texto,'')) LIKE '%valor%' OR UPPER(COALESCE(e.tipo_inferido,''))='PENHORA')"

        join_ack = "LEFT JOIN alert_state a ON a.doc = dp.doc" if has_ack else ""

        sql = f"""
            SELECT
              dp.doc AS doc,
              e.cnj AS cnj,
              e.id AS event_id,
              e.data AS data,
              e.created_at AS created_at,
              e.tipo AS tipo,
              e.tipo_inferido AS tipo_inferido,
              e.texto AS texto
            FROM eventos_mov e
            JOIN doc_process dp ON dp.cnj = e.cnj
            {join_ack}
            {where}
            ORDER BY e.id DESC
            LIMIT ?
        """
        cur.execute(sql, (*params, limit))
        items = [dict(r) for r in cur.fetchall()]

    return jsonify({"ok": True, "items": items, "count": len(items)})


@app.get("/ui/api/alerts/export.csv")
def ui_api_alerts_export_csv():
    """Exporta os alertas do painel em CSV (mesmos filtros do /ui/api/alerts)."""
    # reaproveita a listagem (sem duplicar SQL)
    with app.test_request_context():
        pass  # placeholder

    # chama a função de listagem manualmente (sem HTTP)
    # (duplicação mínima para manter simples e robusto)
    doc = (request.args.get("doc") or "").strip()
    qterm = (request.args.get("q") or "").strip().lower()
    types_raw = (request.args.get("types") or "").strip()
    must_vp = (request.args.get("must_value_penhora") or "").strip() in ("1", "true", "True", "yes", "on")
    try:
        limit = int(request.args.get("limit") or "500")
    except Exception:
        limit = 500
    limit = max(1, min(limit, 2000))

    types: list[str] = []
    if types_raw:
        types = [t.strip().upper() for t in types_raw.split(",") if t.strip()]

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='eventos_mov'")
        if not cur.fetchone():
            out = "doc,cnj,event_id,data,created_at,tipo,tipo_inferido,texto\n"
            return Response(out, mimetype="text/csv; charset=utf-8")

        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='doc_process'")
        if not cur.fetchone():
            out = "doc,cnj,event_id,data,created_at,tipo,tipo_inferido,texto\n"
            return Response(out, mimetype="text/csv; charset=utf-8")

        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='alert_state'")
        has_ack = cur.fetchone() is not None

        where = "WHERE 1=1"
        params: list[Any] = []

        if doc:
            where += " AND dp.doc=?"
            params.append(doc)

        if has_ack:
            where += " AND e.id > COALESCE(a.last_event_id, 0)"

        if types:
            placeholders = ",".join(["?"] * len(types))
            where += f" AND UPPER(COALESCE(e.tipo_inferido,'')) IN ({placeholders})"
            params.extend(types)

        if qterm:
            where += " AND LOWER(COALESCE(e.texto,'')) LIKE ?"
            params.append(f"%{qterm}%")

        if must_vp:
            where += " AND (LOWER(COALESCE(e.texto,'')) LIKE '%penhor%' OR LOWER(COALESCE(e.texto,'')) LIKE '%bloque%' OR LOWER(COALESCE(e.texto,'')) LIKE '%bacen%' OR LOWER(COALESCE(e.texto,'')) LIKE '%sisba%' OR LOWER(COALESCE(e.texto,'')) LIKE '%r$%' OR LOWER(COALESCE(e.texto,'')) LIKE '%valor%' OR UPPER(COALESCE(e.tipo_inferido,''))='PENHORA')"

        join_ack = "LEFT JOIN alert_state a ON a.doc = dp.doc" if has_ack else ""
        sql = f"""
            SELECT
              dp.doc AS doc,
              e.cnj AS cnj,
              e.id AS event_id,
              e.data AS data,
              e.created_at AS created_at,
              e.tipo AS tipo,
              e.tipo_inferido AS tipo_inferido,
              e.texto AS texto
            FROM eventos_mov e
            JOIN doc_process dp ON dp.cnj = e.cnj
            {join_ack}
            {where}
            ORDER BY e.id DESC
            LIMIT ?
        """
        cur.execute(sql, (*params, limit))
        rows = cur.fetchall()

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["doc","cnj","event_id","data","created_at","tipo","tipo_inferido","texto"])
    for r in rows:
        w.writerow([
            r["doc"] or "",
            r["cnj"] or "",
            r["event_id"] or "",
            (r["data"] or ""),
            (r["created_at"] or ""),
            (r["tipo"] or ""),
            (r["tipo_inferido"] or ""),
            (r["texto"] or ""),
        ])
    out = buf.getvalue()
    return Response(out, mimetype="text/csv; charset=utf-8")


def _table_exists(conn: sqlite3.Connection, name: str) -> bool:
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1", (name,))
    return cur.fetchone() is not None


def _list_local_monitoramentos() -> list[dict]:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT
                p.cnj,
                p.created_at,
                p.last_sync_at,
                p.last_event_at,
                COUNT(DISTINCT dp.doc) AS links_count,
                GROUP_CONCAT(DISTINCT dp.doc) AS docs_csv,
                COUNT(DISTINCT ev.id) AS eventos_count
            FROM processos p
            LEFT JOIN doc_process dp ON dp.cnj = p.cnj
            LEFT JOIN eventos_mov ev ON ev.cnj = p.cnj
            GROUP BY p.cnj, p.created_at, p.last_sync_at, p.last_event_at
            ORDER BY COALESCE(p.last_event_at, p.last_sync_at, p.created_at) DESC, p.cnj ASC
        """)
        rows = []
        for r in cur.fetchall():
            item = dict(r)
            docs = [d for d in (item.get("docs_csv") or "").split(",") if d]
            item["docs"] = docs
            item["is_orfao_local"] = len(docs) == 0
            rows.append(item)
        return rows


def _list_watchlist_local() -> list[dict]:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT
                w.id, w.doc, w.tipo_doc, w.created_at,
                COUNT(DISTINCT dp.cnj) AS processos_count
            FROM watchlist w
            LEFT JOIN doc_process dp ON dp.doc = w.doc
            GROUP BY w.id, w.doc, w.tipo_doc, w.created_at
            ORDER BY w.id DESC
        """)
        return [dict(r) for r in cur.fetchall()]


def _cleanup_local_processo(cnj: str) -> dict:
    cnj = (cnj or "").strip()
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM eventos_mov WHERE cnj=?", (cnj,))
        eventos = cur.rowcount or 0
        cur.execute("DELETE FROM documentos_cache WHERE cnj=?", (cnj,))
        docs_cache = cur.rowcount or 0
        cur.execute("DELETE FROM processo_updates WHERE cnj=?", (cnj,))
        updates = cur.rowcount or 0
        cur.execute("DELETE FROM doc_process WHERE cnj=?", (cnj,))
        links = cur.rowcount or 0
        cur.execute("DELETE FROM processos WHERE cnj=?", (cnj,))
        processos = cur.rowcount or 0
        conn.commit()
    return {"cnj": cnj, "deleted_processos": processos, "deleted_links": links, "deleted_eventos": eventos, "deleted_docs_cache": docs_cache, "deleted_updates": updates}


def _cleanup_local_doc(doc: str) -> dict:
    doc = normalize_doc(doc)
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT cnj FROM doc_process WHERE doc=?", (doc,))
        cnjs = [r[0] for r in cur.fetchall()]
        cur.execute("DELETE FROM doc_process WHERE doc=?", (doc,))
        links = cur.rowcount or 0
        cur.execute("DELETE FROM watchlist WHERE doc=?", (doc,))
        watch = cur.rowcount or 0
        deleted_orfaos = 0
        for cnj in cnjs:
            cur.execute("SELECT 1 FROM doc_process WHERE cnj=? LIMIT 1", (cnj,))
            if cur.fetchone() is None:
                cur.execute("DELETE FROM eventos_mov WHERE cnj=?", (cnj,))
                cur.execute("DELETE FROM documentos_cache WHERE cnj=?", (cnj,))
                cur.execute("DELETE FROM processo_updates WHERE cnj=?", (cnj,))
                cur.execute("DELETE FROM processos WHERE cnj=?", (cnj,))
                deleted_orfaos += cur.rowcount or 0
        conn.commit()
    return {"doc": doc, "deleted_watchlist": watch, "deleted_links": links, "deleted_orphan_processes": deleted_orfaos}


def _cleanup_local_orfaos() -> dict:
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT p.cnj FROM processos p LEFT JOIN doc_process dp ON dp.cnj = p.cnj WHERE dp.cnj IS NULL")
        cnjs = [r[0] for r in cur.fetchall()]
        total = 0
        for cnj in cnjs:
            cur.execute("DELETE FROM eventos_mov WHERE cnj=?", (cnj,))
            cur.execute("DELETE FROM documentos_cache WHERE cnj=?", (cnj,))
            cur.execute("DELETE FROM processo_updates WHERE cnj=?", (cnj,))
            cur.execute("DELETE FROM processos WHERE cnj=?", (cnj,))
            total += cur.rowcount or 0
        conn.commit()
    return {"deleted_orphan_processes": total, "cnjs": cnjs}


def _extract_remote_monitor_rows(payload: Any) -> list[dict]:
    rows = extract_list(payload)
    if not rows and isinstance(payload, dict):
        for key in ("monitoramentos", "data", "items", "results"):
            val = payload.get(key)
            if isinstance(val, list):
                rows = val
                break
    out = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        monitor_id = item.get("id")
        cnj = (item.get("numero") or item.get("numero_cnj") or item.get("cnj") or "").strip()
        out.append({"id": monitor_id, "cnj": cnj, "raw": item})
    return out


def _find_remote_monitor_ids_by_cnj(cnj: str) -> list[int]:
    require_token_configured()
    payload = client.listar_monitoramentos_processos(limit=100, page=1)  # type: ignore
    rows = _extract_remote_monitor_rows(payload)
    wanted = (cnj or "").strip()
    out = []
    for r in rows:
        if str(r.get("cnj") or "").strip() == wanted and r.get("id") not in (None, ""):
            try:
                out.append(int(r["id"]))
            except Exception:
                pass
    return out

@app.get("/admin/monitoramentos/local")
def admin_monitoramentos_local_json():
    return jsonify({"ok": True, "items": _list_local_monitoramentos(), "watchlist": _list_watchlist_local()})


@app.post("/admin/monitoramentos/local/processo/cleanup")
def admin_monitoramentos_local_cleanup_processo():
    payload = request.get_json(force=True, silent=True) or {}
    cnj = (payload.get("cnj") or request.form.get("cnj") or "").strip()
    if not cnj:
        return jsonify({"ok": False, "error": "CNJ é obrigatório."}), 400
    return jsonify({"ok": True, **_cleanup_local_processo(cnj)})


@app.post("/admin/monitoramentos/local/doc/cleanup")
def admin_monitoramentos_local_cleanup_doc():
    payload = request.get_json(force=True, silent=True) or {}
    doc = (payload.get("doc") or request.form.get("doc") or "").strip()
    if not doc:
        return jsonify({"ok": False, "error": "Doc é obrigatório."}), 400
    return jsonify({"ok": True, **_cleanup_local_doc(doc)})


@app.post("/admin/monitoramentos/local/orfaos/cleanup")
def admin_monitoramentos_local_cleanup_orfaos():
    return jsonify({"ok": True, **_cleanup_local_orfaos()})


@app.get("/admin/monitoramentos/remote")
def admin_monitoramentos_remote_list():
    require_token_configured()
    try:
        payload = client.listar_monitoramentos_processos(limit=100, page=1)  # type: ignore
        return jsonify({"ok": True, "items": _extract_remote_monitor_rows(payload), "raw": payload})
    except EscavadorUnauthorized as e:
        return jsonify({"ok": False, "error": f"Token sem autorização para listar monitoramentos: {e}"}), 403
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.post("/admin/monitoramentos/remote/remove")
def admin_monitoramentos_remote_remove():
    require_token_configured()
    payload = request.get_json(force=True, silent=True) or {}
    mid = payload.get("id") or request.form.get("id")
    if mid in (None, ""):
        return jsonify({"ok": False, "error": "ID do monitoramento é obrigatório."}), 400
    try:
        client.remover_monitoramento_processo(int(mid))  # type: ignore
        return jsonify({"ok": True, "removed_id": int(mid)})
    except EscavadorUnauthorized as e:
        return jsonify({"ok": False, "error": f"Token sem autorização para remover monitoramentos: {e}"}), 403
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500



@app.post("/admin/monitoramentos/local/processo/remove-completo")
def admin_monitoramentos_local_remove_completo():
    payload = request.get_json(force=True, silent=True) or {}
    cnj = (payload.get("cnj") or request.form.get("cnj") or "").strip()
    if not cnj:
        return jsonify({"ok": False, "error": "CNJ é obrigatório."}), 400

    remote_removed = []
    remote_error = None
    try:
        ids = _find_remote_monitor_ids_by_cnj(cnj)
        for mid in ids:
            try:
                client.remover_monitoramento_processo(int(mid))  # type: ignore
                remote_removed.append(int(mid))
            except Exception as ex:
                remote_error = str(ex)
                break
    except Exception as ex:
        remote_error = str(ex)

    local = _cleanup_local_processo(cnj)
    return jsonify({
        "ok": True,
        "cnj": cnj,
        "remote_removed_ids": remote_removed,
        "remote_error": remote_error,
        **local
    })

@app.post("/admin/monitoramentos/local/monitorar-cnj")
def admin_monitoramentos_local_monitorar_cnj():
    require_token_configured()
    payload = request.get_json(force=True, silent=False) or {}
    cnj_in = (payload.get("cnj") or "").strip()
    if not cnj_in:
        return jsonify({"ok": False, "error": "CNJ é obrigatório."}), 400
    try:
        cnj = normalize_cnj(cnj_in)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    escavador_status = "created"
    escavador_detail = None
    try:
        client.criar_monitor_processo(cnj)  # type: ignore
    except EscavadorAlreadyMonitored as e:
        escavador_status = "already_monitored"
        escavador_detail = str(e)
    except EscavadorUnauthorized as e:
        return jsonify({"ok": False, "error": f"Token sem autorização para monitorar o CNJ: {e}"}), 403
    except requests.HTTPError as e:
        resp = getattr(e, "response", None)
        detail = None
        try:
            detail = resp.text if resp is not None else str(e)
        except Exception:
            detail = str(e)
        return jsonify({"ok": False, "error": "Falha ao criar monitoramento no Escavador.", "detail": detail}), 502

    with sqlite3.connect(DB_PATH) as conn:
        upsert_processo(conn, cnj)
        cur = conn.cursor()
        cur.execute("UPDATE processos SET last_sync_at=COALESCE(last_sync_at, ? ) WHERE cnj=?", (utcnow_iso(), cnj))
        conn.commit()

    synced_events = None
    sync_error = None
    try:
        st = sync_process_movements(client, cnj, limit=100)  # type: ignore
        synced_events = st.new_events
    except Exception as e:
        sync_error = str(e)
        logger.warning("Monitorar CNJ %s: não foi possível sincronizar movimentações iniciais: %s", cnj, sync_error)

    return jsonify({
        "ok": True,
        "cnj": cnj,
        "escavador_status": escavador_status,
        "escavador_detail": escavador_detail,
        "synced_events": synced_events,
        "sync_error": sync_error,
    })


@app.get("/ui/admin/monitoramentos")
def ui_admin_monitoramentos():
    local_items = _list_local_monitoramentos()
    watch_items = _list_watchlist_local()
    orphan_items = [it for it in local_items if it.get("is_orfao_local")]

    total_local = len(local_items)
    total_orfaos = len(orphan_items)
    total_watch = len(watch_items)
    total_vinculados = max(total_local - total_orfaos, 0)

    local_rows = ""
    orphan_rows = ""
    for item in local_items:
        badge = "<span class='badge text-bg-warning'>Órfão local</span>" if item.get("is_orfao_local") else "<span class='badge text-bg-success'>Vinculado</span>"
        docs = ", ".join(item.get("docs") or []) or "—"
        search_text = esc((((item.get("cnj") or "") + " " + docs).lower()))
        detail_href = f"/ui/processo/{quote(str(item.get('cnj') or ''))}"
        local_row = f"""
        <tr data-search='{search_text}'>
          <td class='mono small'><a href='{detail_href}'>{esc(item.get('cnj'))}</a></td>
          <td>{badge}</td>
          <td class='small'>{esc(docs)}</td>
          <td class='text-end mono small'>{int(item.get('links_count') or 0)}</td>
          <td class='text-end mono small'>{int(item.get('eventos_count') or 0)}</td>
          <td class='mono small'>{esc((item.get('last_sync_at') or item.get('created_at') or '')[:19]).replace('T',' ')}</td>
          <td class='text-end'>
            <div class='btn-group btn-group-sm flex-wrap justify-content-end'>
              <a class='btn btn-outline-primary' href='{detail_href}'>Abrir</a>
              <button class='btn btn-outline-danger' onclick="removeFull('{esc(item.get('cnj'))}')">Excluir monitoramento</button>
              <button class='btn btn-outline-secondary' onclick="cleanupProcess('{esc(item.get('cnj'))}')">Só limpar local</button>
            </div>
          </td>
        </tr>
        """
        local_rows += local_row
        if item.get("is_orfao_local"):
            orphan_rows += local_row
    if not local_rows:
        local_rows = "<tr><td colspan='7' class='muted'>Nenhum processo local encontrado.</td></tr>"
    if not orphan_rows:
        orphan_rows = "<tr><td colspan='7' class='muted'>Nenhum órfão local encontrado.</td></tr>"

    watch_rows = ""
    for item in watch_items:
        doc = str(item.get("doc") or "")
        watch_rows += f"""
        <tr>
          <td class='mono small'>{esc(doc)}</td>
          <td>{esc(item.get('tipo_doc') or '')}</td>
          <td class='text-end mono small'>{int(item.get('processos_count') or 0)}</td>
          <td class='mono small'>{esc((item.get('created_at') or '')[:19]).replace('T',' ')}</td>
          <td class='text-end'>
            <div class='btn-group btn-group-sm'>
              <button class='btn btn-outline-primary' onclick="filterByDoc('{esc(doc)}')">Ver processos</button>
              <button class='btn btn-outline-danger' onclick="cleanupDoc('{esc(doc)}')">Excluir watchlist</button>
            </div>
          </td>
        </tr>
        """
    if not watch_rows:
        watch_rows = "<tr><td colspan='5' class='muted'>Watchlist vazia.</td></tr>"

    body = f"""
    <div class='d-flex align-items-center justify-content-between mb-3'>
      <div>
        <div class='h4 mb-0'>Admin / Monitoramentos</div>
        <div class='muted small'>Manutenção local e remota dos monitoramentos, com foco em exclusão, órfãos e vínculos.</div>
      </div>
      <div class='d-flex gap-2 flex-wrap'>
        <a class='btn btn-outline-secondary' href='/ui'>Dashboard</a>
        <a class='btn btn-outline-secondary' href='/ui/watchlist'>CNPJs / CPFs</a>
        <a class='btn btn-outline-secondary' href='/ui/financeiro'>Financeiro</a>
        <button class='btn btn-outline-danger' onclick='cleanupOrphans()'>Limpar órfãos locais</button>
        <button class='btn btn-outline-primary' onclick='loadRemote()'>Carregar remoto</button>
      </div>
    </div>

    <div id='adminMsg' class='mb-3'></div>

    <div class='row g-3 mb-3'>
      <div class='col-lg-3 col-md-6'><div class='card p-3'><div class='muted small'>Processos locais</div><div class='h3 m-0'>{total_local}</div></div></div>
      <div class='col-lg-3 col-md-6'><div class='card p-3'><div class='muted small'>Vinculados</div><div class='h3 m-0'>{total_vinculados}</div></div></div>
      <div class='col-lg-3 col-md-6'><div class='card p-3'><div class='muted small'>Órfãos locais</div><div class='h3 m-0'>{total_orfaos}</div></div></div>
      <div class='col-lg-3 col-md-6'><div class='card p-3'><div class='muted small'>Docs monitorados</div><div class='h3 m-0'>{total_watch}</div></div></div>
    </div>

    <div class='card p-3 mb-3'>
      <div class='d-flex flex-wrap gap-2 align-items-center justify-content-between'>
        <div>
          <div class='h6 mb-1'>Ações rápidas</div>
          <div class='muted small'>CNJ direto, filtros e navegação rápida.</div>
        </div>
        <div class='d-flex gap-2 flex-wrap'>
          <button class='btn btn-outline-secondary active' data-tab-btn='tab-processos' onclick="showAdminTab('tab-processos', this)">Processos</button>
          <button class='btn btn-outline-secondary' data-tab-btn='tab-orfaos' onclick="showAdminTab('tab-orfaos', this)">Órfãos</button>
          <button class='btn btn-outline-secondary' data-tab-btn='tab-docs' onclick="showAdminTab('tab-docs', this)">CNPJs / CPFs</button>
          <button class='btn btn-outline-secondary' data-tab-btn='tab-remoto' onclick="showAdminTab('tab-remoto', this)">Remoto</button>
        </div>
      </div>
      <div class='row g-3 align-items-end mt-1'>
        <div class='col-xl-4'>
          <label class='form-label small muted'>Monitorar CNJ diretamente</label>
          <input id='cnjDirectInput' class='form-control mono' placeholder='1018484-34.2015.8.26.0224'>
        </div>
        <div class='col-xl-2'>
          <button class='btn btn-primary w-100' onclick='monitorDirectCnj()'>Monitorar CNJ</button>
        </div>
        <div class='col-xl-3'>
          <label class='form-label small muted'>Filtro de processos</label>
          <input id='localFilterInput' class='form-control' placeholder='Digite CNJ ou documento vinculado'>
        </div>
        <div class='col-xl-3'>
          <label class='form-label small muted'>Consulta rápida</label>
          <div class='d-flex gap-2'>
            <input id='quickCnjInput' class='form-control mono' placeholder='CNJ'>
            <button class='btn btn-outline-primary' onclick='openQuickProcess()'>Abrir</button>
          </div>
        </div>
      </div>
    </div>

    <div id='tab-processos' class='admin-tab'>
      <div class='card p-3 mb-3'>
        <div class='d-flex align-items-center justify-content-between mb-2'>
          <div class='h6 mb-0'>Processos monitorados</div>
          <div class='muted small'>Excluir monitoramento tenta remover remoto por CNJ e depois limpa o local.</div>
        </div>
        <div class='table-responsive'>
          <table class='table table-sm align-middle'>
            <thead><tr class='muted small'><th>CNJ</th><th>Status</th><th>Docs vinculados</th><th class='text-end'>Links</th><th class='text-end'>Eventos</th><th>Última atividade</th><th class='text-end'>Ações</th></tr></thead>
            <tbody id='localRows'>{local_rows}</tbody>
          </table>
        </div>
      </div>
    </div>

    <div id='tab-orfaos' class='admin-tab' style='display:none'>
      <div class='card p-3 mb-3'>
        <div class='d-flex align-items-center justify-content-between mb-2'>
          <div class='h6 mb-0'>Órfãos locais</div>
          <div class='muted small'>Processos sem vínculo com CNPJ/CPF monitorado.</div>
        </div>
        <div class='table-responsive'>
          <table class='table table-sm align-middle'>
            <thead><tr class='muted small'><th>CNJ</th><th>Status</th><th>Docs vinculados</th><th class='text-end'>Links</th><th class='text-end'>Eventos</th><th>Última atividade</th><th class='text-end'>Ações</th></tr></thead>
            <tbody id='orphanRows'>{orphan_rows}</tbody>
          </table>
        </div>
      </div>
    </div>

    <div id='tab-docs' class='admin-tab' style='display:none'>
      <div class='card p-3 mb-3'>
        <div class='d-flex align-items-center justify-content-between mb-2'>
          <div class='h6 mb-0'>CNPJs / CPFs monitorados</div>
          <div class='muted small'>Manutenção da watchlist local.</div>
        </div>
        <div class='table-responsive'>
          <table class='table table-sm align-middle'>
            <thead><tr class='muted small'><th>Doc</th><th>Tipo</th><th class='text-end'>Processos</th><th>Criado em</th><th class='text-end'>Ações</th></tr></thead>
            <tbody>{watch_rows}</tbody>
          </table>
        </div>
      </div>
    </div>

    <div id='tab-remoto' class='admin-tab' style='display:none'>
      <div class='card p-3'>
        <div class='d-flex align-items-center justify-content-between mb-2'>
          <div class='h6 mb-0'>Monitoramentos remotos no Escavador</div>
          <div class='muted small'>Listagem e exclusão remota por ID.</div>
        </div>
        <div class='table-responsive'>
          <table class='table table-sm align-middle'>
            <thead><tr class='muted small'><th>ID remoto</th><th>CNJ</th><th class='text-end'>Ações</th></tr></thead>
            <tbody id='remoteRows'><tr><td colspan='3' class='muted'>Clique em <b>Carregar remoto</b> para consultar o Escavador.</td></tr></tbody>
          </table>
        </div>
      </div>
    </div>

    <script>
    function msg(kind, html){{
      const el = document.getElementById('adminMsg');
      const cls = kind==='ok' ? 'alert alert-success' : (kind==='warn' ? 'alert alert-warning' : 'alert alert-danger');
      el.innerHTML = `<div class="${{cls}}">${{html}}</div>`;
    }}

    function showAdminTab(id, btn){{
      document.querySelectorAll('.admin-tab').forEach(el => el.style.display = (el.id===id ? '' : 'none'));
      document.querySelectorAll('[data-tab-btn]').forEach(b => b.classList.remove('active'));
      if(btn) btn.classList.add('active');
      if(id === 'tab-remoto') loadRemote();
    }}

    async function postJson(url, payload){{
      const r = await fetch(url, {{method:'POST', headers:{{'Content-Type':'application/json'}}, body: JSON.stringify(payload||{{}})}});
      const j = await r.json();
      if(!r.ok || !j.ok) throw new Error(j.error || `Falha em ${{url}}`);
      return j;
    }}

    async function cleanupProcess(cnj){{
      if(!confirm(`Limpar apenas localmente o processo ${{cnj}}?`)) return;
      try {{
        await postJson('/admin/monitoramentos/local/processo/cleanup', {{cnj}});
        msg('ok', `Processo <b>${{cnj}}</b> limpo localmente.`);
        setTimeout(()=>location.reload(), 300);
      }} catch(ex) {{ msg('err', ex.message); }}
    }}

    async function removeFull(cnj){{
      if(!confirm(`Excluir o monitoramento do CNJ ${{cnj}}? Vou tentar remover no Escavador e depois limpar o local.`)) return;
      try {{
        const j = await postJson('/admin/monitoramentos/local/processo/remove-completo', {{cnj}});
        let extra = '';
        if((j.remote_removed_ids || []).length) extra += ` Remotos removidos: <b>${{j.remote_removed_ids.join(', ')}}</b>.`;
        if(j.remote_error) extra += ` <span class="text-warning">Remoto: ${{j.remote_error}}</span>`;
        msg('ok', `Monitoramento de <b>${{cnj}}</b> excluído.${{extra}}`);
        setTimeout(()=>location.reload(), 500);
      }} catch(ex) {{ msg('err', ex.message); }}
    }}

    async function cleanupDoc(doc){{
      if(!confirm(`Excluir a watchlist local de ${{doc}}?`)) return;
      try {{
        await postJson('/admin/monitoramentos/local/doc/cleanup', {{doc}});
        msg('ok', `Watchlist <b>${{doc}}</b> excluída localmente.`);
        setTimeout(()=>location.reload(), 300);
      }} catch(ex) {{ msg('err', ex.message); }}
    }}

    async function cleanupOrphans(){{
      if(!confirm('Remover todos os processos órfãos locais?')) return;
      try {{
        const j = await postJson('/admin/monitoramentos/local/orfaos/cleanup', {{}});
        msg('ok', `Órfãos removidos: <b>${{j.deleted_orphan_processes || 0}}</b>.`);
        setTimeout(()=>location.reload(), 300);
      }} catch(ex) {{ msg('err', ex.message); }}
    }}

    async function loadRemote(){{
      const tbody = document.getElementById('remoteRows');
      if(!tbody) return;
      tbody.innerHTML = `<tr><td colspan='3' class='muted'>Consultando Escavador...</td></tr>`;
      try {{
        const r = await fetch('/admin/monitoramentos/remote');
        const j = await r.json();
        if(!r.ok || !j.ok) throw new Error(j.error || 'Falha ao consultar remoto');
        const items = j.items || [];
        if(!items.length) {{
          tbody.innerHTML = `<tr><td colspan='3' class='muted'>Nenhum monitoramento remoto retornado.</td></tr>`;
          return;
        }}
        tbody.innerHTML = items.map(it => `
          <tr>
            <td class='mono small'>${{it.id ?? ''}}</td>
            <td class='mono small'>${{it.cnj || '—'}}</td>
            <td class='text-end'><button class='btn btn-sm btn-outline-danger' onclick='removeRemote(${{it.id}}, "${{(it.cnj||'').replace(/"/g,'&quot;')}}")'>Excluir remoto</button></td>
          </tr>`).join('');
      }} catch(ex) {{
        tbody.innerHTML = `<tr><td colspan='3' class='text-danger'>${{ex.message}}</td></tr>`;
      }}
    }}

    async function removeRemote(id, cnj){{
      if(!confirm(`Excluir o monitoramento remoto ID ${{id}}${{cnj ? ' do CNJ ' + cnj : ''}}?`)) return;
      try {{
        await postJson('/admin/monitoramentos/remote/remove', {{id}});
        msg('ok', `Monitoramento remoto <b>${{id}}</b> removido.`);
        loadRemote();
      }} catch(ex) {{ msg('err', ex.message); }}
    }}

    function getQuickCnj(){{
      const v = (document.getElementById('quickCnjInput')?.value || document.getElementById('cnjDirectInput')?.value || '').trim();
      if(!v){{ msg('warn', 'Informe um CNJ para consultar.'); return ''; }}
      return v;
    }}

    function openQuickProcess(){{
      const cnj = getQuickCnj();
      if(!cnj) return;
      window.location.href = '/ui/processo/' + encodeURIComponent(cnj);
    }}

    function filterByDoc(doc){{
      const f = document.getElementById('localFilterInput');
      if(f) f.value = doc || '';
      showAdminTab('tab-processos', document.querySelector('[data-tab-btn="tab-processos"]'));
      applyLocalFilter();
    }}

    async function monitorDirectCnj(){{
      const el = document.getElementById('cnjDirectInput');
      const cnj = (el.value || '').trim();
      if(!cnj){{ msg('warn', 'Informe um CNJ para monitorar.'); return; }}
      try {{
        const j = await postJson('/admin/monitoramentos/local/monitorar-cnj', {{cnj}});
        let extra = '';
        if (j.escavador_status === 'already_monitored') {{
          extra = ' O Escavador já monitorava esse CNJ.';
        }} else if (j.escavador_status === 'created') {{
          extra = ' Monitoramento criado no Escavador.';
        }}
        const detailLink = '/ui/processo/' + encodeURIComponent(j.cnj || cnj);
        msg('ok', `CNJ <b>${{j.cnj || cnj}}</b> monitorado.${{extra}} <a href="${{detailLink}}">Abrir processo</a>`);
        document.getElementById('quickCnjInput').value = j.cnj || cnj;
        el.value = '';
        setTimeout(()=>location.reload(), 900);
      }} catch(ex) {{ msg('err', ex.message); }}
    }}

    function applyLocalFilter(){{
      const q = (document.getElementById('localFilterInput').value || '').trim().toLowerCase();
      const rows = Array.from(document.querySelectorAll('#localRows tr, #orphanRows tr'));
      let visible = 0;
      rows.forEach(tr => {{
        const hay = (tr.getAttribute('data-search') || tr.innerText || '').toLowerCase();
        const show = !q || hay.includes(q);
        tr.style.display = show ? '' : 'none';
        if(show) visible++;
      }});
    }}

    document.getElementById('localFilterInput')?.addEventListener('input', applyLocalFilter);
    </script>
    """
    return render_template_string(UI_BASE, body=body)


@app.get("/ui/dashboard")
def ui_dashboard():
    return redirect("/ui#dashboard")


@app.get("/ui/financeiro")
def ui_financeiro():
    return redirect("/ui/costs")




@app.get("/ui/costs")
def ui_costs():
    # Página para importar extrato (XLSX) e visualizar custos reais vs estimados
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        def table_exists(name: str) -> bool:
            cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1", (name,))
            return cur.fetchone() is not None

        last_real = []
        if table_exists("api_usage_real"):
            try:
                cur.execute("SELECT ts, imported_at, source, method, endpoint, doc, cnj, cost_brl FROM api_usage_real ORDER BY ts DESC LIMIT 25")
                last_real = [dict(r) for r in cur.fetchall()]
            except Exception:
                last_real = []

        finance_proc = []
        if table_exists("processos"):
            try:
                proc_cols = [r[1] for r in cur.execute("PRAGMA table_info(processos)").fetchall()]
                if "valor_causa" in proc_cols:
                    sel = "SELECT cnj, valor_causa, updated_at, classe FROM processos WHERE COALESCE(valor_causa,'')<>'' ORDER BY updated_at DESC LIMIT 200"
                    for r in cur.execute(sel).fetchall():
                        raw = str(r["valor_causa"] or "").strip()
                        if not raw:
                            continue
                        m = re.search(r'(\d[\d\.,]*)', raw)
                        if not m:
                            continue
                        try:
                            val = float(m.group(1).replace('.', '').replace(',', '.'))
                        except Exception:
                            continue
                        finance_proc.append({
                            "cnj": r["cnj"],
                            "tipo": "Valor da causa",
                            "valor": val,
                            "origem": "capa",
                            "data": str(r["updated_at"] or "")[:10],
                            "descricao": str(r["classe"] or "Valor extraído da capa"),
                        })
                if table_exists("eventos_mov"):
                    em_cols = [r[1] for r in cur.execute("PRAGMA table_info(eventos_mov)").fetchall()]
                    text_col = "texto" if "texto" in em_cols else ("conteudo" if "conteudo" in em_cols else None)
                    date_col = "data" if "data" in em_cols else ("created_at" if "created_at" in em_cols else None)
                    if text_col:
                        money_rx = re.compile(r'R\$\s*([0-9\.\,]+)')
                        kind_rx = re.compile(r'(custas|honor[áa]rios|multa|indeniza[cç][ãa]o|valor da causa|penhora|bloqueio)', re.I)
                        sql = f"SELECT cnj, {date_col or 'NULL'} as datax, {text_col} as textox FROM eventos_mov WHERE LOWER(COALESCE({text_col},'')) LIKE '%r$%' ORDER BY COALESCE({date_col or 'NULL'},'') DESC LIMIT 500"
                        seen = set((it["cnj"], it["tipo"], round(float(it["valor"]),2), it["data"]) for it in finance_proc)
                        for r in cur.execute(sql).fetchall():
                            txt = str(r["textox"] or "")
                            mm = money_rx.search(txt)
                            if not mm:
                                continue
                            try:
                                val = float(mm.group(1).replace('.', '').replace(',', '.'))
                            except Exception:
                                continue
                            km = kind_rx.search(txt)
                            kind = km.group(1).strip().title() if km else "Valor identificado"
                            datax = str(r["datax"] or "")[:10]
                            key = (r["cnj"], kind, round(val,2), datax)
                            if key in seen:
                                continue
                            seen.add(key)
                            finance_proc.append({
                                "cnj": r["cnj"],
                                "tipo": kind,
                                "valor": val,
                                "origem": "movimentação",
                                "data": datax,
                                "descricao": (txt[:160] + "…") if len(txt) > 160 else txt,
                            })
            except Exception:
                logger.exception("ui_costs finance_proc failed")

    rows_html = ""
    if last_real:
        for rr in last_real:
            rows_html += (
                "<tr>"
                f"<td class='mono small'>{esc(str(rr.get('ts',''))[:19]).replace('T',' ')}</td>"
                f"<td class='mono small'>{esc(str(rr.get('imported_at','') or '')[:19]).replace('T',' ')}</td>"
                f"<td class='mono small'>{esc(rr.get('source','') or '')}</td>"
                f"<td class='mono small'>{esc(rr.get('method',''))}</td>"
                f"<td class='mono small wrap'>{esc(rr.get('endpoint',''))}</td>"
                f"<td class='mono small'>{esc(rr.get('doc','') or '')}</td>"
                f"<td class='mono small'>{esc(rr.get('cnj','') or '')}</td>"
                f"<td class='text-end mono small'>R$ {float(rr.get('cost_brl') or 0.0):.2f}</td>"
                "</tr>"
            )
    else:
        rows_html = "<tr><td colspan='8' class='muted'>Nenhum extrato importado ainda.</td></tr>"

    fin_rows_html = ""
    if finance_proc:
        finance_proc = sorted(finance_proc, key=lambda x: ((x.get("data") or ""), float(x.get("valor") or 0)), reverse=True)[:250]
        for rr in finance_proc:
            valor_fmt = f"{float(rr.get('valor') or 0.0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            fin_rows_html += (
                "<tr>"
                f"<td class='mono small'><a href='/ui/processo/{quote(str(rr.get('cnj','')))}'>{esc(rr.get('cnj','') or '')}</a></td>"
                f"<td>{esc(rr.get('tipo','') or '')}</td>"
                f"<td class='text-end mono small'>R$ {valor_fmt}</td>"
                f"<td><span class='badge text-bg-light border'>{esc(rr.get('origem','') or '')}</span></td>"
                f"<td class='mono small'>{esc(rr.get('data','') or '')}</td>"
                f"<td>{esc(rr.get('descricao','') or '')}</td>"
                f"<td class='text-end'><a class='btn btn-outline-primary btn-sm' href='/ui/processo/{quote(str(rr.get('cnj','')))}'>Abrir</a></td>"
                "</tr>"
            )
    else:
        fin_rows_html = "<tr><td colspan='7' class='muted'>Nenhum valor local identificado ainda.</td></tr>"

    body = f"""
    <div class="d-flex align-items-center justify-content-between mb-3">
      <div>
        <div class="h4 mb-0">Financeiro</div>
        <div class="muted small">Visão financeira do uso do Escavador. O custo <b>Real</b> vem do XLSX importado e a análise por processo usa dados locais já extraídos.</div>
      </div>
      <div class="d-flex gap-2">
        <a class="btn btn-outline-secondary" href="/ui">Dashboard</a>
        <button class="btn btn-outline-danger" id="btnClearReal" type="button">Limpar histórico importado</button>
      </div>
    </div>

    <div id="importMsg" class="mb-3"></div>

    <div class="card p-3 mb-3">
      <div class="h6 mb-2">Importar XLSX</div>
      <form id="xlsxForm">
        <div class="row g-2 align-items-end">
          <div class="col-md-6">
            <label class="form-label small muted">Arquivo XLSX (Relatório de Consumo da API)</label>
            <input class="form-control" type="file" name="file" accept=".xlsx" required />
          </div>
          <div class="col-md-3">
            <button class="btn btn-primary w-100" type="submit">Importar</button>
          </div>
          <div class="col-md-3">
            <button class="btn btn-outline-secondary w-100" type="button" id="btnReload">Recarregar</button>
          </div>
        </div>
        <div class="muted small mt-2">Dica: importar o mesmo arquivo novamente não duplica (dedupe automático).</div>
      </form>
    </div>

    <div class="card p-3 mb-3">
      <div class="d-flex align-items-center justify-content-between mb-2">
        <div class="h6 mb-0">Financeiro por processo</div>
        <div class="muted small">Valores locais identificados em capa e movimentações.</div>
      </div>
      <div class="table-responsive">
        <table class="table table-sm align-middle">
          <thead>
            <tr class="muted small">
              <th>CNJ</th>
              <th>Tipo</th>
              <th class="text-end">Valor</th>
              <th>Origem</th>
              <th>Data</th>
              <th>Descrição</th>
              <th class="text-end">Ação</th>
            </tr>
          </thead>
          <tbody>
            {fin_rows_html}
          </tbody>
        </table>
      </div>
    </div>

    <div class="card p-3">
      <div class="h6 mb-2">Últimos lançamentos importados (Real)</div>
      <div class="table-responsive">
        <table class="table table-sm align-middle">
          <thead>
            <tr class="muted small">
              <th>Utilização</th>
              <th>Importado</th>
              <th>Fonte</th>
              <th>Método</th>
              <th>Endpoint</th>
              <th>Doc</th>
              <th>CNJ</th>
              <th class="text-end">Custo</th>
            </tr>
          </thead>
          <tbody>
            {rows_html}
          </tbody>
        </table>
      </div>
    </div>

    <script>
    function showMsg(kind, html){{
      const el = document.getElementById('importMsg');
      const cls = kind==='ok' ? 'alert alert-success' : 'alert alert-danger';
      el.innerHTML = `<div class="${{cls}}">${{html}}</div>`;
    }}

    document.getElementById('btnReload').addEventListener('click', ()=>location.reload());

    document.getElementById('xlsxForm').addEventListener('submit', async (e)=>{{
      e.preventDefault();
      const fd = new FormData(e.target);
      showMsg('ok', 'Importando...');

      try {{
        const r = await fetch('/admin/costs/import-xlsx', {{ method:'POST', body: fd }});
        const j = await r.json();
        if(!j.ok) {{
          showMsg('err', 'Falha ao importar: ' + (j.error || 'erro desconhecido'));
          return;
        }}
        showMsg('ok', `Importação concluída. Inseridos: <b>${{j.inserted}}</b> | Duplicados: <b>${{j.skipped}}</b> | Erros: <b>${{j.errors}}</b>`);
        setTimeout(()=>location.reload(), 600);
      }} catch(ex) {{
        showMsg('err', 'Falha de rede ao importar: ' + ex);
      }}
    }});

    document.getElementById('btnClearReal').addEventListener('click', async ()=>{{
      if(!confirm('Tem certeza que deseja limpar TODO o histórico importado (custos reais)?')) return;
      try {{
        const r = await fetch('/admin/costs/clear-real', {{method:'POST'}});
        const j = await r.json();
        if(!j.ok) {{
          showMsg('err', 'Falha ao limpar histórico: ' + (j.error || 'erro'));
          return;
        }}
        showMsg('ok', 'Histórico importado limpo.');
        setTimeout(()=>location.reload(), 400);
      }} catch(ex) {{
        showMsg('err', 'Falha de rede ao limpar: ' + ex);
      }}
    }});
    </script>
    """
    return render_template_string(UI_BASE, body=body)


if __name__ == "__main__":
    start_background_tasks()
    logger.info("Starting Flask on %s:%s (token=%s)", HOST, PORT, _mask(ESCAVADOR_TOKEN))
    app.run(host=HOST, port=PORT)